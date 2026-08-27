import os
import json
import urllib.parse
from datetime import datetime
from io import BytesIO
from functools import wraps

import requests as http_requests
from itsdangerous import URLSafeTimedSerializer
from flask import (Flask, render_template, request, jsonify,
                   send_file, redirect, url_for, flash, session, abort)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from dotenv import load_dotenv

load_dotenv()

from models import (db, User, AuditLog, Todo, Guest, RoomContract, EmailLog,
                     PartiviaQuote, PartiviaRoomRate,
                     PartiviaMeetingRoom, PartiviaFBOption,
                     BudgetOverride, PnrGroup,
                     TourHotel, TourRoomCategory,
                     TourGuest, TourRoomAssignment,
                     TourHotelToken, TourHotelAccessLog,
                     TourClientToken, TourGuestDocument)


def _parse_bool(val):
    """Converte vari formati in booleano."""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ('1', 'true', 'sì', 'si', 'yes', 'x', 'v', '✓')


def create_app():
    app = Flask(__name__)
    app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev')
    app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
        'DATABASE_URL', 'postgresql://postgres:123456@localhost:5432/saba_rooming'
    ).replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
    app.config['SESSION_COOKIE_SECURE'] = bool(os.environ.get('RAILWAY_ENVIRONMENT'))

    # ── Microsoft Entra ID (SSO) ───────────────────────────────────────────
    app.config['MS_CLIENT_ID'] = os.environ.get('MS_CLIENT_ID', '')
    app.config['MS_CLIENT_SECRET'] = os.environ.get('MS_CLIENT_SECRET', '')
    app.config['MS_TENANT_ID'] = os.environ.get('MS_TENANT_ID', '')
    app.config['MS_AUTHORITY'] = f"https://login.microsoftonline.com/{app.config['MS_TENANT_ID']}"
    app.config['MS_REDIRECT_URI'] = os.environ.get(
        'MS_REDIRECT_URI',
        'http://localhost:5005/auth/callback' if not os.environ.get('RAILWAY_ENVIRONMENT')
        else 'https://web-production-394e6.up.railway.app/auth/callback'
    )
    app.config['MS_SCOPES'] = ['User.Read', 'offline_access']

    db.init_app(app)

    login_manager = LoginManager()
    login_manager.init_app(app)
    login_manager.login_view = 'login'

    @login_manager.user_loader
    def load_user(user_id):
        return db.session.get(User, int(user_id))

    import re as _re
    def _parse_eur(s):
        if not s:
            return 999999
        cleaned = _re.sub(r'[^\d.,]', '', s).replace('.', '', s.count('.') - 1).replace(',', '.')
        try:
            return float(cleaned)
        except ValueError:
            return 999999

    app.jinja_env.filters['sort_prices'] = lambda lst: sorted(lst, key=_parse_eur)

    with app.app_context():
        db.create_all()

        # Auto-migrate missing columns
        with db.engine.connect() as conn:
            from sqlalchemy import text, inspect
            email_cols = [c['name'] for c in inspect(db.engine).get_columns('email_logs')]
            if 'log_type' not in email_cols:
                conn.execute(text("ALTER TABLE email_logs ADD COLUMN log_type VARCHAR(20) DEFAULT 'rooming'"))
                conn.commit()
            quote_cols = [c['name'] for c in inspect(db.engine).get_columns('partivia_quotes')]
            if 'vat_included' not in quote_cols:
                conn.execute(text("ALTER TABLE partivia_quotes ADD COLUMN vat_included VARCHAR(20)"))
                conn.commit()
            if 'address' not in quote_cols:
                conn.execute(text("ALTER TABLE partivia_quotes ADD COLUMN address TEXT"))
                conn.commit()
            if 'image_url' not in quote_cols:
                conn.execute(text("ALTER TABLE partivia_quotes ADD COLUMN image_url TEXT"))
                conn.commit()
            if 'website_url' not in quote_cols:
                conn.execute(text("ALTER TABLE partivia_quotes ADD COLUMN website_url TEXT"))
                conn.commit()
            if 'hidden' not in quote_cols:
                conn.execute(text("ALTER TABLE partivia_quotes ADD COLUMN hidden BOOLEAN DEFAULT FALSE"))
                conn.commit()
            guest_cols = [c['name'] for c in inspect(db.engine).get_columns('guests')]
            if 'pnr_group_id' not in guest_cols:
                conn.execute(text("ALTER TABLE guests ADD COLUMN pnr_group_id INTEGER REFERENCES pnr_groups(id)"))
                conn.commit()
            if 'data_nascita' not in guest_cols:
                conn.execute(text("ALTER TABLE guests ADD COLUMN data_nascita VARCHAR(20)"))
                conn.commit()
            # Tour room categories: add sort_order if missing
            trc_tables = [t['name'] for t in inspect(db.engine).get_table_names()] if False else []
            try:
                trc_cols = [c['name'] for c in inspect(db.engine).get_columns('tour_room_categories')]
                if 'sort_order' not in trc_cols:
                    conn.execute(text("ALTER TABLE tour_room_categories ADD COLUMN sort_order INTEGER DEFAULT 0"))
                    conn.commit()
            except Exception:
                pass  # table may not exist yet
            try:
                tg_cols = [c['name'] for c in inspect(db.engine).get_columns('tour_guests')]
                for col in ('passport_file', 'driving_file'):
                    if col not in tg_cols:
                        conn.execute(text(f"ALTER TABLE tour_guests ADD COLUMN {col} VARCHAR(300)"))
                        conn.commit()
            except Exception:
                pass

            # Microsoft SSO columns on users
            try:
                u_cols = [c['name'] for c in inspect(db.engine).get_columns('users')]
                for col, col_type in (('microsoft_id', 'VARCHAR(100)'),
                                       ('ms_access_token', 'TEXT'),
                                       ('ms_refresh_token', 'TEXT')):
                    if col not in u_cols:
                        conn.execute(text(f"ALTER TABLE users ADD COLUMN {col} {col_type}"))
                        conn.commit()
            except Exception:
                pass

            # Soft-delete columns
            for _tbl in ('guests', 'partivia_quotes', 'tour_guests'):
                try:
                    _cols = [c['name'] for c in inspect(db.engine).get_columns(_tbl)]
                    if 'deleted' not in _cols:
                        conn.execute(text(f"ALTER TABLE {_tbl} ADD COLUMN deleted BOOLEAN DEFAULT FALSE"))
                        conn.commit()
                    if 'deleted_at' not in _cols:
                        conn.execute(text(f"ALTER TABLE {_tbl} ADD COLUMN deleted_at TIMESTAMP"))
                        conn.commit()
                except Exception:
                    pass

        # Migrate Italian statuses to English (one-time)
        _status_map = {
            'da_valutare': 'pending_review',
            'in_trattativa': 'negotiating',
            'confermato': 'confirmed',
            'rifiutato': 'declined',
            'scaduto': 'expired',
        }
        _migrated = 0
        for _q in PartiviaQuote.query.filter(
                PartiviaQuote.quote_status.in_(_status_map.keys())).all():
            _q.quote_status = _status_map[_q.quote_status]
            _migrated += 1
        if _migrated:
            db.session.commit()

        # Seed contratti camere se non esistono
        if RoomContract.query.count() == 0:
            CONTRATTI = [
                ('DUS Standard',          36, 218.50, 230.00),
                ('DUS Superior',          50, 266.00, 280.00),
                ('DUS Superior Sea View', 10, 289.75, 305.00),
                ('DUS Deluxe',            30, 308.75, 325.00),
                ('DUS Deluxe Sea View',   31, 332.50, 350.00),
            ]
            for notte in (8, 9):
                for tipo, disp, netta, lorda in CONTRATTI:
                    db.session.add(RoomContract(
                        tipo=tipo, disponibili=disp,
                        tariffa_netta=netta, tariffa_lorda=lorda, notte=notte))
            db.session.commit()

        # Seed default superuser
        if User.query.count() == 0:
            _admin = User(username='stefano', email='stefano.cocchi@sabae20.it',
                          is_superuser=True, role='superuser',
                          must_change_password=True)
            _admin.set_password('changeme')
            _admin.must_change_password = True
            db.session.add(_admin)
            db.session.commit()

    # ── AUTH HELPERS ───────────────────────────────────────────────────────

    def superuser_required(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.is_authenticated or not current_user.is_superuser:
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify(ok=False, error='Accesso riservato'), 403
                flash('Accesso riservato al superuser.', 'error')
                return redirect(url_for('landing'))
            return f(*args, **kwargs)
        return decorated

    @app.before_request
    def require_login():
        open_prefixes = ('/login', '/set-password', '/static',
                         '/auth/microsoft', '/auth/callback',
                         '/tour/client/', '/tour/docs/')
        if any(request.path.startswith(p) for p in open_prefixes):
            return None
        if request.path == '/favicon.ico':
            return None
        if not current_user.is_authenticated:
            if request.is_json or request.path.startswith('/api/'):
                return jsonify(ok=False, error='Non autenticato'), 401
            return redirect(url_for('login'))
        if current_user.must_change_password and request.path != '/set-password':
            return redirect(url_for('set_password'))

    # ── AUTH ROUTES ────────────────────────────────────────────────────────

    def _get_signer():
        return URLSafeTimedSerializer(app.config['SECRET_KEY'])

    @app.route('/auth/microsoft')
    def microsoft_login():
        """Avvia il flusso OAuth2 con Microsoft Entra ID."""
        if not app.config.get('MS_CLIENT_ID'):
            flash('SSO Microsoft non configurato.', 'error')
            return redirect(url_for('login', mode='local'))
        next_url = request.args.get('next', '')
        state = _get_signer().dumps({'next': next_url})
        params = {
            'client_id': app.config['MS_CLIENT_ID'],
            'response_type': 'code',
            'redirect_uri': app.config['MS_REDIRECT_URI'],
            'scope': 'openid profile email ' + ' '.join(app.config['MS_SCOPES']),
            'state': state,
            'response_mode': 'query',
        }
        auth_url = f"{app.config['MS_AUTHORITY']}/oauth2/v2.0/authorize?{urllib.parse.urlencode(params)}"
        return redirect(auth_url)

    @app.route('/auth/callback')
    def microsoft_callback():
        """Callback OAuth2 da Microsoft."""
        state = request.args.get('state', '')
        try:
            state_data = _get_signer().loads(state, max_age=600)
            next_url = state_data.get('next', '')
        except Exception:
            flash('Sessione di autenticazione scaduta, riprova.', 'error')
            return redirect(url_for('login', mode='local'))

        code = request.args.get('code')
        if not code:
            error = request.args.get('error_description',
                                     request.args.get('error', 'Errore sconosciuto'))
            flash(f'Errore Microsoft: {error}', 'error')
            return redirect(url_for('login', mode='local'))

        # Scambia code per token
        token_url = f"{app.config['MS_AUTHORITY']}/oauth2/v2.0/token"
        token_data = {
            'client_id': app.config['MS_CLIENT_ID'],
            'client_secret': app.config['MS_CLIENT_SECRET'],
            'code': code,
            'redirect_uri': app.config['MS_REDIRECT_URI'],
            'grant_type': 'authorization_code',
            'scope': 'openid profile email ' + ' '.join(app.config['MS_SCOPES']),
        }
        r = http_requests.post(token_url, data=token_data, timeout=15)
        if r.status_code != 200:
            app.logger.error('SSO token exchange failed: status=%s', r.status_code)
            flash('Errore nello scambio del token con Microsoft.', 'error')
            return redirect(url_for('login', mode='local'))

        tokens = r.json()
        access_token = tokens.get('access_token')

        # Profilo utente da Graph API
        headers = {'Authorization': f'Bearer {access_token}'}
        me = http_requests.get('https://graph.microsoft.com/v1.0/me',
                               headers=headers, timeout=10).json()
        ms_id = me.get('id')
        ms_email = (me.get('mail') or me.get('userPrincipalName') or '').lower()

        if not ms_id or not ms_email:
            flash('Impossibile recuperare dati utente da Microsoft.', 'error')
            return redirect(url_for('login', mode='local'))

        # Cerca utente per microsoft_id oppure email
        user = User.query.filter_by(microsoft_id=ms_id, is_active=True).first()
        if not user:
            user = User.query.filter_by(email=ms_email, is_active=True).first()
            if user:
                user.microsoft_id = ms_id
                db.session.commit()

        if not user:
            app.logger.warning('SSO: nessun utente trovato per ms_id=%s email=%s', ms_id, ms_email)
            flash(f'Account Microsoft ({ms_email}) non autorizzato. '
                  'Chiedi a un amministratore di creare il tuo utente.', 'error')
            return redirect(url_for('login', mode='local'))

        # Salva token
        user.ms_access_token = access_token
        if tokens.get('refresh_token'):
            user.ms_refresh_token = tokens['refresh_token']
        db.session.commit()

        login_user(user)
        log_audit('system', 'User', user.id, 'login',
                  summary=f'Login SSO Microsoft: {ms_email}')

        if user.must_change_password:
            # SSO non richiede cambio password locale
            user.must_change_password = False
            db.session.commit()

        return redirect(next_url or url_for('landing'))

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for('landing'))

        sso_enabled = bool(app.config.get('MS_CLIENT_ID'))

        # GET senza mode=local e SSO attivo → redirect automatico a Microsoft
        if request.method == 'GET' and request.args.get('mode') != 'local' and sso_enabled:
            return redirect(url_for('microsoft_login',
                                    next=request.args.get('next', '')))

        if request.method == 'POST':
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')
            user = User.query.filter_by(username=username).first()
            if user and not user.is_active:
                flash('Account disattivato. Contatta un amministratore.', 'error')
                return render_template('login.html', sso_enabled=sso_enabled)
            if user and not user.password_hash:
                flash('Password resettata. Contatta un amministratore.', 'error')
                return render_template('login.html', sso_enabled=sso_enabled)
            if user and user.check_password(password):
                user.reset_failed_logins()
                db.session.commit()
                login_user(user)
                log_audit('system', 'User', user.id, 'login',
                          summary=f'Login: {user.username}')
                if user.must_change_password:
                    return redirect(url_for('set_password'))
                return redirect(url_for('landing'))
            if user:
                user.register_failed_login()
                db.session.commit()
                if not user.password_hash:
                    flash('Troppi tentativi falliti. Password resettata — contatta un amministratore.', 'error')
                else:
                    flash('Password errata.', 'error')
            else:
                flash('Utente non trovato.', 'error')
            return render_template('login.html', sso_enabled=sso_enabled)
        return render_template('login.html', sso_enabled=sso_enabled)

    @app.route('/set-password', methods=['GET', 'POST'])
    def set_password():
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if request.method == 'POST':
            pw = request.form.get('password', '')
            pw2 = request.form.get('password2', '')
            if len(pw) < 6:
                flash('La password deve avere almeno 6 caratteri.', 'error')
                return render_template('set_password.html')
            if pw != pw2:
                flash('Le password non corrispondono.', 'error')
                return render_template('set_password.html')
            current_user.set_password(pw)
            current_user.must_change_password = False
            db.session.commit()
            flash('Password impostata con successo.', 'success')
            return redirect(url_for('landing'))
        return render_template('set_password.html')

    @app.route('/logout')
    def logout():
        if current_user.is_authenticated:
            log_audit('system', 'User', current_user.id, 'logout',
                      summary=f'Logout: {current_user.username}')
        logout_user()
        return redirect(url_for('login', mode='local'))

    # ── ADMIN: USER MANAGEMENT ─────────────────────────────────────────────

    @app.route('/admin/users')
    @superuser_required
    def admin_users():
        users = User.query.order_by(User.created_at.desc()).all()
        return render_template('admin_users.html', users=users)

    @app.route('/admin/users', methods=['POST'])
    @superuser_required
    def admin_create_user():
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        role = request.form.get('role', 'user')
        if not username or not email:
            flash('Username e email sono obbligatori.', 'error')
            return redirect(url_for('admin_users'))
        if User.query.filter((User.username == username) | (User.email == email)).first():
            flash('Username o email già esistente.', 'error')
            return redirect(url_for('admin_users'))
        u = User(username=username, email=email, role=role,
                 is_superuser=(role == 'superuser'),
                 must_change_password=True)
        u.set_password('changeme')
        u.must_change_password = True
        db.session.add(u)
        db.session.commit()
        log_audit('system', 'User', u.id, 'create',
                  summary=f'Creato utente {u.username} ({u.role})')
        flash(f'Utente {username} creato. Password temporanea: changeme', 'success')
        return redirect(url_for('admin_users'))

    @app.post('/admin/users/<int:uid>/toggle-role')
    @superuser_required
    def admin_toggle_role(uid):
        u = User.query.get_or_404(uid)
        if u.id == current_user.id:
            flash('Non puoi cambiare il tuo stesso ruolo.', 'error')
            return redirect(url_for('admin_users'))
        cycle = {'user': 'superuser', 'superuser': 'client', 'client': 'user'}
        old_role = u.role or 'user'
        u.role = cycle.get(old_role, 'user')
        u.is_superuser = (u.role == 'superuser')
        db.session.commit()
        log_audit('system', 'User', u.id, 'update',
                  changes={'role': {'old': old_role, 'new': u.role}},
                  summary=f'Ruolo {u.username}: {old_role} → {u.role}')
        flash(f'Ruolo di {u.username} cambiato a {u.role}.', 'success')
        return redirect(url_for('admin_users'))

    @app.post('/admin/users/<int:uid>/reset-password')
    @superuser_required
    def admin_reset_password(uid):
        u = User.query.get_or_404(uid)
        u.set_password('changeme')
        u.must_change_password = True
        u.failed_login_attempts = 0
        db.session.commit()
        log_audit('system', 'User', u.id, 'update',
                  summary=f'Password resettata per {u.username}')
        flash(f'Password di {u.username} resettata a "changeme".', 'success')
        return redirect(url_for('admin_users'))

    @app.post('/admin/users/<int:uid>/toggle-active')
    @superuser_required
    def admin_toggle_active(uid):
        u = User.query.get_or_404(uid)
        if u.id == current_user.id:
            flash('Non puoi disattivare te stesso.', 'error')
            return redirect(url_for('admin_users'))
        u.is_active = not u.is_active
        db.session.commit()
        status = 'attivato' if u.is_active else 'disattivato'
        log_audit('system', 'User', u.id, 'update',
                  changes={'is_active': {'old': not u.is_active, 'new': u.is_active}},
                  summary=f'Utente {u.username} {status}')
        flash(f'Utente {u.username} {status}.', 'success')
        return redirect(url_for('admin_users'))

    # ── AUDIT HELPERS ──────────────────────────────────────────────────────

    def log_audit(section, entity_type, entity_id, action, changes=None, summary=None):
        try:
            ip = request.headers.get('X-Forwarded-For', request.remote_addr or '')
            entry = AuditLog(
                timestamp=datetime.utcnow(),
                user_id=current_user.id if current_user.is_authenticated else None,
                user_email=(current_user.email if current_user.is_authenticated
                            else 'system'),
                section=section,
                entity_type=entity_type,
                entity_id=entity_id,
                action=action,
                changes=changes,
                summary=summary or f'{action} {entity_type} #{entity_id}',
                ip_address=str(ip)[:45]
            )
            db.session.add(entry)
            db.session.commit()
        except Exception:
            db.session.rollback()

    def _snapshot(obj, fields):
        """Dict of current field values for audit logging."""
        result = {}
        for f in fields:
            v = getattr(obj, f, None)
            if isinstance(v, datetime):
                v = v.isoformat()
            result[f] = v
        return result

    def _diff(obj, data, str_fields, bool_fields=()):
        """Return {field: {old, new}} for fields that actually changed."""
        changes = {}
        for f in str_fields:
            if f in data:
                old = getattr(obj, f, None) or ''
                new = data[f] if data[f] is not None else ''
                if str(old) != str(new):
                    changes[f] = {'old': old, 'new': new}
        for f in bool_fields:
            if f in data:
                old = getattr(obj, f, None) or False
                new = _parse_bool(data[f])
                if old != new:
                    changes[f] = {'old': old, 'new': new}
        return changes

    # ── TO-DO LIST API ─────────────────────────────────────────────────────

    @app.get('/api/todos/<section>')
    def list_todos(section):
        if section not in ('rooming', 'partivia', 'tour'):
            return jsonify(ok=False, error='Sezione non valida'), 400
        status_filter = request.args.get('status')
        q = Todo.query.filter_by(section=section).order_by(
            Todo.status.asc(), Todo.priority.desc(), Todo.due_date.asc().nullslast(), Todo.created_at.desc())
        if status_filter:
            q = q.filter_by(status=status_filter)
        todos = q.all()
        return jsonify(ok=True, todos=[{
            'id': t.id, 'title': t.title, 'description': t.description or '',
            'owner': t.owner or '', 'priority': t.priority or 'normal',
            'status': t.status or 'todo',
            'due_date': t.due_date.isoformat() if t.due_date else None,
            'created_by': t.created_by or '',
            'created_at': t.created_at.isoformat() if t.created_at else None,
            'completed_at': t.completed_at.isoformat() if t.completed_at else None,
        } for t in todos])

    @app.post('/api/todos/<section>')
    def create_todo(section):
        if section not in ('rooming', 'partivia', 'tour'):
            return jsonify(ok=False, error='Sezione non valida'), 400
        data = request.get_json() or {}
        title = (data.get('title') or '').strip()
        if not title:
            return jsonify(ok=False, error='Titolo obbligatorio'), 400
        due = None
        if data.get('due_date'):
            try:
                due = datetime.strptime(data['due_date'], '%Y-%m-%d').date()
            except ValueError:
                pass
        t = Todo(
            section=section, title=title,
            description=(data.get('description') or '').strip(),
            owner=(data.get('owner') or '').strip(),
            priority=data.get('priority', 'normal'),
            status='todo', due_date=due,
            created_by=current_user.username if current_user.is_authenticated else '')
        db.session.add(t)
        db.session.commit()
        log_audit(section, 'Todo', t.id, 'create',
                  summary=f'To-Do: {t.title}')
        return jsonify(ok=True, id=t.id)

    @app.put('/api/todos/<int:todo_id>')
    def update_todo(todo_id):
        t = Todo.query.get_or_404(todo_id)
        data = request.get_json() or {}
        changes = {}
        for f in ('title', 'description', 'owner', 'priority', 'status'):
            if f in data and data[f] is not None:
                old = getattr(t, f)
                new = data[f].strip() if isinstance(data[f], str) else data[f]
                if old != new:
                    changes[f] = {'old': old, 'new': new}
                    setattr(t, f, new)
        if 'due_date' in data:
            old_due = t.due_date.isoformat() if t.due_date else None
            if data['due_date']:
                try:
                    t.due_date = datetime.strptime(data['due_date'], '%Y-%m-%d').date()
                except ValueError:
                    pass
            else:
                t.due_date = None
            new_due = t.due_date.isoformat() if t.due_date else None
            if old_due != new_due:
                changes['due_date'] = {'old': old_due, 'new': new_due}
        if t.status == 'done' and not t.completed_at:
            t.completed_at = datetime.utcnow()
        elif t.status != 'done':
            t.completed_at = None
        db.session.commit()
        if changes:
            log_audit(t.section, 'Todo', t.id, 'update', changes=changes,
                      summary=f'To-Do aggiornato: {t.title}')
        return jsonify(ok=True)

    @app.delete('/api/todos/<int:todo_id>')
    def delete_todo(todo_id):
        t = Todo.query.get_or_404(todo_id)
        title = t.title
        section = t.section
        db.session.delete(t)
        db.session.commit()
        log_audit(section, 'Todo', todo_id, 'delete',
                  summary=f'To-Do eliminato: {title}')
        return jsonify(ok=True)

    @app.post('/api/todos/<int:todo_id>/toggle')
    def toggle_todo(todo_id):
        t = Todo.query.get_or_404(todo_id)
        old_status = t.status
        if t.status == 'done':
            t.status = 'todo'
            t.completed_at = None
        else:
            t.status = 'done'
            t.completed_at = datetime.utcnow()
        db.session.commit()
        log_audit(t.section, 'Todo', t.id, 'update',
                  changes={'status': {'old': old_status, 'new': t.status}},
                  summary=f'To-Do {"completato" if t.status == "done" else "riaperto"}: {t.title}')
        return jsonify(ok=True, status=t.status)

    # ── LANDING PAGE ────────────────────────────────────────────────────────

    @app.route('/')
    def landing():
        return render_template('landing.html')

    # ── LLM GENERICO ───────────────────────────────────────────────────────

    TOUR_GUEST_STR_FIELDS = (
        'cognome', 'nome', 'email', 'telefono', 'nazionalita', 'titolo',
        'arrivo_mezzo', 'arrivo_data', 'room_with', 'car_number', 'car_with',
        'vip', 'client_room_note', 'payment', 'cloth_size', 'diet',
        'notes', 'email_requests',
    )
    TOUR_GUEST_BOOL_FIELDS = ('dinner', 'sept2')

    @app.post('/api/llm/process')
    def llm_process():
        import anthropic
        payload = request.get_json() or {}
        text = payload.get('text', '').strip()
        section = payload.get('section', '')
        if not text:
            return jsonify(ok=False, error='Nessun testo inserito'), 400
        if section not in ('rooming', 'partivia', 'tour'):
            return jsonify(ok=False, error='Sezione non valida'), 400

        api_key = os.environ.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify(ok=False, error='ANTHROPIC_API_KEY non configurata'), 500

        # Build context only for the active section
        if section == 'rooming':
            guests = Guest.query.filter_by(deleted=False).order_by(Guest.cognome).all()
            entity_list = '\n'.join(
                f'  ID={g.id} {g.cognome} {g.nome} (email={g.email or ""}, sede={g.sede_lavoro or ""})'
                for g in guests) or '  (nessun ospite)'
            section_desc = """SEZIONE: ROOMING (Equans, 8-9 Ottobre 2026) — Gestione ospiti evento aziendale
Entità: Guest
Campi stringa: cognome (MAIUSCOLO), nome, email, telefono, sede_lavoro, volo_arrivo, volo_partenza, aeroporto_partenza, aeroporto_arrivo, pickup_bus_andata, pickup_bus_ritorno, divide_stanza_con, restrizioni_alimentari, tipo_camera, camera_assegnata, note_form, note, data_nascita
Campi booleani: presenza_8, presenza_9, presenza_10, presenza_11, parcheggio_linate, parcheggio_hotel"""

        elif section == 'partivia':
            quotes = PartiviaQuote.query.filter_by(deleted=False).order_by(PartiviaQuote.hotel_name).all()
            entity_list = '\n'.join(
                f'  ID={q.id} {q.hotel_name} ({q.city}, {q.stars or "?"}★, status={q.quote_status})'
                for q in quotes) or '  (nessun preventivo)'
            section_desc = """SEZIONE: PARTIVIA (N!Partivia Spain) — Preventivi hotel per viaggio incentive in Spagna
Entità: PartiviaQuote
Campi: hotel_name, city (Barcellona/Madrid/Siviglia/Valencia), stars (1-5), contact_name, contact_email, website_url, address, dates_proposed, rooms_available, min_rooms_required, cancellation_policy, payment_terms, validity_date, commission, total_estimate, included_services, notes, raw_summary, quote_status (pending_review/negotiating/confirmed/declined/expired), vat_included (yes/no/unknown)
Sub-tabelle: room_rates (room_type, rate_per_night con €, breakfast_included, notes), meeting_rooms (name, capacity, rate, notes), fb_options (meal_type, price_per_person, menu_description)
IMPORTANTE: Tutti i testi DEVONO essere in inglese."""

        else:  # tour
            tour_guests = TourGuest.query.filter_by(deleted=False).order_by(TourGuest.cognome).all()
            # Build guest list with hotel assignments
            hotels = TourHotel.query.order_by(TourHotel.night_date).all()
            assignments = TourRoomAssignment.query.all()
            # Map guest_id → list of hotel assignments
            guest_hotels = {}
            hotel_map = {h.id: h for h in hotels}
            for a in assignments:
                h = hotel_map.get(a.hotel_id)
                if h:
                    guest_hotels.setdefault(a.guest_id, []).append(
                        f'{h.hotel_name} ({h.night_label}, room={a.room_code or "?"})')
            entity_list = '\n'.join(
                f'  ID={g.id} {g.cognome} {g.nome} (payment={g.payment or ""}, car={g.car_number or ""}, '
                f'hotels=[{", ".join(guest_hotels.get(g.id, ["non assegnato"]))}])'
                for g in tour_guests) or '  (nessun partecipante)'
            # Hotel list
            hotel_list = '\n'.join(
                f'  {h.hotel_name} ({h.city}, notte {h.night_label}, {h.rooms_blocked} camere)'
                for h in hotels) or '  (nessun hotel)'
            section_desc = f"""SEZIONE: TOUR (Liqui Moly Nexus Auto Tour, 1-5 Settembre 2026) — Tour itinerante multi-città
Entità: TourGuest
Campi stringa: cognome (MAIUSCOLO), nome, email, telefono, nazionalita, titolo (Mr/Mrs), arrivo_mezzo (Airplane/Car/Train/Other), arrivo_data, room_with, car_number, car_with, vip (VIP/ULTRA VIP), client_room_note, payment (PAID/TO COLLECT/NO NEED/PAY ON SITE), cloth_size (S-XXXL), diet, notes, email_requests
Campi booleani: dinner, sept2

Hotel del tour:
{hotel_list}

Ogni partecipante ha le assegnazioni hotel indicate tra parentesi quadre."""

        system_prompt = f"""Sei un assistente per la gestione eventi Saba. Analizzi richieste in linguaggio naturale e determini le operazioni CRUD da eseguire.
Lavori ESCLUSIVAMENTE sulla sezione indicata. Non mischiare dati di altre sezioni.

{section_desc}

Record attuali:
{entity_list}

REGOLE:
- Per update/delete: usa match_id con l'ID esatto dell'entità dalla lista sopra. Fai matching fuzzy su cognome+nome.
- Per create: match_id = null
- cognome sempre MAIUSCOLO
- Campi non menzionati nel testo → null (non includerli in data)
- Se la richiesta è solo una query/domanda (non un'operazione CRUD), rispondi con operations vuoto e metti la risposta nel summary.

Rispondi SOLO con JSON valido (no markdown, no commenti):
{{
  "section": "{section}",
  "operations": [
    {{
      "action": "create|update|delete",
      "entity_type": "Guest|PartiviaQuote|TourGuest",
      "data": {{}},
      "match_id": null,
      "preview": "Descrizione leggibile dell'operazione"
    }}
  ],
  "summary": "Riassunto complessivo delle operazioni"
}}"""

        try:
            client = anthropic.Anthropic(api_key=api_key)
            response = client.messages.create(
                model='claude-haiku-4-5-20251001',
                max_tokens=8192,
                system=system_prompt,
                messages=[{'role': 'user', 'content': text}],
            )
            raw = response.content[0].text.strip()
            if raw.startswith('```'):
                raw = raw.split('\n', 1)[1].rsplit('```', 1)[0].strip()
            result = json.loads(raw)
        except json.JSONDecodeError:
            return jsonify(ok=False, error='Risposta LLM non valida', raw=raw), 500
        except Exception as e:
            return jsonify(ok=False, error=str(e)), 500

        # Enrich operations with current data for diff preview
        for op in result.get('operations', []):
            if op.get('match_id') and op.get('action') == 'update':
                etype = op.get('entity_type')
                mid = op['match_id']
                if etype == 'Guest':
                    obj = Guest.query.get(mid)
                    if obj:
                        op['current_data'] = {f: getattr(obj, f, None)
                                               for f in list(op.get('data', {}).keys())}
                elif etype == 'PartiviaQuote':
                    obj = PartiviaQuote.query.get(mid)
                    if obj:
                        op['current_data'] = {f: getattr(obj, f, None)
                                               for f in list(op.get('data', {}).keys())}
                elif etype == 'TourGuest':
                    obj = TourGuest.query.get(mid)
                    if obj:
                        op['current_data'] = {f: getattr(obj, f, None)
                                               for f in list(op.get('data', {}).keys())}

        # Save to email log
        log = EmailLog(testo=text, summary=result.get('summary', ''), log_type='llm')
        db.session.add(log)
        db.session.commit()

        inp = response.usage.input_tokens
        out = response.usage.output_tokens
        cost = (inp * 0.80 + out * 4.00) / 1_000_000

        return jsonify(ok=True, **result, email_log_id=log.id,
                       tokens={'input': inp, 'output': out, 'cost_eur': round(cost * 0.92, 4)})

    @app.post('/api/llm/apply')
    def llm_apply():
        data = request.get_json() or {}
        operations = data.get('operations', [])
        email_log_id = data.get('email_log_id')
        results = []

        for op in operations:
            action = op.get('action')
            etype = op.get('entity_type')
            opdata = op.get('data', {})
            match_id = op.get('match_id')

            try:
                if etype == 'Guest':
                    if action == 'create':
                        g = Guest(
                            cognome=opdata.get('cognome', ''),
                            nome=opdata.get('nome', ''),
                            source='llm', email_log_id=email_log_id)
                        for f in ('email', 'telefono', 'sede_lavoro', 'volo_arrivo',
                                  'volo_partenza', 'aeroporto_partenza', 'aeroporto_arrivo',
                                  'pickup_bus_andata', 'pickup_bus_ritorno',
                                  'divide_stanza_con', 'restrizioni_alimentari',
                                  'tipo_camera', 'camera_assegnata', 'note_form', 'note',
                                  'data_nascita'):
                            if opdata.get(f) is not None:
                                setattr(g, f, opdata[f])
                        for f in ('presenza_8', 'presenza_9', 'presenza_10', 'presenza_11',
                                  'parcheggio_linate', 'parcheggio_hotel'):
                            if opdata.get(f) is not None:
                                setattr(g, f, _parse_bool(opdata[f]))
                        db.session.add(g)
                        db.session.flush()
                        log_audit('rooming', 'Guest', g.id, 'create',
                                  summary=f'LLM: Aggiunto {g.nome_completo}')
                        results.append({'ok': True, 'action': 'created', 'id': g.id,
                                        'name': g.nome_completo})

                    elif action == 'update' and match_id:
                        g = Guest.query.get(match_id)
                        if not g or g.deleted:
                            results.append({'ok': False, 'error': f'Guest {match_id} non trovato'})
                            continue
                        changes = _diff(g, opdata,
                                        [f for f in opdata if f not in ('presenza_8', 'presenza_9',
                                         'presenza_10', 'presenza_11', 'parcheggio_linate', 'parcheggio_hotel')],
                                        [f for f in opdata if f in ('presenza_8', 'presenza_9',
                                         'presenza_10', 'presenza_11', 'parcheggio_linate', 'parcheggio_hotel')])
                        for f in opdata:
                            if opdata[f] is not None:
                                if f in ('presenza_8', 'presenza_9', 'presenza_10', 'presenza_11',
                                         'parcheggio_linate', 'parcheggio_hotel'):
                                    setattr(g, f, _parse_bool(opdata[f]))
                                else:
                                    setattr(g, f, opdata[f])
                        g.updated_at = datetime.utcnow()
                        if changes:
                            log_audit('rooming', 'Guest', g.id, 'update', changes=changes,
                                      summary=f'LLM: Modificato {g.nome_completo}')
                        results.append({'ok': True, 'action': 'updated', 'id': g.id,
                                        'name': g.nome_completo})

                    elif action == 'delete' and match_id:
                        g = Guest.query.get(match_id)
                        if not g or g.deleted:
                            results.append({'ok': False, 'error': f'Guest {match_id} non trovato'})
                            continue
                        g.deleted = True
                        g.deleted_at = datetime.utcnow()
                        log_audit('rooming', 'Guest', g.id, 'delete',
                                  summary=f'LLM: Eliminato {g.nome_completo}')
                        results.append({'ok': True, 'action': 'deleted', 'id': g.id,
                                        'name': g.nome_completo})

                elif etype == 'PartiviaQuote':
                    if action == 'create':
                        q = PartiviaQuote(
                            hotel_name=opdata.get('hotel_name', ''),
                            city=opdata.get('city', ''),
                            source='llm', email_log_id=email_log_id)
                        for f in ('stars', 'contact_name', 'contact_email', 'website_url',
                                  'address', 'dates_proposed', 'rooms_available',
                                  'min_rooms_required', 'cancellation_policy', 'payment_terms',
                                  'validity_date', 'commission', 'total_estimate',
                                  'included_services', 'notes', 'raw_summary',
                                  'quote_status', 'vat_included'):
                            if opdata.get(f) is not None:
                                setattr(q, f, opdata[f])
                        db.session.add(q)
                        db.session.flush()
                        for rr in opdata.get('room_rates', []):
                            db.session.add(PartiviaRoomRate(
                                quote_id=q.id, room_type=rr.get('room_type', ''),
                                rate_per_night=rr.get('rate_per_night'),
                                breakfast_included=rr.get('breakfast_included'),
                                notes=rr.get('notes')))
                        for mr in opdata.get('meeting_rooms', []):
                            db.session.add(PartiviaMeetingRoom(
                                quote_id=q.id, name=mr.get('name', ''),
                                capacity=mr.get('capacity'),
                                rate=mr.get('rate'), notes=mr.get('notes')))
                        for fb in opdata.get('fb_options', []):
                            db.session.add(PartiviaFBOption(
                                quote_id=q.id, meal_type=fb.get('meal_type', ''),
                                price_per_person=fb.get('price_per_person'),
                                menu_description=fb.get('menu_description')))
                        log_audit('partivia', 'PartiviaQuote', q.id, 'create',
                                  summary=f'LLM: Creato preventivo {q.hotel_name}')
                        results.append({'ok': True, 'action': 'created', 'id': q.id,
                                        'name': q.hotel_name})

                    elif action == 'update' and match_id:
                        q = PartiviaQuote.query.get(match_id)
                        if not q or q.deleted:
                            results.append({'ok': False, 'error': f'Quote {match_id} non trovata'})
                            continue
                        for f in opdata:
                            if f not in ('room_rates', 'meeting_rooms', 'fb_options') and opdata[f] is not None:
                                setattr(q, f, opdata[f])
                        q.updated_at = datetime.utcnow()
                        log_audit('partivia', 'PartiviaQuote', q.id, 'update',
                                  summary=f'LLM: Modificato {q.hotel_name}')
                        results.append({'ok': True, 'action': 'updated', 'id': q.id,
                                        'name': q.hotel_name})

                    elif action == 'delete' and match_id:
                        q = PartiviaQuote.query.get(match_id)
                        if not q or q.deleted:
                            results.append({'ok': False, 'error': f'Quote {match_id} non trovata'})
                            continue
                        q.deleted = True
                        q.deleted_at = datetime.utcnow()
                        log_audit('partivia', 'PartiviaQuote', q.id, 'delete',
                                  summary=f'LLM: Eliminato {q.hotel_name}')
                        results.append({'ok': True, 'action': 'deleted', 'id': q.id,
                                        'name': q.hotel_name})

                elif etype == 'TourGuest':
                    if action == 'update' and match_id:
                        g = TourGuest.query.get(match_id)
                        if not g or g.deleted:
                            results.append({'ok': False, 'error': f'TourGuest {match_id} non trovato'})
                            continue
                        str_fields = [f for f in opdata if f in TOUR_GUEST_STR_FIELDS]
                        bool_fields = [f for f in opdata if f in TOUR_GUEST_BOOL_FIELDS]
                        changes = _diff(g, opdata, str_fields, bool_fields)
                        for f in str_fields:
                            if opdata[f] is not None:
                                setattr(g, f, opdata[f])
                        for f in bool_fields:
                            if opdata[f] is not None:
                                setattr(g, f, _parse_bool(opdata[f]))
                        g.updated_at = datetime.utcnow()
                        if changes:
                            log_audit('tour', 'TourGuest', g.id, 'update', changes=changes,
                                      summary=f'LLM: Modificato {g.nome_completo}')
                        results.append({'ok': True, 'action': 'updated', 'id': g.id,
                                        'name': g.nome_completo})

                    elif action == 'delete' and match_id:
                        g = TourGuest.query.get(match_id)
                        if not g or g.deleted:
                            results.append({'ok': False, 'error': f'TourGuest {match_id} non trovato'})
                            continue
                        g.deleted = True
                        g.deleted_at = datetime.utcnow()
                        log_audit('tour', 'TourGuest', g.id, 'delete',
                                  summary=f'LLM: Eliminato {g.nome_completo}')
                        results.append({'ok': True, 'action': 'deleted', 'id': g.id,
                                        'name': g.nome_completo})

            except Exception as e:
                results.append({'ok': False, 'error': str(e)})

        db.session.commit()
        return jsonify(ok=True, results=results)

    # ── PAGINA ROOMING ──────────────────────────────────────────────────────

    @app.route('/rooming/client')
    def rooming_client():
        return index(client_view=True)

    @app.route('/rooming')
    def index(client_view=False):
        guests = Guest.query.filter_by(deleted=False).order_by(Guest.cognome, Guest.nome).all()
        return render_template('index.html', guests=guests, client_view=client_view)

    # ── CRUD API ─────────────────────────────────────────────────────────────

    # Campi stringa editabili
    GUEST_STR_FIELDS = (
        'cognome', 'nome', 'email', 'telefono', 'sede_lavoro',
        'volo_arrivo', 'volo_partenza',
        'aeroporto_partenza', 'aeroporto_arrivo',
        'pickup_bus_andata', 'pickup_bus_ritorno',
        'divide_stanza_con', 'restrizioni_alimentari',
        'tipo_camera', 'camera_assegnata', 'note_form', 'note',
        'data_nascita',
    )
    GUEST_BOOL_FIELDS = (
        'presenza_8', 'presenza_9', 'presenza_10', 'presenza_11',
        'parcheggio_linate', 'parcheggio_hotel',
    )

    @app.post('/api/guest')
    def add_guest():
        data = request.get_json()
        kwargs = {'source': 'manual'}
        for f in GUEST_STR_FIELDS:
            v = data.get(f, '').strip() if data.get(f) else None
            kwargs[f] = v
        for f in GUEST_BOOL_FIELDS:
            kwargs[f] = _parse_bool(data.get(f))
        g = Guest(**kwargs)
        if not g.cognome:
            return jsonify(ok=False, error='Cognome obbligatorio'), 400
        db.session.add(g)
        db.session.commit()
        log_audit('rooming', 'Guest', g.id, 'create',
                  changes=_snapshot(g, GUEST_STR_FIELDS),
                  summary=f'Aggiunto {g.nome_completo}')
        return jsonify(ok=True, id=g.id)

    @app.put('/api/guest/<int:gid>')
    def update_guest(gid):
        g = Guest.query.get_or_404(gid)
        data = request.get_json()
        changes = _diff(g, data, GUEST_STR_FIELDS, GUEST_BOOL_FIELDS)
        for field in GUEST_STR_FIELDS:
            if field in data:
                setattr(g, field, data[field].strip() if data[field] else None)
        for field in GUEST_BOOL_FIELDS:
            if field in data:
                setattr(g, field, _parse_bool(data[field]))
        g.updated_at = datetime.utcnow()
        db.session.commit()
        if changes:
            log_audit('rooming', 'Guest', g.id, 'update',
                      changes=changes,
                      summary=f'Modificato {g.nome_completo}')
        return jsonify(ok=True)

    @app.delete('/api/guest/<int:gid>')
    def delete_guest(gid):
        g = Guest.query.get_or_404(gid)
        g.deleted = True
        g.deleted_at = datetime.utcnow()
        db.session.commit()
        log_audit('rooming', 'Guest', g.id, 'delete',
                  changes=_snapshot(g, GUEST_STR_FIELDS),
                  summary=f'Eliminato {g.nome_completo}')
        return jsonify(ok=True)

    # ── IMPORT XLSX (LLM-guided) ─────────────────────────────────────────────

    @app.post('/api/import/preview')
    def import_preview():
        """Step 1: Upload XLSX, LLM analizza headers e prime righe, propone mapping."""
        import anthropic

        f = request.files.get('file')
        if not f or not f.filename.endswith(('.xlsx', '.xls')):
            return jsonify(ok=False, error='File XLSX richiesto'), 400

        from openpyxl import load_workbook
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify(ok=False, error='File vuoto'), 400

        # Leggi header + prime 5 righe di dati per contesto
        header = [str(c).strip() if c else '' for c in rows[0]]
        sample_rows = []
        for row in rows[1:6]:
            sample_rows.append([str(c).strip() if c else '' for c in row])

        # Prepara tutte le righe dati per salvarle in sessione
        all_rows = []
        for row in rows[1:]:
            all_rows.append([str(c).strip() if c else '' for c in row])

        # Chiedi a Claude di mappare le colonne
        api_key = os.environ.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify(ok=False, error='ANTHROPIC_API_KEY non configurata'), 500

        system_prompt = """Sei un assistente che analizza file Excel di rooming list per un evento che si svolge dall'8 all'11 ottobre.

Ti vengono dati gli header delle colonne e alcune righe di esempio.

STEP 1 — MAPPING COLONNE
Mappa ogni colonna del file a uno dei seguenti campi del database:

- cognome, nome (o "nome_completo" se in una sola colonna, con campo "formato": "nome cognome" o "cognome nome")
- email, telefono
- sede_lavoro (città/sede di lavoro, es. MILANO, CATANIA)
- presenza_8, presenza_9, presenza_10, presenza_11 (giorni 8-11 ottobre)
- volo_arrivo (volo di andata), volo_partenza (volo di ritorno)
- aeroporto_partenza, aeroporto_arrivo
- pickup_bus_andata (orario pickup bus andata), pickup_bus_ritorno (orario pickup bus ritorno)
- parcheggio_linate (booleano)
- parcheggio_hotel (booleano)
- divide_stanza_con (con chi condivide la stanza)
- restrizioni_alimentari (allergie, intolleranze, diete religiose, ecc.)
- tipo_camera (singola, doppia, twin, suite, etc.)
- note_form (note inserite dall'utente nel form di registrazione)
- note (note operative/gestionali)
- IGNORA: colonne che non servono

STEP 2 — INTERPRETAZIONE DATI (FONDAMENTALE)
I dati nel file possono NON corrispondere 1:1 ai campi. Devi capire il significato reale.

REGOLE DI INTERPRETAZIONE DATE/PRESENZE:
- L'evento è dall'8 all'11 ottobre. I giorni sono: 8, 9, 10, 11.
- "arrivo gio 8/10" o "arrivo 8 ott" = la persona ARRIVA il giorno 8 ottobre
- "partenza ven 10/10" o "riparte 10" = la persona RIPARTE il giorno 10 ottobre
- Se una persona arriva il giorno X e riparte il giorno Y, è PRESENTE tutti i giorni da X a Y-1 (l'ultimo giorno riparte, non è presente all'evento)
  - Esempio: arrivo 8, partenza 10 → presenza_8=sì, presenza_9=sì, presenza_10=no, presenza_11=no
- Se c'è solo "arrivo 8" senza partenza, assumi che resti fino alla fine (presenza_8=sì, presenza_9=sì, presenza_10=sì, presenza_11=sì)
- "8/10" in una colonna di date può significare "8 ottobre" (giorno/mese) — NON "dall'8 al 10"
- Se ci sono colonne separate per ogni giorno (es. "8 ott", "9 ott"), mappale direttamente a presenza_8, presenza_9, ecc.
- Se c'è UNA sola colonna con date di arrivo/partenza, NON mapparla a un singolo campo presenza. Segnalala come "date_soggiorno" e nella sezione "trasformazioni" spiega come derivare le presenze.

REGOLE GENERALI:
- Analizza i DATI nelle righe, non solo gli header
- Se non sei sicuro, mappa come IGNORA
- Colonne vuote o con solo formattazione → IGNORA

Rispondi SOLO con JSON valido:
{
  "mapping": {
    "0": {"campo": "cognome", "header_originale": "Surname", "confidenza": "alta"},
    "1": {"campo": "nome_completo", "header_originale": "Nome", "confidenza": "alta", "formato": "nome cognome"}
  },
  "trasformazioni": [
    {
      "descrizione": "La colonna X contiene date di arrivo nel formato 'gio 8/10'. Derivare presenza_8..11 dal range arrivo-partenza.",
      "colonne_coinvolte": [3, 5],
      "tipo": "date_to_presenze"
    }
  ],
  "note_mapping": "spiegazione breve"
}

Le chiavi di "mapping" sono gli indici delle colonne (0, 1, 2, ...).
"confidenza" può essere: "alta", "media", "bassa".
"trasformazioni" è opzionale — usalo quando i dati richiedono interpretazione oltre al semplice mapping."""

        sample_text = f"HEADER: {json.dumps(header, ensure_ascii=False)}\n\n"
        sample_text += "RIGHE DI ESEMPIO:\n"
        for i, row in enumerate(sample_rows):
            sample_text += f"Riga {i+1}: {json.dumps(row, ensure_ascii=False)}\n"

        client = anthropic.Anthropic(api_key=api_key)

        try:
            response = client.messages.create(
                model='claude-haiku-4-5-20251001',
                max_tokens=2048,
                system=system_prompt,
                messages=[{'role': 'user', 'content': sample_text}],
            )
            raw = response.content[0].text.strip()
            if raw.startswith('```'):
                raw = raw.split('\n', 1)[1] if '\n' in raw else raw[3:]
                if raw.endswith('```'):
                    raw = raw[:-3]
                raw = raw.strip()

            mapping_result = json.loads(raw)

            inp = response.usage.input_tokens
            out = response.usage.output_tokens
            cost = (inp * 0.80 + out * 4.00) / 1_000_000

            # Salva i dati in un file temporaneo per lo step 2
            import tempfile, uuid
            import_id = str(uuid.uuid4())
            tmp_path = os.path.join(tempfile.gettempdir(), f'saba_import_{import_id}.json')
            with open(tmp_path, 'w') as tf:
                json.dump({'header': header, 'rows': all_rows}, tf, ensure_ascii=False)

            return jsonify(
                ok=True,
                import_id=import_id,
                header=header,
                mapping=mapping_result.get('mapping', {}),
                trasformazioni=mapping_result.get('trasformazioni', []),
                note_mapping=mapping_result.get('note_mapping', ''),
                sample_rows=sample_rows,
                total_rows=len(all_rows),
                usage={'input': inp, 'output': out, 'cost_eur': round(cost * 0.92, 4)},
            )

        except json.JSONDecodeError:
            return jsonify(ok=False, error=f'Risposta LLM non valida: {raw[:300]}'), 500
        except Exception as e:
            return jsonify(ok=False, error=str(e)), 500

    @app.post('/api/import/confirm')
    def import_confirm():
        """Step 2: Usa il LLM per interpretare OGNI riga e produrre i record Guest."""
        import anthropic, tempfile

        data = request.get_json()
        import_id = data.get('import_id')
        mapping = data.get('mapping', {})
        trasformazioni = data.get('trasformazioni', [])

        if not import_id:
            return jsonify(ok=False, error='import_id mancante'), 400

        tmp_path = os.path.join(tempfile.gettempdir(), f'saba_import_{import_id}.json')
        if not os.path.exists(tmp_path):
            return jsonify(ok=False, error='Sessione di import scaduta. Ricarica il file.'), 400

        with open(tmp_path) as tf:
            file_data = json.load(tf)

        header = file_data['header']
        rows = file_data['rows']

        # Controlla se servono trasformazioni complesse (date → presenze)
        needs_llm = any(t.get('tipo') == 'date_to_presenze' for t in trasformazioni)

        if needs_llm:
            # Manda TUTTE le righe al LLM per interpretazione
            api_key = os.environ.get('ANTHROPIC_API_KEY')
            if not api_key:
                return jsonify(ok=False, error='ANTHROPIC_API_KEY non configurata'), 500

            # Costruisci mapping descrittivo
            mapping_desc = {}
            for idx_str, info in mapping.items():
                campo = info.get('campo', 'IGNORA') if isinstance(info, dict) else info
                if campo != 'IGNORA':
                    mapping_desc[header[int(idx_str)]] = campo

            system_prompt = f"""Sei un assistente che converte righe di un Excel in record JSON per un database di rooming.

L'evento è dall'8 all'11 ottobre.

Mapping colonne stabilito: {json.dumps(mapping_desc, ensure_ascii=False)}

Trasformazioni richieste: {json.dumps(trasformazioni, ensure_ascii=False)}

Per ogni riga, produci un oggetto JSON con TUTTI questi campi:
- cognome (MAIUSCOLO), nome, email, telefono, sede_lavoro
- presenza_8, presenza_9, presenza_10, presenza_11 (true/false)
- volo_arrivo, volo_partenza
- aeroporto_partenza, aeroporto_arrivo
- pickup_bus_andata, pickup_bus_ritorno
- parcheggio_linate (true/false), parcheggio_hotel (true/false)
- divide_stanza_con, restrizioni_alimentari
- tipo_camera, note_form, note
- data_nascita (formato GG/MM/AAAA)

REGOLE PRESENZE:
- Se arriva giorno X e parte giorno Y: presente da X a Y-1
- "arrivo gio 8/10" = arriva l'8 ottobre (gio=giovedì, 8/10=8 ottobre)
- "partenza 10/10" = parte il 10 ottobre
- Se solo arrivo senza partenza: presente dall'arrivo fino all'11
- Se solo partenza senza arrivo: presente dall'8 fino a partenza-1
- Campi non presenti nel file → null (stringhe) o false (booleani)
- Righe vuote (senza cognome/nome) → skippa, non includerle

Rispondi SOLO con JSON valido (array di oggetti), niente markdown."""

            # Manda a blocchi di 50 righe per non superare i limiti
            all_guests = []
            CHUNK = 50
            client = anthropic.Anthropic(api_key=api_key)
            total_inp, total_out = 0, 0

            for chunk_start in range(0, len(rows), CHUNK):
                chunk = rows[chunk_start:chunk_start + CHUNK]
                rows_text = f"HEADER: {json.dumps(header, ensure_ascii=False)}\n\n"
                for i, row in enumerate(chunk):
                    rows_text += f"Riga {chunk_start + i + 1}: {json.dumps(row, ensure_ascii=False)}\n"

                response = client.messages.create(
                    model='claude-haiku-4-5-20251001',
                    max_tokens=4096,
                    system=system_prompt,
                    messages=[{'role': 'user', 'content': rows_text}],
                )
                raw = response.content[0].text.strip()
                if raw.startswith('```'):
                    raw = raw.split('\n', 1)[1] if '\n' in raw else raw[3:]
                    if raw.endswith('```'):
                        raw = raw[:-3]
                    raw = raw.strip()

                chunk_guests = json.loads(raw)
                if isinstance(chunk_guests, dict) and 'guests' in chunk_guests:
                    chunk_guests = chunk_guests['guests']
                all_guests.extend(chunk_guests)
                total_inp += response.usage.input_tokens
                total_out += response.usage.output_tokens

            # Inserisci nel DB
            added = 0
            skipped = 0
            for gd in all_guests:
                cognome = (gd.get('cognome') or '').strip()
                if not cognome:
                    skipped += 1
                    continue
                g = Guest(
                    cognome=cognome,
                    nome=(gd.get('nome') or '').strip(),
                    email=gd.get('email'),
                    telefono=gd.get('telefono'),
                    sede_lavoro=gd.get('sede_lavoro'),
                    presenza_8=_parse_bool(gd.get('presenza_8')),
                    presenza_9=_parse_bool(gd.get('presenza_9')),
                    presenza_10=_parse_bool(gd.get('presenza_10')),
                    presenza_11=_parse_bool(gd.get('presenza_11')),
                    volo_arrivo=gd.get('volo_arrivo'),
                    volo_partenza=gd.get('volo_partenza'),
                    aeroporto_partenza=gd.get('aeroporto_partenza'),
                    aeroporto_arrivo=gd.get('aeroporto_arrivo'),
                    pickup_bus_andata=gd.get('pickup_bus_andata'),
                    pickup_bus_ritorno=gd.get('pickup_bus_ritorno'),
                    parcheggio_linate=_parse_bool(gd.get('parcheggio_linate')),
                    parcheggio_hotel=_parse_bool(gd.get('parcheggio_hotel')),
                    divide_stanza_con=gd.get('divide_stanza_con'),
                    restrizioni_alimentari=gd.get('restrizioni_alimentari'),
                    tipo_camera=gd.get('tipo_camera'),
                    note_form=gd.get('note_form'),
                    note=gd.get('note'),
                    data_nascita=gd.get('data_nascita'),
                    source='xlsx',
                )
                db.session.add(g)
                added += 1

            db.session.commit()
            log_audit('rooming', 'Guest', None, 'import',
                      changes={'count': added},
                      summary=f'Importati {added} ospiti da XLSX')
            cost = (total_inp * 0.80 + total_out * 4.00) / 1_000_000

            try:
                os.remove(tmp_path)
            except OSError:
                pass

            return jsonify(ok=True, added=added, skipped=skipped,
                           usage={'input': total_inp, 'output': total_out,
                                  'cost_eur': round(cost * 0.92, 4)})

        # ── Fallback: mapping diretto senza LLM (nessuna trasformazione) ──
        col_map = {}
        for idx_str, info in mapping.items():
            campo = info.get('campo', 'IGNORA') if isinstance(info, dict) else info
            if campo != 'IGNORA':
                col_map[campo] = int(idx_str)

        if 'cognome' not in col_map and 'nome_completo' not in col_map:
            return jsonify(ok=False, error='Nessuna colonna mappata a cognome o nome completo'), 400

        added = 0
        skipped = 0

        for row in rows:
            if 'nome_completo' in col_map:
                full = row[col_map['nome_completo']].strip() if col_map['nome_completo'] < len(row) else ''
                if not full:
                    skipped += 1
                    continue
                formato = None
                for idx_str, info in mapping.items():
                    if isinstance(info, dict) and info.get('campo') == 'nome_completo':
                        formato = info.get('formato', 'nome cognome')
                        break
                parts = full.split(None, 1)
                if formato and 'cognome' in formato.split()[0].lower():
                    cognome = parts[0] if parts else full
                    nome = parts[1] if len(parts) > 1 else ''
                else:
                    nome = parts[0] if parts else ''
                    cognome = parts[1] if len(parts) > 1 else full
            else:
                cognome = row[col_map['cognome']].strip() if col_map.get('cognome') is not None and col_map['cognome'] < len(row) else ''
                if not cognome:
                    skipped += 1
                    continue
                nome = row[col_map['nome']].strip() if col_map.get('nome') is not None and col_map['nome'] < len(row) else ''

            def get_val(campo):
                idx = col_map.get(campo)
                if idx is not None and idx < len(row):
                    v = row[idx].strip()
                    return v if v else None
                return None

            g = Guest(
                cognome=cognome,
                nome=nome,
                email=get_val('email'),
                telefono=get_val('telefono'),
                sede_lavoro=get_val('sede_lavoro'),
                presenza_8=_parse_bool(get_val('presenza_8')),
                presenza_9=_parse_bool(get_val('presenza_9')),
                presenza_10=_parse_bool(get_val('presenza_10')),
                presenza_11=_parse_bool(get_val('presenza_11')),
                volo_arrivo=get_val('volo_arrivo'),
                volo_partenza=get_val('volo_partenza'),
                aeroporto_partenza=get_val('aeroporto_partenza'),
                aeroporto_arrivo=get_val('aeroporto_arrivo'),
                pickup_bus_andata=get_val('pickup_bus_andata'),
                pickup_bus_ritorno=get_val('pickup_bus_ritorno'),
                parcheggio_linate=_parse_bool(get_val('parcheggio_linate')),
                parcheggio_hotel=_parse_bool(get_val('parcheggio_hotel')),
                divide_stanza_con=get_val('divide_stanza_con'),
                restrizioni_alimentari=get_val('restrizioni_alimentari'),
                tipo_camera=get_val('tipo_camera'),
                note_form=get_val('note_form'),
                note=get_val('note'),
                data_nascita=get_val('data_nascita'),
                source='xlsx',
            )
            db.session.add(g)
            added += 1

        db.session.commit()
        log_audit('rooming', 'Guest', None, 'import',
                  changes={'count': added},
                  summary=f'Importati {added} ospiti da XLSX')

        try:
            os.remove(tmp_path)
        except OSError:
            pass

        return jsonify(ok=True, added=added, skipped=skipped)

    # ── STANZE PER GIORNO ───────────────────────────────────────────────────

    @app.get('/api/stanze/<int:giorno>')
    def stanze_giorno(giorno):
        """Calcola stanze necessarie per un giorno (8, 9, 10, 11)."""
        if giorno not in (8, 9, 10, 11):
            return jsonify(ok=False, error='Giorno non valido'), 400

        campo = f'presenza_{giorno}'
        presenti = Guest.query.filter(Guest.deleted==False, getattr(Guest, campo) == True).order_by(
            Guest.cognome, Guest.nome).all()

        # Raggruppa per stanze: chi condivide conta come 1 stanza
        stanze = []       # lista di liste di nomi
        assegnati = set() # id già assegnati a una stanza

        for g in presenti:
            if g.id in assegnati:
                continue

            stanza = [g]
            assegnati.add(g.id)

            if g.divide_stanza_con and g.divide_stanza_con.strip():
                # Cerca i compagni di stanza tra i presenti
                compagni_nomi = [n.strip().lower() for n in g.divide_stanza_con.split(',')]
                for p in presenti:
                    if p.id in assegnati:
                        continue
                    nome_completo = f'{p.nome} {p.cognome}'.lower()
                    cognome_lower = p.cognome.lower()
                    nome_lower = p.nome.lower()
                    # Match flessibile: nome completo, solo cognome, o solo nome
                    for cn in compagni_nomi:
                        if cn and (cn in nome_completo or cn in cognome_lower
                                or cn in nome_lower or (cognome_lower and cognome_lower in cn)
                                or (nome_lower and nome_lower in cn)):
                            stanza.append(p)
                            assegnati.add(p.id)
                            break

            stanze.append(stanza)

        # Serializza
        result = []
        for stanza in stanze:
            result.append({
                'ospiti': [
                    {'id': g.id, 'cognome': g.cognome, 'nome': g.nome,
                     'tipo_camera': g.tipo_camera or '',
                     'divide_stanza_con': g.divide_stanza_con or ''}
                    for g in stanza
                ],
                'tipo_camera': stanza[0].tipo_camera or '',
            })

        return jsonify(
            ok=True,
            giorno=giorno,
            totale_presenti=len(presenti),
            totale_stanze=len(stanze),
            stanze=result,
        )

    # ── VOLI RAGGRUPPATI ────────────────────────────────────────────────────

    @app.get('/api/voli/<tipo>')
    def voli_raggruppati(tipo):
        """Raggruppa ospiti per volo. tipo = 'andata' o 'ritorno'."""
        if tipo not in ('andata', 'ritorno'):
            return jsonify(ok=False, error='Tipo non valido (andata/ritorno)'), 400

        campo = Guest.volo_arrivo if tipo == 'andata' else Guest.volo_partenza
        guests = Guest.query.filter(Guest.deleted==False, campo.isnot(None), campo != '').order_by(
            campo, Guest.cognome, Guest.nome).all()

        # Raggruppa per codice volo
        gruppi = {}
        for g in guests:
            volo = (g.volo_arrivo if tipo == 'andata' else g.volo_partenza).strip()
            if not volo:
                continue
            if volo not in gruppi:
                gruppi[volo] = []
            gruppi[volo].append({
                'id': g.id,
                'cognome': g.cognome,
                'nome': g.nome,
                'sede_lavoro': g.sede_lavoro or '',
                'aeroporto': (g.aeroporto_partenza if tipo == 'andata' else g.aeroporto_arrivo) or '',
            })

        # Ordina per codice volo
        result = []
        for volo in sorted(gruppi.keys()):
            result.append({
                'volo': volo,
                'passeggeri': gruppi[volo],
                'totale': len(gruppi[volo]),
            })

        senza_volo = Guest.query.filter(Guest.deleted==False).filter(
            (campo.is_(None)) | (campo == '')
        ).count()

        return jsonify(
            ok=True,
            tipo=tipo,
            gruppi=result,
            totale_con_volo=len(guests),
            totale_senza_volo=senza_volo,
            totale_voli=len(result),
        )

    # ── PNR GROUPS ──────────────────────────────────────────────────────────

    @app.get('/api/pnr')
    def pnr_list():
        """Lista PNR groups con ospiti assegnati e statistiche."""
        import re as _re
        groups = PnrGroup.query.order_by(PnrGroup.volo_andata, PnrGroup.pnr_code).all()
        senza_pnr = Guest.query.filter(Guest.deleted==False, Guest.pnr_group_id.is_(None)).count()

        result = []
        totale_assegnati = 0
        for g in groups:
            assigned = [{
                'id': p.id, 'cognome': p.cognome, 'nome': p.nome,
                'sede_lavoro': p.sede_lavoro or '',
            } for p in g.guests]
            totale_assegnati += len(assigned)
            # Parse rotta in origin/dest (es. LINPMO → LIN / PMO)
            orig_a = g.rotta_andata[:3] if g.rotta_andata and len(g.rotta_andata) >= 6 else ''
            dest_a = g.rotta_andata[3:] if g.rotta_andata and len(g.rotta_andata) >= 6 else ''
            orig_r = g.rotta_ritorno[:3] if g.rotta_ritorno and len(g.rotta_ritorno) >= 6 else ''
            dest_r = g.rotta_ritorno[3:] if g.rotta_ritorno and len(g.rotta_ritorno) >= 6 else ''
            result.append({
                'id': g.id,
                'pnr_code': g.pnr_code,
                'group_name': g.group_name or '',
                'seats': g.seats,
                'assigned': len(assigned),
                'available': g.seats - len(assigned),
                'volo_andata': g.volo_andata or '',
                'data_andata': g.data_andata or '',
                'rotta_andata': g.rotta_andata or '',
                'origin_andata': orig_a,
                'dest_andata': dest_a,
                'orario_andata': g.orario_andata or '',
                'volo_ritorno': g.volo_ritorno or '',
                'data_ritorno': g.data_ritorno or '',
                'rotta_ritorno': g.rotta_ritorno or '',
                'origin_ritorno': orig_r,
                'dest_ritorno': dest_r,
                'orario_ritorno': g.orario_ritorno or '',
                'passeggeri': assigned,
            })

        return jsonify(
            ok=True,
            groups=result,
            totale_groups=len(result),
            totale_seats=sum(g.seats for g in groups),
            totale_assegnati=totale_assegnati,
            senza_pnr=senza_pnr,
        )

    @app.post('/api/pnr/parse')
    def pnr_parse():
        """Parsa testo PNR Amadeus e restituisce gruppi estratti."""
        import re
        text = (request.get_json() or {}).get('text', '')
        if not text.strip():
            return jsonify(ok=False, error='Testo vuoto'), 400

        # Split per blocchi RLR
        blocks = re.split(r'---?\s*(?:AXR\s+)?RLR\s*---?', text)
        groups = []

        for block in blocks:
            block = block.strip()
            if not block:
                continue

            # PNR code: ultimo token della riga RP/
            pnr_match = re.search(r'RP/\S+\s+\S+\s+\S+\s+(\w{6})', block)
            pnr_code = pnr_match.group(1) if pnr_match else None
            if not pnr_code:
                continue

            # Seats + group name: "0. 32PAOLACATANIADOS  NM: 0"
            group_match = re.search(r'0\.\s*(\d+)(\w+)\s+NM:', block)
            seats = int(group_match.group(1)) if group_match else 0
            group_name = group_match.group(2) if group_match else ''

            # Voli: "1  AZ1765 S 08OCT 4 LINPMO HK32  0955 1135  *1A/E*"
            flights = re.findall(
                r'\d+\s+(\w{2}\d{3,5})\s+\w\s+(\d{2}[A-Z]{3})\s+\d\s+(\w{6})\s+HK\d+\s+(\d{4})\s+(\d{4})',
                block
            )

            volo_andata = volo_ritorno = ''
            data_andata = data_ritorno = ''
            rotta_andata = rotta_ritorno = ''
            orario_andata = orario_ritorno = ''

            if len(flights) >= 1:
                volo_andata = flights[0][0]
                data_andata = flights[0][1]
                rotta_andata = flights[0][2]
                orario_andata = f"{flights[0][3]}-{flights[0][4]}"
            if len(flights) >= 2:
                volo_ritorno = flights[1][0]
                data_ritorno = flights[1][1]
                rotta_ritorno = flights[1][2]
                orario_ritorno = f"{flights[1][3]}-{flights[1][4]}"

            groups.append({
                'pnr_code': pnr_code,
                'group_name': group_name,
                'seats': seats,
                'volo_andata': volo_andata,
                'data_andata': data_andata,
                'rotta_andata': rotta_andata,
                'orario_andata': orario_andata,
                'volo_ritorno': volo_ritorno,
                'data_ritorno': data_ritorno,
                'rotta_ritorno': rotta_ritorno,
                'orario_ritorno': orario_ritorno,
            })

        return jsonify(ok=True, groups=groups)

    @app.post('/api/pnr/import')
    def pnr_import():
        """Salva i PNR groups parsati nel DB."""
        data = request.get_json()
        groups_data = data.get('groups', [])
        results = []

        for gd in groups_data:
            pnr_code = (gd.get('pnr_code') or '').strip()
            if not pnr_code:
                continue
            # Dedup: se esiste già, aggiorna
            existing = PnrGroup.query.filter_by(pnr_code=pnr_code).first()
            if existing:
                for f in ('group_name', 'seats', 'volo_andata', 'data_andata',
                          'rotta_andata', 'orario_andata', 'volo_ritorno',
                          'data_ritorno', 'rotta_ritorno', 'orario_ritorno'):
                    if gd.get(f) is not None:
                        setattr(existing, f, gd[f])
                results.append({'pnr_code': pnr_code, 'action': 'updated', 'id': existing.id})
            else:
                pg = PnrGroup(
                    pnr_code=pnr_code,
                    group_name=gd.get('group_name', ''),
                    seats=gd.get('seats', 0),
                    volo_andata=gd.get('volo_andata'),
                    data_andata=gd.get('data_andata'),
                    rotta_andata=gd.get('rotta_andata'),
                    orario_andata=gd.get('orario_andata'),
                    volo_ritorno=gd.get('volo_ritorno'),
                    data_ritorno=gd.get('data_ritorno'),
                    rotta_ritorno=gd.get('rotta_ritorno'),
                    orario_ritorno=gd.get('orario_ritorno'),
                )
                db.session.add(pg)
                db.session.flush()
                results.append({'pnr_code': pnr_code, 'action': 'added', 'id': pg.id})

        db.session.commit()
        log_audit('rooming', 'PnrGroup', None, 'import',
                  changes={'count': len(results)},
                  summary=f'Importati {len(results)} gruppi PNR')
        return jsonify(ok=True, results=results)

    @app.post('/api/pnr/<int:group_id>/assign')
    def pnr_assign(group_id):
        """Assegna ospiti a un PNR group."""
        pg = PnrGroup.query.get_or_404(group_id)
        data = request.get_json()
        guest_ids = data.get('guest_ids', [])

        assigned = 0
        for gid in guest_ids:
            guest = Guest.query.get(gid)
            if guest:
                guest.pnr_group_id = group_id
                # Popola anche i campi volo sul guest
                guest.volo_arrivo = pg.volo_andata
                guest.volo_partenza = pg.volo_ritorno
                origin_a = pg.rotta_andata[:3] if pg.rotta_andata and len(pg.rotta_andata) >= 6 else ''
                dest_r = pg.rotta_ritorno[3:] if pg.rotta_ritorno and len(pg.rotta_ritorno) >= 6 else ''
                if origin_a:
                    guest.aeroporto_partenza = origin_a
                if dest_r:
                    guest.aeroporto_arrivo = dest_r
                assigned += 1

        current_count = Guest.query.filter_by(deleted=False, pnr_group_id=group_id).count()
        overbooking = current_count > pg.seats

        db.session.commit()
        for gid in guest_ids:
            guest = Guest.query.get(gid)
            if guest:
                log_audit('rooming', 'Guest', guest.id, 'assign',
                          changes={'pnr_group_id': {'old': None, 'new': group_id}},
                          summary=f'{guest.nome_completo} assegnato a PNR {pg.pnr_code}')
        return jsonify(ok=True, assigned=assigned, total=current_count,
                       seats=pg.seats, overbooking=overbooking)

    @app.post('/api/pnr/<int:group_id>/unassign')
    def pnr_unassign(group_id):
        """Rimuovi ospiti da un PNR group."""
        data = request.get_json()
        guest_ids = data.get('guest_ids', [])

        unassigned_guests = []
        for gid in guest_ids:
            guest = Guest.query.get(gid)
            if guest and guest.pnr_group_id == group_id:
                unassigned_guests.append(guest)
                guest.pnr_group_id = None

        db.session.commit()
        for guest in unassigned_guests:
            log_audit('rooming', 'Guest', guest.id, 'unassign',
                      changes={'pnr_group_id': {'old': group_id, 'new': None}},
                      summary=f'{guest.nome_completo} rimosso da PNR')
        return jsonify(ok=True)

    @app.delete('/api/pnr/<int:group_id>')
    def pnr_delete(group_id):
        """Elimina un PNR group (scollega ospiti)."""
        pg = PnrGroup.query.get_or_404(group_id)
        Guest.query.filter_by(deleted=False, pnr_group_id=group_id).update({'pnr_group_id': None})
        db.session.delete(pg)
        db.session.commit()
        return jsonify(ok=True)

    @app.post('/api/pnr/auto-assign')
    def pnr_auto_assign():
        """Auto-assegna ospiti ai PNR in base ai voli indicati.
        Se confirm=true applica, altrimenti restituisce solo preview."""
        import re

        def normalize_flight(s):
            """Estrae codice volo puro: 'AZ 1773 - 08:25' → 'AZ1773'."""
            if not s:
                return ''
            # Rimuovi tutto dopo il trattino/spazio con orario
            s = re.split(r'\s*[-–]\s*\d{2}[.:]\d{2}', s)[0]
            # Rimuovi spazi interni e normalizza
            s = re.sub(r'\s+', '', s).upper()
            return s

        data = request.get_json() or {}
        confirm = data.get('confirm', False)

        groups = PnrGroup.query.all()
        # Indice: (volo_andata, volo_ritorno) → lista PnrGroup ordinata per posti
        pnr_index = {}
        for pg in groups:
            key = (normalize_flight(pg.volo_andata),
                   normalize_flight(pg.volo_ritorno))
            pnr_index.setdefault(key, []).append(pg)

        # Conta posti già occupati per PNR
        seats_used = {}
        for pg in groups:
            seats_used[pg.id] = Guest.query.filter_by(deleted=False, pnr_group_id=pg.id).count()

        # Ospiti non ancora assegnati
        unassigned = Guest.query.filter(
            Guest.deleted==False, Guest.pnr_group_id.is_(None)
        ).order_by(Guest.cognome, Guest.nome).all()

        matched = []       # match esatto andata+ritorno
        partial = []       # solo andata o solo ritorno matcha
        no_match = []      # ha voli ma nessun PNR corrisponde
        no_flights = []    # nessun volo compilato
        overflow = []      # matchato ma PNR pieno

        def _guest_base(g):
            notes = ' | '.join(filter(None, [g.note_form, g.note]))
            return {
                'id': g.id, 'cognome': g.cognome, 'nome': g.nome,
                'sede_lavoro': g.sede_lavoro or '',
                'note': notes,
                'email_log_id': g.email_log_id,
            }

        for g in unassigned:
            andata = normalize_flight(g.volo_arrivo)
            ritorno = normalize_flight(g.volo_partenza)

            if not andata and not ritorno:
                no_flights.append({
                    **_guest_base(g),
                    'reason': 'Nessun volo compilato',
                })
                continue

            # Prova match esatto
            key = (andata, ritorno)
            candidates = pnr_index.get(key, [])

            assigned_pg = None
            for pg in candidates:
                if seats_used.get(pg.id, 0) < pg.seats:
                    assigned_pg = pg
                    break

            if assigned_pg:
                orig_a = assigned_pg.rotta_andata[:3] if assigned_pg.rotta_andata and len(assigned_pg.rotta_andata) >= 6 else ''
                dest_a = assigned_pg.rotta_andata[3:] if assigned_pg.rotta_andata and len(assigned_pg.rotta_andata) >= 6 else ''
                orig_r = assigned_pg.rotta_ritorno[:3] if assigned_pg.rotta_ritorno and len(assigned_pg.rotta_ritorno) >= 6 else ''
                dest_r = assigned_pg.rotta_ritorno[3:] if assigned_pg.rotta_ritorno and len(assigned_pg.rotta_ritorno) >= 6 else ''
                matched.append({
                    **_guest_base(g),
                    'pnr_id': assigned_pg.id, 'pnr_code': assigned_pg.pnr_code,
                    'volo_andata': assigned_pg.volo_andata,
                    'volo_ritorno': assigned_pg.volo_ritorno,
                    'rotta_andata': f'{orig_a}→{dest_a}' if orig_a else '',
                    'rotta_ritorno': f'{orig_r}→{dest_r}' if orig_r else '',
                    'orario_andata': assigned_pg.orario_andata or '',
                    'orario_ritorno': assigned_pg.orario_ritorno or '',
                })
                seats_used[assigned_pg.id] = seats_used.get(assigned_pg.id, 0) + 1
                continue

            # Match esatto ma tutti pieni → overflow
            if candidates:
                overflow.append({
                    **_guest_base(g),
                    'volo_arrivo': andata, 'volo_partenza': ritorno,
                    'pnr_codes': [pg.pnr_code for pg in candidates],
                    'reason': f'PNR pieni: {", ".join(pg.pnr_code for pg in candidates)}',
                })
                continue

            # Prova match parziale
            partial_matches = []
            for (k_a, k_r), pgs in pnr_index.items():
                if andata and andata == k_a and ritorno != k_r:
                    for pg in pgs:
                        partial_matches.append({
                            'pnr_code': pg.pnr_code, 'type': 'andata',
                            'match_flight': k_a,
                            'mismatch': f'Ritorno: ospite={ritorno or "—"} PNR={k_r}',
                        })
                elif ritorno and ritorno == k_r and andata != k_a:
                    for pg in pgs:
                        partial_matches.append({
                            'pnr_code': pg.pnr_code, 'type': 'ritorno',
                            'match_flight': k_r,
                            'mismatch': f'Andata: ospite={andata or "—"} PNR={k_a}',
                        })

            if partial_matches:
                partial.append({
                    **_guest_base(g),
                    'volo_arrivo': andata, 'volo_partenza': ritorno,
                    'partial_matches': partial_matches,
                })
            else:
                no_match.append({
                    **_guest_base(g),
                    'volo_arrivo': andata, 'volo_partenza': ritorno,
                    'reason': 'Nessun PNR con questi voli',
                })

        # Se confirm, applica i match esatti
        applied = 0
        if confirm and matched:
            for m in matched:
                guest = Guest.query.get(m['id'])
                if guest:
                    guest.pnr_group_id = m['pnr_id']
                    pg = PnrGroup.query.get(m['pnr_id'])
                    if pg:
                        guest.volo_arrivo = pg.volo_andata
                        guest.volo_partenza = pg.volo_ritorno
                        origin = pg.rotta_andata[:3] if pg.rotta_andata and len(pg.rotta_andata) >= 6 else ''
                        dest = pg.rotta_ritorno[3:] if pg.rotta_ritorno and len(pg.rotta_ritorno) >= 6 else ''
                        if origin:
                            guest.aeroporto_partenza = origin
                        if dest:
                            guest.aeroporto_arrivo = dest
                    applied += 1
            db.session.commit()
            log_audit('rooming', 'PnrGroup', None, 'assign',
                      changes={'count': applied},
                      summary=f'Auto-assegnazione PNR: {applied} ospiti')

        # Riepilogo posti per PNR
        pnr_summary = []
        for pg in sorted(groups, key=lambda x: x.pnr_code):
            used = seats_used.get(pg.id, 0)
            pnr_summary.append({
                'pnr_code': pg.pnr_code,
                'seats': pg.seats,
                'used': used,
                'available': pg.seats - used,
                'overbooking': used > pg.seats,
                'volo_andata': pg.volo_andata,
                'volo_ritorno': pg.volo_ritorno,
                'rotta_andata': pg.rotta_andata or '',
                'rotta_ritorno': pg.rotta_ritorno or '',
                'data_andata': pg.data_andata or '',
                'data_ritorno': pg.data_ritorno or '',
                'orario_andata': pg.orario_andata or '',
                'orario_ritorno': pg.orario_ritorno or '',
            })

        return jsonify(
            ok=True,
            confirmed=confirm,
            applied=applied,
            matched=matched,
            partial=partial,
            overflow=overflow,
            no_match=no_match,
            no_flights=no_flights,
            pnr_summary=pnr_summary,
            totals={
                'matched': len(matched),
                'partial': len(partial),
                'overflow': len(overflow),
                'no_match': len(no_match),
                'no_flights': len(no_flights),
            },
        )

    @app.post('/api/pnr/assign-by-code')
    def pnr_assign_by_code():
        """Assegna un singolo ospite a un PNR tramite pnr_code."""
        data = request.get_json()
        pnr_code = (data.get('pnr_code') or '').strip()
        guest_id = data.get('guest_id')
        if not pnr_code or not guest_id:
            return jsonify(ok=False, error='Dati mancanti'), 400

        pg = PnrGroup.query.filter_by(pnr_code=pnr_code).first()
        if not pg:
            return jsonify(ok=False, error=f'PNR {pnr_code} non trovato'), 404

        guest = Guest.query.get(guest_id)
        if not guest:
            return jsonify(ok=False, error='Ospite non trovato'), 404

        guest.pnr_group_id = pg.id
        guest.volo_arrivo = pg.volo_andata
        guest.volo_partenza = pg.volo_ritorno
        origin = pg.rotta_andata[:3] if pg.rotta_andata and len(pg.rotta_andata) >= 6 else ''
        dest = pg.rotta_ritorno[3:] if pg.rotta_ritorno and len(pg.rotta_ritorno) >= 6 else ''
        if origin:
            guest.aeroporto_partenza = origin
        if dest:
            guest.aeroporto_arrivo = dest
        db.session.commit()
        return jsonify(ok=True)

    @app.get('/api/pnr/unassigned')
    def pnr_unassigned():
        """Lista ospiti non assegnati a nessun PNR."""
        guests = Guest.query.filter(
            Guest.deleted==False, Guest.pnr_group_id.is_(None)
        ).order_by(Guest.cognome, Guest.nome).all()
        return jsonify(ok=True, guests=[{
            'id': g.id, 'cognome': g.cognome, 'nome': g.nome,
            'sede_lavoro': g.sede_lavoro or '',
            'aeroporto_partenza': g.aeroporto_partenza or '',
        } for g in guests])

    # ── ASSEGNAZIONE CAMERE ─────────────────────────────────────────────────

    @app.get('/api/camere/<int:notte>')
    def camere_disponibilita(notte):
        """Mostra disponibilità camere vs assegnazioni per una notte."""
        if notte not in (8, 9, 10, 11):
            return jsonify(ok=False, error='Notte non valida'), 400

        contratti = RoomContract.query.filter_by(notte=notte).order_by(
            RoomContract.tariffa_netta).all()

        campo = f'presenza_{notte}'
        presenti = Guest.query.filter(Guest.deleted==False, getattr(Guest, campo) == True).order_by(
            Guest.cognome, Guest.nome).all()

        # Calcola stanze necessarie (come endpoint stanze)
        assegnati_ids = set()
        stanze_necessarie = []
        for g in presenti:
            if g.id in assegnati_ids:
                continue
            stanza = [g]
            assegnati_ids.add(g.id)
            if g.divide_stanza_con and g.divide_stanza_con.strip():
                compagni = [n.strip().lower() for n in g.divide_stanza_con.split(',')]
                for p in presenti:
                    if p.id in assegnati_ids:
                        continue
                    nc = f'{p.nome} {p.cognome}'.lower()
                    cl = p.cognome.lower()
                    nl = p.nome.lower()
                    for cn in compagni:
                        if cn and (cn in nc or cn in cl or cn in nl or (cl and cl in cn) or (nl and nl in cn)):
                            stanza.append(p)
                            assegnati_ids.add(p.id)
                            break
            stanze_necessarie.append(stanza)

        # Conta assegnazioni per tipo
        assegnazioni_per_tipo = {}
        non_assegnati = []
        for stanza in stanze_necessarie:
            camera = stanza[0].camera_assegnata
            if camera:
                assegnazioni_per_tipo[camera] = assegnazioni_per_tipo.get(camera, 0) + 1
            else:
                non_assegnati.append(stanza)

        result_contratti = []
        for c in contratti:
            usate = assegnazioni_per_tipo.get(c.tipo, 0)
            result_contratti.append({
                'id': c.id,
                'tipo': c.tipo,
                'disponibili': c.disponibili,
                'assegnate': usate,
                'libere': c.disponibili - usate,
                'tariffa_netta': c.tariffa_netta,
                'tariffa_lorda': c.tariffa_lorda,
            })

        result_non_assegnati = []
        for stanza in non_assegnati:
            result_non_assegnati.append({
                'ospiti': [{'id': g.id, 'cognome': g.cognome, 'nome': g.nome,
                            'tipo_camera': g.tipo_camera or ''}
                           for g in stanza],
            })

        return jsonify(
            ok=True,
            notte=notte,
            contratti=result_contratti,
            totale_stanze_necessarie=len(stanze_necessarie),
            totale_assegnate=len(stanze_necessarie) - len(non_assegnati),
            totale_non_assegnate=len(non_assegnati),
            non_assegnati=result_non_assegnati,
        )

    @app.post('/api/camere/assegna')
    def assegna_camera():
        """Assegna manualmente un tipo camera a un ospite (e al suo compagno di stanza)."""
        data = request.get_json()
        guest_id = data.get('guest_id')
        tipo_camera = data.get('tipo_camera')

        if not guest_id or not tipo_camera:
            return jsonify(ok=False, error='guest_id e tipo_camera obbligatori'), 400

        g = Guest.query.get_or_404(guest_id)
        old_room = g.camera_assegnata
        g.camera_assegnata = tipo_camera
        g.updated_at = datetime.utcnow()

        # Assegna anche ai compagni di stanza
        assegnati = [g.id]
        if g.divide_stanza_con and g.divide_stanza_con.strip():
            compagni = [n.strip().lower() for n in g.divide_stanza_con.split(',')]
            tutti = Guest.query.filter_by(deleted=False).all()
            for p in tutti:
                if p.id == g.id:
                    continue
                nc = f'{p.nome} {p.cognome}'.lower()
                cl = p.cognome.lower()
                nl = p.nome.lower()
                for cn in compagni:
                    if cn and (cn in nc or cn in cl or cn in nl or (cl and cl in cn) or (nl and nl in cn)):
                        p.camera_assegnata = tipo_camera
                        p.updated_at = datetime.utcnow()
                        assegnati.append(p.id)
                        break

        db.session.commit()
        log_audit('rooming', 'Guest', g.id, 'assign',
                  changes={'camera_assegnata': {'old': old_room, 'new': tipo_camera}},
                  summary=f'Camera assegnata a {g.nome_completo}')
        return jsonify(ok=True, assegnati=assegnati)

    @app.post('/api/camere/auto-assegna/<int:notte>')
    def auto_assegna(notte):
        """Assegna automaticamente le camere per una notte, dal tipo più economico."""
        if notte not in (8, 9, 10, 11):
            return jsonify(ok=False, error='Notte non valida'), 400

        contratti = RoomContract.query.filter_by(notte=notte).order_by(
            RoomContract.tariffa_netta).all()

        campo = f'presenza_{notte}'
        presenti = Guest.query.filter(Guest.deleted==False, getattr(Guest, campo) == True).order_by(
            Guest.cognome, Guest.nome).all()

        # Calcola stanze
        assegnati_ids = set()
        stanze = []
        for g in presenti:
            if g.id in assegnati_ids:
                continue
            stanza = [g]
            assegnati_ids.add(g.id)
            if g.divide_stanza_con and g.divide_stanza_con.strip():
                compagni = [n.strip().lower() for n in g.divide_stanza_con.split(',')]
                for p in presenti:
                    if p.id in assegnati_ids:
                        continue
                    nc = f'{p.nome} {p.cognome}'.lower()
                    cl = p.cognome.lower()
                    nl = p.nome.lower()
                    for cn in compagni:
                        if cn and (cn in nc or cn in cl or cn in nl or (cl and cl in cn) or (nl and nl in cn)):
                            stanza.append(p)
                            assegnati_ids.add(p.id)
                            break
            stanze.append(stanza)

        # Filtra solo stanze non ancora assegnate
        stanze_da_assegnare = [s for s in stanze if not s[0].camera_assegnata]

        # Assegna partendo dal tipo più economico
        assegnate = 0
        overflow = 0
        for contratto in contratti:
            # Quante già usate di questo tipo?
            gia_usate = sum(1 for s in stanze if s[0].camera_assegnata == contratto.tipo)
            libere = contratto.disponibili - gia_usate

            while libere > 0 and stanze_da_assegnare:
                stanza = stanze_da_assegnare.pop(0)
                for g in stanza:
                    g.camera_assegnata = contratto.tipo
                    g.updated_at = datetime.utcnow()
                libere -= 1
                assegnate += 1

        overflow = len(stanze_da_assegnare)
        db.session.commit()
        log_audit('rooming', 'RoomContract', None, 'assign',
                  changes={'count': assegnate, 'notte': notte},
                  summary=f'Auto-assegnazione camere notte {notte}: {assegnate} ospiti')

        return jsonify(ok=True, assegnate=assegnate, overflow=overflow,
                       messaggio=f'{assegnate} stanze assegnate' +
                       (f', {overflow} senza camera disponibile' if overflow else ''))

    @app.post('/api/camere/reset/<int:notte>')
    def reset_assegnazioni(notte):
        """Rimuove tutte le assegnazioni camera per una notte."""
        if notte not in (8, 9, 10, 11):
            return jsonify(ok=False, error='Notte non valida'), 400

        campo = f'presenza_{notte}'
        presenti = Guest.query.filter(Guest.deleted==False, getattr(Guest, campo) == True).all()
        for g in presenti:
            g.camera_assegnata = None
            g.updated_at = datetime.utcnow()
        db.session.commit()
        log_audit('rooming', 'RoomContract', None, 'update',
                  changes={'notte': notte},
                  summary=f'Reset assegnazioni camere notte {notte}')
        return jsonify(ok=True)

    # ── EXPORT XLSX ──────────────────────────────────────────────────────────

    @app.get('/api/export')
    def export_xlsx():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        guests = Guest.query.filter_by(deleted=False).order_by(Guest.cognome, Guest.nome).all()
        wb = Workbook()

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='795548')
        header_fill2 = PatternFill('solid', fgColor='6D4C41')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        def write_sheet(ws, headers, row_fn, fill=header_fill):
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            for r, g in enumerate(guests, 2):
                for c, v in enumerate(row_fn(g), 1):
                    cell = ws.cell(row=r, column=c, value=v if v is not None else '')
                    cell.border = thin_border
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        def bool_label(v):
            return 'Sì' if v else 'No'

        # ── Sheet 1: Anagrafica completa ──────────────────────────────────────
        ws = wb.active
        ws.title = 'Anagrafica'
        write_sheet(ws,
            ['Cognome', 'Nome', 'Data Nascita', 'Email', 'Telefono', 'Sede Lavoro',
             '8 Ott', '9 Ott', '10 Ott', '11 Ott',
             'Tipo Camera', 'Divide stanza con',
             'Parcheggio Linate', 'Parcheggio Hotel',
             'Restrizioni Alimentari', 'Note Form', 'Note'],
            lambda g: [g.cognome, g.nome, g.data_nascita, g.email, g.telefono, g.sede_lavoro,
                       bool_label(g.presenza_8), bool_label(g.presenza_9),
                       bool_label(g.presenza_10), bool_label(g.presenza_11),
                       g.tipo_camera, g.divide_stanza_con,
                       bool_label(g.parcheggio_linate), bool_label(g.parcheggio_hotel),
                       g.restrizioni_alimentari, g.note_form, g.note])

        # ── Sheet 2: Voli e Trasporti ─────────────────────────────────────────
        ws2 = wb.create_sheet('Voli e Trasporti')
        write_sheet(ws2,
            ['Cognome', 'Nome', 'Sede Lavoro',
             'Aeroporto Partenza', 'Volo Andata',
             'Aeroporto Arrivo', 'Volo Ritorno',
             'Pickup Bus Andata', 'Pickup Bus Ritorno'],
            lambda g: [g.cognome, g.nome, g.sede_lavoro,
                       g.aeroporto_partenza, g.volo_arrivo,
                       g.aeroporto_arrivo, g.volo_partenza,
                       g.pickup_bus_andata, g.pickup_bus_ritorno],
            fill=header_fill2)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        today = datetime.now().strftime('%Y-%m-%d')
        return send_file(buf, as_attachment=True,
                         download_name=f'rooming_flight_{today}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ── EXPORT STANZE / VOLI / CAMERE ────────────────────────────────────────

    @app.get('/api/export/stanze/<int:giorno>')
    def export_stanze(giorno):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if giorno not in (8, 9, 10, 11):
            return jsonify(ok=False, error='Giorno non valido'), 400

        campo = f'presenza_{giorno}'
        presenti = Guest.query.filter(Guest.deleted==False, getattr(Guest, campo) == True).order_by(
            Guest.cognome, Guest.nome).all()

        # Raggruppa per stanze
        stanze, assegnati = [], set()
        for g in presenti:
            if g.id in assegnati:
                continue
            stanza = [g]
            assegnati.add(g.id)
            if g.divide_stanza_con and g.divide_stanza_con.strip():
                compagni = [n.strip().lower() for n in g.divide_stanza_con.split(',')]
                for p in presenti:
                    if p.id in assegnati:
                        continue
                    nc = f'{p.nome} {p.cognome}'.lower()
                    cl, nl = p.cognome.lower(), p.nome.lower()
                    for cn in compagni:
                        if cn and (cn in nc or cn in cl or cn in nl or (cl and cl in cn) or (nl and nl in cn)):
                            stanza.append(p)
                            assegnati.add(p.id)
                            break
            stanze.append(stanza)

        wb = Workbook()
        ws = wb.active
        ws.title = f'Stanze {giorno} Ott'
        hfont = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill('solid', fgColor='795548')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['#', 'Cognome', 'Nome', 'Tipo Camera', 'Divide stanza con']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        row = 2
        for i, stanza in enumerate(stanze, 1):
            for g in stanza:
                ws.cell(row=row, column=1, value=i).border = border
                ws.cell(row=row, column=2, value=g.cognome).border = border
                ws.cell(row=row, column=3, value=g.nome).border = border
                ws.cell(row=row, column=4, value=g.tipo_camera or '').border = border
                ws.cell(row=row, column=5, value=g.divide_stanza_con or '').border = border
                row += 1

        for col in ws.columns:
            mx = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 40)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'stanze_{giorno}_ott.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.get('/api/export/voli/<tipo>')
    def export_voli(tipo):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if tipo not in ('andata', 'ritorno'):
            return jsonify(ok=False, error='Tipo non valido'), 400

        campo = Guest.volo_arrivo if tipo == 'andata' else Guest.volo_partenza
        guests = Guest.query.filter(Guest.deleted==False, campo.isnot(None), campo != '').order_by(
            campo, Guest.cognome, Guest.nome).all()

        wb = Workbook()
        ws = wb.active
        label = 'Andata' if tipo == 'andata' else 'Ritorno'
        ws.title = f'Voli {label}'
        hfont = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill('solid', fgColor='6D4C41')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['Volo', 'Cognome', 'Nome', 'Sede Lavoro', 'Aeroporto']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for r, g in enumerate(guests, 2):
            volo = (g.volo_arrivo if tipo == 'andata' else g.volo_partenza) or ''
            aeroporto = (g.aeroporto_partenza if tipo == 'andata' else g.aeroporto_arrivo) or ''
            vals = [volo, g.cognome, g.nome, g.sede_lavoro or '', aeroporto]
            for c, v in enumerate(vals, 1):
                ws.cell(row=r, column=c, value=v).border = border

        for col in ws.columns:
            mx = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 40)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'voli_{tipo}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.get('/api/export/pnr')
    def export_pnr():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        groups = PnrGroup.query.order_by(PnrGroup.volo_andata, PnrGroup.pnr_code).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'PNR Groups'
        hfont = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill('solid', fgColor='6D4C41')
        cluster_fill = PatternFill('solid', fgColor='EFEBE9')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['PNR', 'Posti', 'Volo Andata', 'Rotta', 'Data', 'Orario',
                    'Volo Ritorno', 'Rotta', 'Data', 'Orario',
                    'Cognome', 'Nome', 'Sede Lavoro']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        row = 2
        for pg in groups:
            guests = Guest.query.filter_by(deleted=False, pnr_group_id=pg.id).order_by(
                Guest.cognome, Guest.nome).all()
            if not guests:
                vals = [pg.pnr_code, pg.seats, pg.volo_andata, pg.rotta_andata,
                        pg.data_andata, pg.orario_andata, pg.volo_ritorno,
                        pg.rotta_ritorno, pg.data_ritorno, pg.orario_ritorno,
                        '', '', '']
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.border = border
                    cell.fill = cluster_fill
                row += 1
            else:
                for i, g in enumerate(guests):
                    vals = [
                        pg.pnr_code if i == 0 else '',
                        pg.seats if i == 0 else '',
                        pg.volo_andata if i == 0 else '',
                        pg.rotta_andata if i == 0 else '',
                        pg.data_andata if i == 0 else '',
                        pg.orario_andata if i == 0 else '',
                        pg.volo_ritorno if i == 0 else '',
                        pg.rotta_ritorno if i == 0 else '',
                        pg.data_ritorno if i == 0 else '',
                        pg.orario_ritorno if i == 0 else '',
                        g.cognome, g.nome, g.sede_lavoro or '',
                    ]
                    for c, v in enumerate(vals, 1):
                        cell = ws.cell(row=row, column=c, value=v)
                        cell.border = border
                        if i == 0:
                            cell.fill = cluster_fill
                    row += 1

        # Unassigned guests
        unassigned = Guest.query.filter(Guest.deleted==False, Guest.pnr_group_id.is_(None)).order_by(
            Guest.cognome, Guest.nome).all()
        if unassigned:
            row += 1
            cell = ws.cell(row=row, column=1, value='SENZA PNR')
            cell.font = Font(bold=True, color='CC0000')
            row += 1
            for g in unassigned:
                ws.cell(row=row, column=11, value=g.cognome).border = border
                ws.cell(row=row, column=12, value=g.nome).border = border
                ws.cell(row=row, column=13, value=g.sede_lavoro or '').border = border
                row += 1

        for col in ws.columns:
            mx = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 40)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name='pnr_groups.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.get('/api/export/camere/<int:notte>')
    def export_camere(notte):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if notte not in (8, 9, 10, 11):
            return jsonify(ok=False, error='Notte non valida'), 400

        contratti = RoomContract.query.filter_by(notte=notte).order_by(
            RoomContract.tariffa_netta).all()

        campo = f'presenza_{notte}'
        presenti = Guest.query.filter(Guest.deleted==False, getattr(Guest, campo) == True).order_by(
            Guest.cognome, Guest.nome).all()

        # Raggruppa per stanze
        assegnati_ids = set()
        stanze = []
        for g in presenti:
            if g.id in assegnati_ids:
                continue
            stanza = [g]
            assegnati_ids.add(g.id)
            if g.divide_stanza_con and g.divide_stanza_con.strip():
                compagni = [n.strip().lower() for n in g.divide_stanza_con.split(',')]
                for p in presenti:
                    if p.id in assegnati_ids:
                        continue
                    nc = f'{p.nome} {p.cognome}'.lower()
                    cl, nl = p.cognome.lower(), p.nome.lower()
                    for cn in compagni:
                        if cn and (cn in nc or cn in cl or cn in nl or (cl and cl in cn) or (nl and nl in cn)):
                            stanza.append(p)
                            assegnati_ids.add(p.id)
                            break
            stanze.append(stanza)

        assegnazioni_per_tipo = {}
        non_assegnati = []
        for stanza in stanze:
            camera = stanza[0].camera_assegnata
            if camera:
                assegnazioni_per_tipo[camera] = assegnazioni_per_tipo.get(camera, 0) + 1
            else:
                non_assegnati.append(stanza)

        wb = Workbook()
        hfont = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill('solid', fgColor='795548')
        hfill2 = PatternFill('solid', fgColor='6D4C41')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # Sheet 1: Contratti
        ws1 = wb.active
        ws1.title = 'Disponibilità'
        h1 = ['Tipologia', 'Disponibili', 'Assegnate', 'Libere', 'Tariffa Netta', 'Tariffa Lorda']
        for c, h in enumerate(h1, 1):
            cell = ws1.cell(row=1, column=c, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        for r, ct in enumerate(contratti, 2):
            usate = assegnazioni_per_tipo.get(ct.tipo, 0)
            vals = [ct.tipo, ct.disponibili, usate, ct.disponibili - usate,
                    ct.tariffa_netta, ct.tariffa_lorda]
            for c, v in enumerate(vals, 1):
                cell = ws1.cell(row=r, column=c, value=v)
                cell.border = border
                if c >= 5:
                    cell.number_format = '#,##0.00 €'
        for col in ws1.columns:
            mx = max(len(str(c.value or '')) for c in col)
            ws1.column_dimensions[col[0].column_letter].width = min(mx + 4, 40)

        # Sheet 2: Assegnazioni
        ws2 = wb.create_sheet('Assegnazioni')
        h2 = ['#', 'Cognome', 'Nome', 'Camera Assegnata', 'Tipo Richiesto']
        for c, h in enumerate(h2, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = hfont
            cell.fill = hfill2
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        row = 2
        for i, stanza in enumerate(stanze, 1):
            for g in stanza:
                ws2.cell(row=row, column=1, value=i).border = border
                ws2.cell(row=row, column=2, value=g.cognome).border = border
                ws2.cell(row=row, column=3, value=g.nome).border = border
                ws2.cell(row=row, column=4, value=g.camera_assegnata or '').border = border
                ws2.cell(row=row, column=5, value=g.tipo_camera or '').border = border
                row += 1
        for col in ws2.columns:
            mx = max(len(str(c.value or '')) for c in col)
            ws2.column_dimensions[col[0].column_letter].width = min(mx + 4, 40)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'camere_notte_{notte}_ott.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ── EMAIL PARSING (LLM) ─────────────────────────────────────────────────

    @app.post('/api/parse-email')
    def parse_email():
        import anthropic

        data = request.get_json()
        text = (data.get('text') or '').strip()
        if not text:
            return jsonify(ok=False, error='Testo vuoto'), 400

        # Raccogli ospiti esistenti per contesto
        guests = Guest.query.filter_by(deleted=False).order_by(Guest.cognome).all()
        guest_list = '\n'.join(
            f'- [id={g.id}] {g.cognome} {g.nome} (camera: {g.tipo_camera or "n/a"}, '
            f'arrivo: {g.volo_arrivo or "n/a"}, partenza: {g.volo_partenza or "n/a"}, '
            f'email: {g.email or "n/a"})'
            for g in guests
        ) or '(nessun ospite ancora registrato)'

        system_prompt = f"""Sei un assistente che estrae dati di rooming e voli da email/messaggi.

Ospiti attualmente in lista:
{guest_list}

Estrai TUTTE le informazioni su ospiti menzionati nel testo. Per ogni persona, determina:
- cognome (MAIUSCOLO)
- nome
- email, telefono
- sede_lavoro (città/sede di lavoro)
- volo_arrivo (codice volo + orario), volo_partenza (codice volo + orario)
- aeroporto_partenza, aeroporto_arrivo
- pickup_bus_andata, pickup_bus_ritorno (orario pickup bus)
- tipo_camera (singola, doppia, twin, suite, etc.)
- presenza_8, presenza_9, presenza_10, presenza_11 (true/false, giorni 8-11 ottobre)
- parcheggio_linate, parcheggio_hotel (true/false)
- divide_stanza_con (con chi condivide la stanza)
- restrizioni_alimentari
- azione: "update" se la persona esiste già in lista, "add" se è nuova

Se un campo non è menzionato nel testo, usa null.
Se la persona esiste già, includi SOLO i campi che vanno aggiornati (gli altri null).

Rispondi SOLO con JSON valido, niente markdown:
{{
  "guests": [
    {{
      "cognome": "ROSSI",
      "nome": "Mario",
      "email": null,
      "telefono": null,
      "sede_lavoro": null,
      "volo_arrivo": "AZ1234 08:25",
      "volo_partenza": null,
      "tipo_camera": "doppia",
      "presenza_8": true,
      "presenza_9": true,
      "presenza_10": null,
      "presenza_11": null,
      "parcheggio_linate": null,
      "parcheggio_hotel": null,
      "divide_stanza_con": null,
      "restrizioni_alimentari": null,
      "azione": "add",
      "match_id": null,
      "nota": "breve spiegazione di cosa hai interpretato"
    }}
  ],
  "summary": "riassunto di cosa dice l'email"
}}

Per "azione": "update", valorizza "match_id" con l'ID dell'ospite corrispondente.
"""

        api_key = os.environ.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify(ok=False, error='ANTHROPIC_API_KEY non configurata'), 500

        client = anthropic.Anthropic(api_key=api_key)

        try:
            response = client.messages.create(
                model='claude-haiku-4-5-20251001',
                max_tokens=8192,
                system=system_prompt,
                messages=[{'role': 'user', 'content': text}],
            )
            raw = response.content[0].text.strip()
            if raw.startswith('```'):
                raw = raw.split('\n', 1)[1] if '\n' in raw else raw[3:]
                if raw.endswith('```'):
                    raw = raw[:-3]
                raw = raw.strip()

            parsed = json.loads(raw)

            # Costo (Haiku 4.5: $0.80/MTok in, $4.00/MTok out)
            inp = response.usage.input_tokens
            out = response.usage.output_tokens
            cost = (inp * 0.80 + out * 4.00) / 1_000_000

            # Se match_id non fornito dal LLM, prova fuzzy match
            for g in parsed.get('guests', []):
                if g.get('azione') == 'update' and not g.get('match_id'):
                    cognome_up = (g.get('cognome') or '').strip().upper()
                    nome_up = (g.get('nome') or '').strip().upper()
                    # Match case-insensitive su cognome + nome
                    match = Guest.query.filter(
                        Guest.deleted==False,
                        db.func.upper(Guest.cognome) == cognome_up,
                        db.func.upper(Guest.nome) == nome_up
                    ).first() if cognome_up and nome_up else None
                    # Fallback: solo cognome se nome non disponibile
                    if not match and cognome_up:
                        match = Guest.query.filter(
                            Guest.deleted==False,
                            db.func.upper(Guest.cognome) == cognome_up
                        ).first()
                    if match:
                        g['match_id'] = match.id
                        g['match_nome'] = match.nome_completo

            # Per ogni update, includi i dati attuali per confronto
            compare_fields = ('cognome', 'nome', 'email', 'telefono', 'sede_lavoro',
                              'volo_arrivo', 'volo_partenza',
                              'aeroporto_partenza', 'aeroporto_arrivo',
                              'pickup_bus_andata', 'pickup_bus_ritorno',
                              'divide_stanza_con', 'restrizioni_alimentari',
                              'tipo_camera', 'note_form',
                              'presenza_8', 'presenza_9', 'presenza_10', 'presenza_11',
                              'parcheggio_linate', 'parcheggio_hotel')
            for g in parsed.get('guests', []):
                if g.get('azione') == 'update' and g.get('match_id'):
                    existing = Guest.query.get(g['match_id'])
                    if existing:
                        g['current_data'] = {f: getattr(existing, f) for f in compare_fields}

            # Salva il messaggio originale nel log (anche se non verrà applicato)
            email_log = EmailLog(testo=text, summary=parsed.get('summary'), log_type='rooming')
            db.session.add(email_log)
            db.session.commit()

            return jsonify(ok=True, parsed=parsed, email_log_id=email_log.id,
                           usage={'input': inp, 'output': out, 'cost_eur': round(cost * 0.92, 4)})

        except json.JSONDecodeError:
            return jsonify(ok=False, error=f'Risposta LLM non valida: {raw[:300]}'), 500
        except Exception as e:
            return jsonify(ok=False, error=str(e)), 500

    @app.post('/api/apply-parsed')
    def apply_parsed():
        """Applica le azioni estratte dal parsing email."""
        data = request.get_json()
        guests_data = data.get('guests', [])
        email_log_id = data.get('email_log_id')

        results = []
        for gd in guests_data:
            azione = gd.get('azione', 'add')
            cognome = (gd.get('cognome') or '').strip()
            nome = (gd.get('nome') or '').strip()

            str_fields = ('email', 'telefono', 'sede_lavoro',
                          'volo_arrivo', 'volo_partenza',
                          'aeroporto_partenza', 'aeroporto_arrivo',
                          'pickup_bus_andata', 'pickup_bus_ritorno',
                          'divide_stanza_con', 'restrizioni_alimentari',
                          'tipo_camera', 'note_form')
            bool_fields = ('presenza_8', 'presenza_9', 'presenza_10', 'presenza_11',
                           'parcheggio_linate', 'parcheggio_hotel')

            if azione == 'update' and gd.get('match_id'):
                g = Guest.query.get(gd['match_id'])
                if not g:
                    results.append({'cognome': cognome, 'ok': False, 'error': 'Non trovato'})
                    continue
                # Non sovrascrivere cognome/nome — il DB ha il dato canonico
                for f in str_fields:
                    if gd.get(f) is not None:
                        setattr(g, f, gd[f])
                for f in bool_fields:
                    if gd.get(f) is not None:
                        setattr(g, f, _parse_bool(gd[f]))
                g.updated_at = datetime.utcnow()
                if email_log_id:
                    g.email_log_id = email_log_id
                results.append({'cognome': cognome, 'ok': True, 'action': 'updated', 'id': g.id})
            else:
                # Controllo duplicati case-insensitive prima di inserire
                existing = None
                if cognome:
                    q = Guest.query.filter(
                        Guest.deleted==False,
                        db.func.upper(Guest.cognome) == cognome.upper()
                    )
                    if nome:
                        q = q.filter(db.func.upper(Guest.nome) == nome.upper())
                    existing = q.first()
                if existing:
                    # Esiste già: aggiorna invece di duplicare
                    for f in str_fields:
                        if gd.get(f) is not None:
                            setattr(existing, f, gd[f])
                    for f in bool_fields:
                        if gd.get(f) is not None:
                            setattr(existing, f, _parse_bool(gd[f]))
                    existing.updated_at = datetime.utcnow()
                    if email_log_id:
                        existing.email_log_id = email_log_id
                    results.append({'cognome': cognome, 'ok': True,
                                    'action': 'updated (dedup)', 'id': existing.id})
                else:
                    kwargs = dict(cognome=cognome, nome=nome, source='email',
                                  note=gd.get('nota'),
                                  email_log_id=email_log_id)
                    for f in str_fields:
                        kwargs[f] = gd.get(f)
                    for f in bool_fields:
                        kwargs[f] = _parse_bool(gd.get(f))
                    g = Guest(**kwargs)
                    db.session.add(g)
                    db.session.flush()
                    results.append({'cognome': cognome, 'ok': True, 'action': 'added', 'id': g.id})

        db.session.commit()
        added_count = sum(1 for r in results if r.get('ok') and r.get('action') == 'added')
        updated_count = sum(1 for r in results if r.get('ok') and 'updated' in (r.get('action') or ''))
        log_audit('rooming', 'Guest', None, 'import',
                  changes={'added': added_count, 'updated': updated_count},
                  summary=f'Email import: {added_count} aggiunti, {updated_count} aggiornati')
        return jsonify(ok=True, results=results)

    # ── EMAIL LOG ─────────────────────────────────────────────────────────

    @app.get('/api/email-logs')
    def list_email_logs():
        logs = EmailLog.query.filter(
            (EmailLog.log_type == 'rooming') | (EmailLog.log_type.is_(None))
        ).order_by(EmailLog.created_at.desc()).all()
        return jsonify([{
            'id': l.id,
            'summary': l.summary,
            'testo': l.testo,
            'created_at': l.created_at.isoformat(),
            'guests': [{'id': g.id, 'nome_completo': g.nome_completo}
                        for g in Guest.query.filter_by(deleted=False, email_log_id=l.id).all()]
        } for l in logs])

    @app.get('/api/partivia/email-logs')
    def list_partivia_email_logs():
        logs = EmailLog.query.filter_by(log_type='partivia').order_by(EmailLog.created_at.desc()).all()
        return jsonify([{
            'id': l.id,
            'summary': l.summary,
            'testo': l.testo,
            'created_at': l.created_at.isoformat(),
            'quotes': [{'id': q.id, 'hotel_name': q.hotel_name}
                        for q in PartiviaQuote.query.filter_by(deleted=False, email_log_id=l.id).all()]
        } for l in logs])

    @app.get('/api/email-log/<int:log_id>')
    def get_email_log(log_id):
        log = EmailLog.query.get_or_404(log_id)
        return jsonify(id=log.id, testo=log.testo, summary=log.summary,
                       created_at=log.created_at.isoformat())

    # ── IMPORT DATE NASCITA DA JSON ─────────────────────────────────────────

    @app.post('/api/import/birth-dates')
    def import_birth_dates():
        """Import birth dates from allegati JSON (key: 'Nome Cognome/file.jpg' → 'DD/MM/YYYY')."""
        import unicodedata

        if 'file' not in request.files:
            return jsonify(ok=False, error='File JSON mancante'), 400

        try:
            data = json.load(request.files['file'])
        except Exception as e:
            return jsonify(ok=False, error=f'JSON non valido: {e}'), 400

        # Extract unique name → birth_date
        birth_dates = {}
        for key, date_val in data.items():
            name = key.split('/')[0].strip()
            if name not in birth_dates and date_val:
                birth_dates[name] = date_val

        def _norm(name):
            s = name.lower().strip().replace('\u2019', "'").replace('\u2018', "'")
            return ''.join(
                c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn'
            )

        guests = Guest.query.filter_by(deleted=False).all()
        lookup = {}
        for g in guests:
            lookup[_norm(f"{g.cognome} {g.nome}")] = g
            rev = _norm(f"{g.nome} {g.cognome}")
            if rev not in lookup:
                lookup[rev] = g

        updated = 0
        not_found = []
        for name, bdate in birth_dates.items():
            guest = lookup.get(_norm(name))
            if guest:
                guest.data_nascita = bdate
                updated += 1
            else:
                not_found.append(name)

        db.session.commit()
        return jsonify(ok=True, updated=updated, total=len(birth_dates),
                       not_found=not_found)

    # ── DELETE ALL ───────────────────────────────────────────────────────────

    @app.delete('/api/guests')
    def delete_all():
        Guest.query.filter_by(deleted=False).update({'deleted': True, 'deleted_at': datetime.utcnow()})
        db.session.commit()
        log_audit('rooming', 'Guest', None, 'delete',
                  summary='Eliminati tutti gli ospiti')
        return jsonify(ok=True)

    @app.delete('/api/partivia/quotes-all')
    def delete_all_partivia():
        """Delete all Partivia quotes and their partivia email logs."""
        PartiviaQuote.query.filter_by(deleted=False).update({'deleted': True, 'deleted_at': datetime.utcnow()})
        EmailLog.query.filter_by(log_type='partivia').delete()
        db.session.commit()
        log_audit('partivia', 'PartiviaQuote', None, 'delete',
                  summary='Eliminati tutti i preventivi')
        return jsonify(ok=True)

    # ══════════════════════════════════════════════════════════════════════════
    # ██  PARTIVIA — Preventivi Hotel                                       ██
    # ══════════════════════════════════════════════════════════════════════════

    @app.route('/partivia/client')
    def partivia_client():
        return partivia(client_view=True)

    @app.route('/partivia')
    def partivia(client_view=False):
        import re

        quotes = (PartiviaQuote.query.filter_by(deleted=False)
                  .order_by(PartiviaQuote.city, PartiviaQuote.hotel_name)
                  .all())

        # ── Normalizzazione tipo camera per pivot ──
        def normalize_room_type(rt):
            rt_l = rt.lower().strip()
            if 'suite' in rt_l and 'junior' not in rt_l:
                return 'Suite'
            if 'junior' in rt_l:
                return 'Junior Suite'
            if 'superior' in rt_l:
                return 'Superior'
            if 'deluxe' in rt_l:
                return 'Deluxe'
            if 'triple' in rt_l or 'tripl' in rt_l:
                return 'Tripla'
            # "Double for single use" / "single occupancy" → Singola
            if any(k in rt_l for k in ('single use', 'single occupancy',
                                        'uso singol', 'for single',
                                        'dui', 'individual use',
                                        'dus')):
                return 'Singola'
            # Pure single
            if any(k in rt_l for k in ('singol', 'single')):
                return 'Singola'
            if any(k in rt_l for k in ('doppi', 'double', 'twin',
                                        'dbl', 'double use',
                                        'double occupancy')):
                return 'Doppia/Twin'
            # Run of House / ROH → treat as Doppia/Twin (most common)
            if 'run of' in rt_l or 'roh' in rt_l:
                return 'Doppia/Twin'
            # Premium/standard mix → keep as-is
            return rt.strip()

        # ── Normalizzazione tipo pasto per pivot ──
        def normalize_meal_type(mt):
            mt_l = mt.lower().strip()
            if 'coffee' in mt_l or 'break' in mt_l:
                return 'Coffee Break'
            if 'cocktail' in mt_l or 'welcome' in mt_l or 'aperitivo' in mt_l:
                return 'Cocktail'
            if 'gala' in mt_l:
                return 'Gala Dinner'
            if 'cena' in mt_l or 'dinner' in mt_l:
                return 'Cena'
            if 'pranzo' in mt_l or 'lunch' in mt_l or 'buffet' in mt_l:
                return 'Pranzo'
            if 'colazione' in mt_l or 'breakfast' in mt_l:
                return 'Colazione'
            if 'ddr' in mt_l or 'delegate' in mt_l:
                return 'DDR'
            return mt.strip()

        # ── Estrai numero da stringa prezzo ──
        def parse_price(s):
            if not s:
                return None
            m = re.search(r'[\d.,]+', s.replace('.', '').replace(',', '.'))
            return float(m.group()) if m else None

        # ── Dati aggregati per le tab ──
        ROOM_COLS = ['Singola', 'Doppia/Twin', 'Superior', 'Deluxe',
                     'Junior Suite', 'Suite']
        MEAL_COLS = ['Colazione', 'Coffee Break', 'Pranzo', 'Cena',
                     'Cocktail', 'Gala Dinner', 'DDR']

        # Set of hidden hotel keys for filtering pivots
        hidden_hotel_keys = {q.hotel_name.lower().strip()
                             for q in quotes if q.hidden}

        room_pivot = []  # lista di dict per ogni quote
        fb_pivot = []
        for q in quotes:
            if q.hotel_name.lower().strip() in hidden_hotel_keys:
                continue
            # Room pivot
            rates_map = {}
            for rr in q.room_rates:
                norm = normalize_room_type(rr.room_type)
                if norm not in rates_map:
                    rates_map[norm] = rr.rate_per_night or ''
            room_pivot.append({
                'id': q.id,
                'hotel': q.hotel_name,
                'city': q.city,
                'stars': q.stars,
                'rooms': q.rooms_available or '',
                'rates': {col: rates_map.get(col, '') for col in ROOM_COLS},
                'price_val': parse_price(rates_map.get('Doppia/Twin')
                                         or rates_map.get('Singola')),
            })

            # F&B pivot
            fb_map = {}
            for fb in q.fb_options:
                norm = normalize_meal_type(fb.meal_type)
                if norm not in fb_map:
                    fb_map[norm] = fb.price_per_person or ''
            fb_pivot.append({
                'id': q.id,
                'hotel': q.hotel_name,
                'city': q.city,
                'meals': {col: fb_map.get(col, '') for col in MEAL_COLS},
            })

        # ── Raggruppa per hotel (per Overview) ──
        hotels_grouped = {}  # key = hotel_name_lower → list of quotes
        for q in quotes:
            key = q.hotel_name.lower().strip()
            hotels_grouped.setdefault(key, []).append(q)

        # Per ogni gruppo, scegli il "best" (più dati) e tieni le versioni
        all_hotels = []  # lista di dict con best + versions (include hidden)
        for key, group in hotels_grouped.items():
            # Ordina per completezza: più room_rates + meeting_rooms + fb_options
            scored = sorted(group, key=lambda q: (
                len(q.room_rates) + len(q.meeting_rooms) + len(q.fb_options)
            ), reverse=True)
            best = scored[0]
            all_hotels.append({
                'best': best,
                'versions': group,
                'count': len(group),
                'hidden': best.hidden or False,
            })

        # Ordina hotels per città + nome
        all_hotels.sort(key=lambda h: (h['best'].city, h['best'].hotel_name))

        # Visible hotels (non-hidden) for display
        hotels = [h for h in all_hotels if not h['hidden']]

        # Stats (only visible quotes)
        visible_quotes = [q for q in quotes
                          if q.hotel_name.lower().strip() not in hidden_hotel_keys]
        cities = {}
        stars_count = {}
        status_count = {}
        for q in visible_quotes:
            cities[q.city] = cities.get(q.city, 0) + 1
            s = q.stars or 0
            stars_count[s] = stars_count.get(s, 0) + 1
            status_count[q.quote_status] = status_count.get(q.quote_status, 0) + 1

        return render_template('partivia.html',
                               quotes=visible_quotes,
                               hotels=hotels,
                               all_hotels=all_hotels,
                               room_pivot=room_pivot,
                               room_cols=ROOM_COLS,
                               fb_pivot=fb_pivot,
                               meal_cols=MEAL_COLS,
                               stats_cities=cities,
                               stats_stars=stars_count,
                               stats_status=status_count,
                               client_view=client_view)

    # ── Budget overrides ─────────────────────────────────────────────────

    @app.get('/api/partivia/budget-overrides')
    def get_budget_overrides():
        row = BudgetOverride.query.first()
        return jsonify(row.data if row else {})

    @app.post('/api/partivia/budget-overrides')
    def save_budget_overrides():
        data = request.get_json()
        row = BudgetOverride.query.first()
        if not row:
            row = BudgetOverride(data=data)
            db.session.add(row)
        else:
            row.data = data
        db.session.commit()
        return jsonify(ok=True)

    # ── Parse email preventivo ────────────────────────────────────────────

    @app.post('/api/partivia/parse-email')
    def partivia_parse_email():
        import anthropic

        data = request.get_json()
        text = (data.get('text') or '').strip()
        if not text:
            return jsonify(ok=False, error='Testo vuoto'), 400

        # Contesto: preventivi già in DB
        existing = PartiviaQuote.query.filter_by(deleted=False).order_by(PartiviaQuote.city).all()
        existing_list = '\n'.join(
            f'- [id={q.id}] {q.hotel_name} ({q.city}, {q.stars or "?"}★) '
            f'— stato: {q.quote_status}, date: {q.dates_proposed or "n/a"}'
            for q in existing
        ) or '(nessun preventivo ancora registrato)'

        system_prompt = f"""You are an assistant that extracts hotel quote data from emails.
The event is "N!Partivia" — a corporate incentive trip in Spain.
Possible destinations: Barcellona, Madrid, Siviglia, Valencia.

Quotes already registered:
{existing_list}

Analyze the email and extract ALL hotel quotes/proposals present.
For each quote, extract:
- hotel_name (hotel name)
- city (Barcellona, Madrid, Siviglia or Valencia — always normalize to Italian spelling)
- stars (integer 1-5 or null)
- contact_name, contact_email (hotel contact)
- website_url (hotel website URL if mentioned, or null)
- address (hotel street address if mentioned, or null)
- dates_proposed (proposed dates, e.g. "10-13 October 2026" — MANDATORY, always extract available dates/periods mentioned in the email)
- rooms_available (available rooms)
- min_rooms_required (minimum rooms required)
- room_rates: list of objects with room_type, rate_per_night (with €, MANDATORY — always extract the nightly rate even if you need to calculate it from a total or package price), breakfast_included (yes/no/not specified), notes (in English, about room specifics only). NEVER leave rate_per_night empty or null — if the email mentions any price for rooms, extract it. If a total/package price is given instead of per-night, divide and note "calculated from total" in notes.
- meeting_rooms: list with name, capacity, rate, notes (in English — technical details: AV equipment, layout, natural light, etc.)
- fb_options: list with meal_type (Breakfast/Lunch/Dinner/Coffee Break/Gala Dinner/DDR), price_per_person, menu_description
- cancellation_policy, payment_terms, validity_date, commission
- total_estimate (total estimate if present)
- included_services (list of included services like WiFi, parking, etc.)
- notes (in English — only about rooms and meeting rooms, not general conditions)
- raw_summary (2-3 sentence summary in English — MUST always include room rates/costs per night, e.g. "Double rooms at €180/night". Room pricing is the most important information in the summary.)
- quote_status: one of "pending_review", "negotiating", "confirmed", "declined", "expired". Determine from the email tone:
  * "confirmed" if the hotel confirms availability/booking
  * "declined" if the hotel declines or says dates are not available
  * "expired" if the option deadline has passed or quote is no longer valid
  * "negotiating" if the hotel is providing a quote/proposal with rates
  * "pending_review" only if unclear
- is_update: true if updating an existing quote (with match_id), false if new
- match_id: ID of existing quote if updating, null if new

IMPORTANT: ALL text fields MUST be in English. No exceptions. This includes:
- raw_summary, notes, room_rates[].notes, meeting_rooms[].notes, fb_options[].menu_description
- cancellation_policy, payment_terms, included_services
- total_estimate (e.g. "Not provided" instead of "Non fornito")
- rooms_available (e.g. "50 DSU (to be confirmed)" instead of "50 DSU (da confermare)")
- rate_per_night notes and descriptions
- breakfast_included (use "Yes"/"No"/"Not specified" only)
Always translate EVERYTHING from Spanish, Italian, or any other language to English. Never leave any field in its original language, even partially. For example: "350€ supplement per night per room" NOT "350€ supplemento per notte per camera".

CRITICAL: Room costs (rate_per_night) and dates_proposed are the MOST important data to extract.
- Every room_rates entry MUST have a rate_per_night value with € symbol. If the email quotes room prices in ANY format (per night, per stay, per person, package), convert to per-night rate and include it.
- dates_proposed MUST always be filled if any dates or periods are mentioned in the email (check-in/check-out, event dates, availability windows).
- The raw_summary MUST always mention the room rates (e.g. "rooms from €X to €Y per night") and the proposed dates.

If the message does NOT contain quotes (e.g. simple follow-up), set is_quote=false.

Reply ONLY with valid JSON (no markdown):
{{
  "quotes": [
    {{
      "hotel_name": "Hotel Example",
      "city": "Barcellona",
      "stars": 4,
      "contact_name": "Mario Rossi",
      "contact_email": "mario@hotel.com",
      "website_url": "https://www.hotelexample.com",
      "dates_proposed": "10-13 October 2026",
      "rooms_available": "80",
      "min_rooms_required": null,
      "room_rates": [
        {{"room_type": "Double", "rate_per_night": "€ 180", "breakfast_included": "yes", "notes": "Sea view upgrade available"}}
      ],
      "meeting_rooms": [
        {{"name": "Grand Hall", "capacity": "200 pax theatre", "rate": "€ 2,000/day", "notes": "AV included, natural daylight, 250sqm"}}
      ],
      "fb_options": [
        {{"meal_type": "Dinner", "price_per_person": "€ 55/pax", "menu_description": "3-course menu"}}
      ],
      "cancellation_policy": "Free cancellation up to 30 days",
      "payment_terms": "30% upon confirmation",
      "validity_date": "30/09/2026",
      "commission": "10%",
      "total_estimate": "€ 45,000",
      "included_services": ["WiFi", "Parking", "Gym"],
      "notes": "Room upgrade available on request",
      "raw_summary": "Hotel Example offers 80 double rooms at €180/night...",
      "is_update": false,
      "match_id": null
    }}
  ],
  "is_quote": true,
  "message_type": "quote",
  "summary": "Received quote from Hotel Example for Barcellona..."
}}"""

        api_key = os.environ.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify(ok=False, error='ANTHROPIC_API_KEY non configurata'), 500

        client = anthropic.Anthropic(api_key=api_key)

        try:
            response = client.messages.create(
                model='claude-haiku-4-5-20251001',
                max_tokens=4096,
                system=system_prompt,
                messages=[{'role': 'user', 'content': text}],
            )
            raw = response.content[0].text.strip()
            if raw.startswith('```'):
                raw = raw.split('\n', 1)[1] if '\n' in raw else raw[3:]
                if raw.endswith('```'):
                    raw = raw[:-3]
                raw = raw.strip()

            parsed = json.loads(raw)

            # Costo (Haiku 4.5)
            inp = response.usage.input_tokens
            out = response.usage.output_tokens
            cost = (inp * 0.80 + out * 4.00) / 1_000_000

            # Salva log
            email_log = EmailLog(testo=text, summary=parsed.get('summary'), log_type='partivia')
            db.session.add(email_log)
            db.session.commit()

            return jsonify(ok=True, parsed=parsed, email_log_id=email_log.id,
                           usage={'input': inp, 'output': out,
                                  'cost_eur': round(cost * 0.92, 4)})

        except json.JSONDecodeError:
            return jsonify(ok=False, error=f'Risposta LLM non valida: {raw[:300]}'), 500
        except Exception as e:
            return jsonify(ok=False, error=str(e)), 500

    # ── Applica preventivi estratti ───────────────────────────────────────

    @app.post('/api/partivia/apply')
    def partivia_apply():
        data = request.get_json()
        quotes_data = data.get('quotes', [])
        email_log_id = data.get('email_log_id')

        results = []
        for qd in quotes_data:
            is_update = qd.get('is_update', False)
            match_id = qd.get('match_id')

            if is_update and match_id:
                q = PartiviaQuote.query.get(match_id)
                if not q:
                    results.append({'hotel': qd.get('hotel_name'),
                                    'ok': False, 'error': 'Non trovato'})
                    continue
                # Aggiorna campi top-level
                for field in ('hotel_name', 'city', 'stars', 'contact_name',
                              'contact_email', 'dates_proposed', 'rooms_available',
                              'min_rooms_required', 'cancellation_policy',
                              'payment_terms', 'validity_date', 'commission',
                              'total_estimate', 'notes', 'raw_summary',
                              'website_url', 'address', 'quote_status'):
                    if qd.get(field) is not None:
                        setattr(q, field, qd[field])
                if qd.get('included_services'):
                    q.included_services = ', '.join(qd['included_services'])
                q.updated_at = datetime.utcnow()
                if email_log_id:
                    q.email_log_id = email_log_id

                # Sostituisci sotto-tabelle se fornite
                if qd.get('room_rates'):
                    PartiviaRoomRate.query.filter_by(quote_id=q.id).delete()
                    for rr in qd['room_rates']:
                        db.session.add(PartiviaRoomRate(
                            quote_id=q.id, room_type=rr.get('room_type', ''),
                            rate_per_night=rr.get('rate_per_night'),
                            breakfast_included=rr.get('breakfast_included'),
                            notes=rr.get('notes')))
                if qd.get('meeting_rooms'):
                    PartiviaMeetingRoom.query.filter_by(quote_id=q.id).delete()
                    for mr in qd['meeting_rooms']:
                        db.session.add(PartiviaMeetingRoom(
                            quote_id=q.id, name=mr.get('name', ''),
                            capacity=mr.get('capacity'),
                            rate=mr.get('rate'), notes=mr.get('notes')))
                if qd.get('fb_options'):
                    PartiviaFBOption.query.filter_by(quote_id=q.id).delete()
                    for fb in qd['fb_options']:
                        db.session.add(PartiviaFBOption(
                            quote_id=q.id, meal_type=fb.get('meal_type', ''),
                            price_per_person=fb.get('price_per_person'),
                            menu_description=fb.get('menu_description')))

                db.session.flush()
                results.append({'hotel': q.hotel_name, 'ok': True,
                                'action': 'updated', 'id': q.id})
            else:
                # Controllo duplicati case-insensitive per hotel_name + city
                hotel_name_raw = (qd.get('hotel_name') or '').strip()
                city_raw = (qd.get('city') or '').strip()
                existing_q = None
                if hotel_name_raw:
                    eq = PartiviaQuote.query.filter(
                        PartiviaQuote.deleted==False,
                        db.func.upper(PartiviaQuote.hotel_name) == hotel_name_raw.upper()
                    )
                    if city_raw:
                        eq = eq.filter(
                            db.func.upper(PartiviaQuote.city) == city_raw.upper()
                        )
                    existing_q = eq.first()
                if existing_q:
                    # Esiste già: aggiorna invece di duplicare
                    for fld in ('contact_name', 'contact_email', 'dates_proposed',
                                'rooms_available', 'min_rooms_required',
                                'cancellation_policy', 'payment_terms',
                                'validity_date', 'commission', 'total_estimate',
                                'notes', 'raw_summary', 'website_url', 'address'):
                        if qd.get(fld) is not None:
                            setattr(existing_q, fld, qd[fld])
                    if qd.get('stars') is not None:
                        existing_q.stars = qd['stars']
                    if qd.get('included_services'):
                        existing_q.included_services = ', '.join(qd['included_services'])
                    existing_q.updated_at = datetime.utcnow()
                    if email_log_id:
                        existing_q.email_log_id = email_log_id
                    # Sostituisci sotto-tabelle se fornite
                    if qd.get('room_rates'):
                        PartiviaRoomRate.query.filter_by(quote_id=existing_q.id).delete()
                        for rr in qd['room_rates']:
                            db.session.add(PartiviaRoomRate(
                                quote_id=existing_q.id, room_type=rr.get('room_type', ''),
                                rate_per_night=rr.get('rate_per_night'),
                                breakfast_included=rr.get('breakfast_included'),
                                notes=rr.get('notes')))
                    if qd.get('meeting_rooms'):
                        PartiviaMeetingRoom.query.filter_by(quote_id=existing_q.id).delete()
                        for mr in qd['meeting_rooms']:
                            db.session.add(PartiviaMeetingRoom(
                                quote_id=existing_q.id, name=mr.get('name', ''),
                                capacity=mr.get('capacity'),
                                rate=mr.get('rate'), notes=mr.get('notes')))
                    if qd.get('fb_options'):
                        PartiviaFBOption.query.filter_by(quote_id=existing_q.id).delete()
                        for fb in qd['fb_options']:
                            db.session.add(PartiviaFBOption(
                                quote_id=existing_q.id, meal_type=fb.get('meal_type', ''),
                                price_per_person=fb.get('price_per_person'),
                                menu_description=fb.get('menu_description')))
                    db.session.flush()
                    results.append({'hotel': existing_q.hotel_name, 'ok': True,
                                    'action': 'updated (dedup)', 'id': existing_q.id})
                    continue

                # Nuovo preventivo
                q = PartiviaQuote(
                    hotel_name=hotel_name_raw,
                    city=city_raw,
                    stars=qd.get('stars'),
                    contact_name=qd.get('contact_name'),
                    contact_email=qd.get('contact_email'),
                    dates_proposed=qd.get('dates_proposed'),
                    rooms_available=qd.get('rooms_available'),
                    min_rooms_required=qd.get('min_rooms_required'),
                    cancellation_policy=qd.get('cancellation_policy'),
                    payment_terms=qd.get('payment_terms'),
                    validity_date=qd.get('validity_date'),
                    commission=qd.get('commission'),
                    total_estimate=qd.get('total_estimate'),
                    included_services=', '.join(qd.get('included_services', [])),
                    notes=qd.get('notes'),
                    raw_summary=qd.get('raw_summary'),
                    website_url=qd.get('website_url'),
                    address=qd.get('address'),
                    quote_status=qd.get('quote_status', 'pending_review'),
                    source='email',
                    email_log_id=email_log_id,
                )
                db.session.add(q)
                db.session.flush()

                for rr in qd.get('room_rates', []):
                    db.session.add(PartiviaRoomRate(
                        quote_id=q.id, room_type=rr.get('room_type', ''),
                        rate_per_night=rr.get('rate_per_night'),
                        breakfast_included=rr.get('breakfast_included'),
                        notes=rr.get('notes')))
                for mr in qd.get('meeting_rooms', []):
                    db.session.add(PartiviaMeetingRoom(
                        quote_id=q.id, name=mr.get('name', ''),
                        capacity=mr.get('capacity'),
                        rate=mr.get('rate'), notes=mr.get('notes')))
                for fb in qd.get('fb_options', []):
                    db.session.add(PartiviaFBOption(
                        quote_id=q.id, meal_type=fb.get('meal_type', ''),
                        price_per_person=fb.get('price_per_person'),
                        menu_description=fb.get('menu_description')))

                db.session.flush()
                results.append({'hotel': q.hotel_name, 'ok': True,
                                'action': 'added', 'id': q.id})

        db.session.commit()
        count = sum(1 for r in results if r.get('ok'))
        log_audit('partivia', 'PartiviaQuote', None, 'import',
                  changes={'count': count},
                  summary=f'Applicati {count} preventivi da email')
        return jsonify(ok=True, results=results)

    # ── Re-parse all quotes from original emails ────────────────────────

    @app.post('/api/partivia/reparse-all')
    @app.post('/api/partivia/extract-missing')
    def partivia_extract_missing():
        """Re-analyze saved emails to extract dinner prices and VAT info."""
        import anthropic

        client = anthropic.Anthropic()
        results = []

        # Get all quotes that have an email_log
        quotes = PartiviaQuote.query.filter(
            PartiviaQuote.deleted==False, PartiviaQuote.email_log_id.isnot(None)
        ).order_by(PartiviaQuote.city, PartiviaQuote.hotel_name).all()

        for q in quotes:
            log = EmailLog.query.get(q.email_log_id)
            if not log:
                results.append({'hotel': q.hotel_name, 'status': 'no email log'})
                continue

            # Check what's missing
            has_dinner = any(
                'dinner' in fb.meal_type.lower() or 'cena' in fb.meal_type.lower() or 'gala' in fb.meal_type.lower()
                for fb in q.fb_options
            )
            has_vat = q.vat_included is not None

            if has_dinner and has_vat:
                results.append({'hotel': q.hotel_name, 'status': 'already complete'})
                continue

            prompt = f"""Analyze this hotel quote email and extract ONLY the following information:

1. DINNER: Is there any dinner option offered? If yes, what is the price per person?
   Look for: dinner, cena, gala dinner, cocktail dinner, evening meal, supper.
   Also check any attached menus or F&B proposals.

2. VAT: Are the ROOM rates VAT included or excluded?
   Look for: VAT, IVA, tax included, tax excluded, +VAT, +IVA, impuestos.
   Check room rate notes, general conditions, and fine print.

Reply ONLY with valid JSON (no markdown):
{{
  "dinner_options": [
    {{"meal_type": "Dinner", "price_per_person": "€ 65", "description": "3-course menu"}},
  ],
  "vat_included": "yes" or "no" or "unknown"
}}

If no dinner is offered at all, return empty list for dinner_options.
If VAT status cannot be determined, use "unknown".

HOTEL: {q.hotel_name} ({q.city})
EMAIL TEXT:
{log.testo[:6000]}"""

            try:
                response = client.messages.create(
                    model='claude-haiku-4-5-20251001',
                    max_tokens=1024,
                    messages=[{'role': 'user', 'content': prompt}],
                )
                raw = response.content[0].text.strip()
                if raw.startswith('```'):
                    raw = raw.split('\n', 1)[1].rsplit('```', 1)[0].strip()

                import json
                parsed = json.loads(raw)

                # Update VAT
                vat = parsed.get('vat_included', 'unknown')
                if vat in ('yes', 'no', 'unknown'):
                    q.vat_included = vat

                # Add dinner options if missing
                dinner_opts = parsed.get('dinner_options', [])
                if not has_dinner and dinner_opts:
                    for d in dinner_opts:
                        fb = PartiviaFBOption(
                            quote_id=q.id,
                            meal_type=d.get('meal_type', 'Dinner'),
                            price_per_person=d.get('price_per_person'),
                            menu_description=d.get('description'),
                        )
                        db.session.add(fb)

                db.session.commit()
                results.append({
                    'hotel': q.hotel_name,
                    'status': 'updated',
                    'vat': vat,
                    'dinners_added': len(dinner_opts) if not has_dinner else 0,
                })

            except Exception as e:
                results.append({'hotel': q.hotel_name, 'status': f'error: {str(e)}'})

        return jsonify(ok=True, results=results)

    @app.post('/api/partivia/reparse-all')
    def partivia_reparse_all():
        """Re-generate raw_summary for ALL quotes using existing DB data.
        Works even without original email text — builds a structured
        description from stored fields and asks the LLM to produce a
        proper summary with room costs and dates."""
        import anthropic
        import time as _time

        api_key = os.environ.get('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify(ok=False, error='ANTHROPIC_API_KEY non configurata'), 500

        # Support batch processing to avoid Railway timeout
        data = request.get_json(silent=True) or {}
        batch_offset = data.get('offset', 0)
        batch_limit = data.get('limit', 5)  # default 5 at a time

        quotes = PartiviaQuote.query.filter_by(deleted=False).order_by(PartiviaQuote.id)\
            .offset(batch_offset).limit(batch_limit).all()
        total_quotes = PartiviaQuote.query.filter_by(deleted=False).count()
        if not quotes:
            return jsonify(ok=False, error='No quotes found (or offset past end)')

        client = anthropic.Anthropic(api_key=api_key)
        results = []
        total_cost = 0.0

        system_prompt = """You are an assistant that processes hotel quote data.
Given structured data about a hotel quote, produce a JSON object with:
- raw_summary: 2-3 sentence summary in English. MUST always include:
  1. Room rates per night (e.g. "Double rooms at €180/night")
  2. Proposed dates if available
  3. Key highlights (capacity, meeting rooms, F&B)
- translated_notes: ALL the text fields below translated to English. Return them as a JSON object.

For each input field that is NOT already in English, translate it. Keep prices, numbers and proper nouns unchanged.

Reply ONLY with valid JSON (no markdown):
{"raw_summary": "...", "room_rates_notes": {"0": "translated note for first room rate", "1": "..."}, "meeting_rooms_notes": {"0": "translated note", ...}, "fb_descriptions": {"0": "translated description", ...}, "cancellation_policy": "...", "payment_terms": "...", "breakfast_labels": {"0": "Yes/No/Not specified", ...}}

Only include fields that needed translation. Omit fields already in English."""

        for q in quotes:
            # Check if this quote has an original email we can re-parse
            has_email = False
            if q.email_log_id:
                email_log = EmailLog.query.get(q.email_log_id)
                if email_log and email_log.testo:
                    has_email = True

            if has_email:
                # Full re-parse from original email
                full_prompt = f"""Extract the hotel quote data from this email. Focus on:
1. room_rates with rate_per_night (MANDATORY, with € symbol)
2. dates_proposed (MANDATORY)
3. raw_summary that includes room costs and dates

The hotel name should be: {q.hotel_name}

Reply ONLY with valid JSON:
{{"hotel_name": "...", "dates_proposed": "...", "room_rates": [{{"room_type": "...", "rate_per_night": "€...", "breakfast_included": "...", "notes": "..."}}], "raw_summary": "...", "meeting_rooms": [...], "fb_options": [...], "cancellation_policy": "...", "payment_terms": "...", "validity_date": "...", "commission": "...", "total_estimate": "...", "included_services": [...], "notes": "..."}}

Email text:
{email_log.testo}"""

                try:
                    response = client.messages.create(
                        model='claude-haiku-4-5-20251001',
                        max_tokens=4096,
                        messages=[{'role': 'user', 'content': full_prompt}],
                    )
                    raw = response.content[0].text.strip()
                    if raw.startswith('```'):
                        raw = raw.split('\n', 1)[1] if '\n' in raw else raw[3:]
                        if raw.endswith('```'):
                            raw = raw[:-3]
                        raw = raw.strip()

                    pq = json.loads(raw)
                    inp = response.usage.input_tokens
                    out = response.usage.output_tokens
                    total_cost += (inp * 0.80 + out * 4.00) / 1_000_000

                    # Update all fields from re-parse
                    for field in ('contact_name', 'contact_email',
                                  'dates_proposed', 'rooms_available',
                                  'min_rooms_required', 'cancellation_policy',
                                  'payment_terms', 'validity_date',
                                  'commission', 'total_estimate', 'notes',
                                  'raw_summary'):
                        val = pq.get(field)
                        if val is not None:
                            setattr(q, field, val)
                    if pq.get('included_services'):
                        svc = pq['included_services']
                        q.included_services = ', '.join(svc) if isinstance(svc, list) else svc

                    if pq.get('room_rates'):
                        PartiviaRoomRate.query.filter_by(quote_id=q.id).delete()
                        for rr in pq['room_rates']:
                            db.session.add(PartiviaRoomRate(
                                quote_id=q.id,
                                room_type=rr.get('room_type', ''),
                                rate_per_night=rr.get('rate_per_night'),
                                breakfast_included=rr.get('breakfast_included'),
                                notes=rr.get('notes')))
                    if pq.get('meeting_rooms'):
                        PartiviaMeetingRoom.query.filter_by(quote_id=q.id).delete()
                        for mr in pq['meeting_rooms']:
                            db.session.add(PartiviaMeetingRoom(
                                quote_id=q.id, name=mr.get('name', ''),
                                capacity=mr.get('capacity'),
                                rate=mr.get('rate'), notes=mr.get('notes')))
                    if pq.get('fb_options'):
                        PartiviaFBOption.query.filter_by(quote_id=q.id).delete()
                        for fb in pq['fb_options']:
                            db.session.add(PartiviaFBOption(
                                quote_id=q.id, meal_type=fb.get('meal_type', ''),
                                price_per_person=fb.get('price_per_person'),
                                menu_description=fb.get('menu_description')))

                    db.session.flush()
                    results.append({
                        'hotel': q.hotel_name, 'ok': True, 'mode': 'email',
                        'rates': [rr.get('rate_per_night', '?') for rr in pq.get('room_rates', [])],
                        'dates': pq.get('dates_proposed'),
                        'summary': (pq.get('raw_summary') or '')[:120],
                    })
                except Exception as e:
                    results.append({'hotel': q.hotel_name, 'ok': False,
                                    'error': str(e)})
            else:
                # No original email — rebuild summary from existing DB data
                room_info = []
                for idx, rr in enumerate(q.room_rates):
                    parts = [f'[{idx}] {rr.room_type}']
                    if rr.rate_per_night:
                        parts.append(f'rate: {rr.rate_per_night}/night')
                    if rr.breakfast_included:
                        parts.append(f'breakfast: {rr.breakfast_included}')
                    if rr.notes:
                        parts.append(f'notes: {rr.notes}')
                    room_info.append(', '.join(parts))

                meeting_info = []
                for idx, mr in enumerate(q.meeting_rooms):
                    parts = [f'[{idx}] {mr.name}']
                    if mr.capacity:
                        parts.append(f'capacity: {mr.capacity}')
                    if mr.rate:
                        parts.append(f'rate: {mr.rate}')
                    if mr.notes:
                        parts.append(f'notes: {mr.notes}')
                    meeting_info.append(', '.join(parts))

                fb_info = []
                for idx, fb in enumerate(q.fb_options):
                    parts = [f'[{idx}] {fb.meal_type}']
                    if fb.price_per_person:
                        parts.append(fb.price_per_person)
                    if fb.menu_description:
                        parts.append(f'desc: {fb.menu_description}')
                    fb_info.append(', '.join(parts))

                user_msg = f"""Hotel: {q.hotel_name}
City: {q.city}
Stars: {q.stars or 'N/A'}
Dates proposed: {q.dates_proposed or 'N/A'}
Rooms available: {q.rooms_available or 'N/A'}
Room rates:
{chr(10).join('  - ' + r for r in room_info) if room_info else '  (none)'}
Meeting rooms:
{chr(10).join('  - ' + m for m in meeting_info) if meeting_info else '  (none)'}
F&B options:
{chr(10).join('  - ' + f for f in fb_info) if fb_info else '  (none)'}
Total estimate: {q.total_estimate or 'N/A'}
Cancellation: {q.cancellation_policy or 'N/A'}
Payment terms: {q.payment_terms or 'N/A'}
Deadline: {q.validity_date or 'N/A'}
Commission: {q.commission or 'N/A'}
Notes: {q.notes or 'N/A'}"""

                try:
                    response = client.messages.create(
                        model='claude-haiku-4-5-20251001',
                        max_tokens=1500,
                        system=system_prompt,
                        messages=[{'role': 'user', 'content': user_msg}],
                    )
                    raw = response.content[0].text.strip()
                    if raw.startswith('```'):
                        raw = raw.split('\n', 1)[1] if '\n' in raw else raw[3:]
                        if raw.endswith('```'):
                            raw = raw[:-3]
                        raw = raw.strip()

                    parsed = json.loads(raw)
                    inp = response.usage.input_tokens
                    out = response.usage.output_tokens
                    total_cost += (inp * 0.80 + out * 4.00) / 1_000_000

                    q.raw_summary = parsed.get('raw_summary', q.raw_summary)

                    # Apply translated notes to room_rates
                    rr_notes = parsed.get('room_rates_notes', {})
                    breakfast_labels = parsed.get('breakfast_labels', {})
                    for idx, rr in enumerate(q.room_rates):
                        if str(idx) in rr_notes:
                            rr.notes = rr_notes[str(idx)]
                        if str(idx) in breakfast_labels:
                            rr.breakfast_included = breakfast_labels[str(idx)]

                    # Apply translated notes to meeting_rooms
                    mr_notes = parsed.get('meeting_rooms_notes', {})
                    for idx, mr in enumerate(q.meeting_rooms):
                        if str(idx) in mr_notes:
                            mr.notes = mr_notes[str(idx)]

                    # Apply translated descriptions to fb_options
                    fb_descs = parsed.get('fb_descriptions', {})
                    for idx, fb in enumerate(q.fb_options):
                        if str(idx) in fb_descs:
                            fb.menu_description = fb_descs[str(idx)]

                    # Apply translated top-level fields
                    if parsed.get('cancellation_policy'):
                        q.cancellation_policy = parsed['cancellation_policy']
                    if parsed.get('payment_terms'):
                        q.payment_terms = parsed['payment_terms']

                    db.session.flush()

                    results.append({
                        'hotel': q.hotel_name, 'ok': True, 'mode': 'rebuild',
                        'summary': (q.raw_summary or '')[:120],
                        'translated': list(rr_notes.keys()) + list(mr_notes.keys()),
                    })
                except Exception as e:
                    results.append({'hotel': q.hotel_name, 'ok': False,
                                    'error': str(e)})

            _time.sleep(0.3)  # rate limiting

        db.session.commit()
        log_audit('partivia', 'PartiviaQuote', None, 'import',
                  summary='Reparse completo di tutti i preventivi')
        next_offset = batch_offset + batch_limit
        return jsonify(ok=True, results=results,
                       cost_eur=round(total_cost * 0.92, 4),
                       processed=len(quotes), total=total_quotes,
                       next_offset=next_offset if next_offset < total_quotes else None)

    # ── Diagnostic: dump room rates ──────────────────────────────────────

    @app.get('/api/partivia/debug-rates')
    def partivia_debug_rates():
        quotes = PartiviaQuote.query.filter_by(deleted=False).order_by(PartiviaQuote.city, PartiviaQuote.hotel_name).all()
        out = []
        for q in quotes:
            out.append({
                'id': q.id, 'hotel': q.hotel_name, 'city': q.city,
                'email_log_id': q.email_log_id,
                'dates_proposed': q.dates_proposed,
                'raw_summary': q.raw_summary,
                'room_rates': [
                    {'room_type': rr.room_type,
                     'rate_per_night': rr.rate_per_night,
                     'breakfast': rr.breakfast_included,
                     'notes': rr.notes}
                    for rr in q.room_rates
                ],
            })
        return jsonify(out)

    # ── Edit inline quote ─────────────────────────────────────────────────

    @app.route('/api/partivia/quote/<int:qid>', methods=['PUT', 'PATCH'])
    def partivia_update_quote(qid):
        q = PartiviaQuote.query.get_or_404(qid)
        data = request.get_json()
        QUOTE_STR_FIELDS = ('hotel_name', 'city', 'stars', 'contact_name',
                            'contact_email', 'dates_proposed', 'rooms_available',
                            'min_rooms_required', 'cancellation_policy',
                            'payment_terms', 'validity_date', 'commission',
                            'total_estimate', 'included_services', 'notes',
                            'raw_summary', 'quote_status', 'image_url',
                            'website_url', 'address', 'vat_included', 'hidden')
        changes = _diff(q, data, QUOTE_STR_FIELDS)
        for field in QUOTE_STR_FIELDS:
            if field in data:
                val = data[field]
                if field == 'stars' and val is not None:
                    val = int(val) if str(val).strip() else None
                setattr(q, field, val)
        q.updated_at = datetime.utcnow()
        db.session.commit()
        if changes:
            log_audit('partivia', 'PartiviaQuote', q.id, 'update',
                      changes=changes,
                      summary=f'Modificato preventivo {q.hotel_name}')
        return jsonify(ok=True)

    @app.delete('/api/partivia/quote/<int:qid>')
    def partivia_delete_quote(qid):
        q = PartiviaQuote.query.get_or_404(qid)
        q.deleted = True
        q.deleted_at = datetime.utcnow()
        db.session.commit()
        log_audit('partivia', 'PartiviaQuote', q.id, 'delete',
                  summary=f'Eliminato preventivo {q.hotel_name}')
        return jsonify(ok=True)

    @app.post('/api/partivia/delete-hotel')
    def partivia_delete_hotel():
        data = request.get_json()
        hotel_key = data.get('hotel_key', '').lower().strip()
        deleted = 0
        for q in PartiviaQuote.query.filter_by(deleted=False).all():
            if q.hotel_name.lower().strip() == hotel_key:
                q.deleted = True
                q.deleted_at = datetime.utcnow()
                deleted += 1
        db.session.commit()
        log_audit('partivia', 'PartiviaQuote', None, 'delete',
                  summary=f'Eliminati preventivi hotel {hotel_key}')
        return jsonify(ok=True, deleted=deleted)

    @app.post('/api/partivia/toggle-hotel')
    def partivia_toggle_hotel():
        """Bulk update hotel visibility. Body: {hotels: {hotel_key: bool(hidden), ...}}"""
        data = request.get_json()
        visibility = data.get('hotels', {})
        updated = 0
        for q in PartiviaQuote.query.filter_by(deleted=False).all():
            key = q.hotel_name.lower().strip()
            if key in visibility:
                q.hidden = visibility[key]
                updated += 1
        db.session.commit()
        hotel_keys = ', '.join(visibility.keys())
        log_audit('partivia', 'PartiviaQuote', None, 'update',
                  summary=f'Visibilità hotel {hotel_keys} cambiata')
        return jsonify(ok=True, updated=updated)

    # ── Edit inline sotto-tabelle ─────────────────────────────────────────

    @app.put('/api/partivia/room-rate/<int:rid>')
    def partivia_update_room_rate(rid):
        rr = PartiviaRoomRate.query.get_or_404(rid)
        data = request.get_json()
        for f in ('room_type', 'rate_per_night', 'breakfast_included', 'notes'):
            if f in data:
                setattr(rr, f, data[f])
        db.session.commit()
        log_audit('partivia', 'PartiviaRoomRate', rid, 'update',
                  summary='Aggiornata tariffa camera')
        return jsonify(ok=True)

    @app.put('/api/partivia/meeting-room/<int:mid>')
    def partivia_update_meeting_room(mid):
        mr = PartiviaMeetingRoom.query.get_or_404(mid)
        data = request.get_json()
        for f in ('name', 'capacity', 'rate', 'notes'):
            if f in data:
                setattr(mr, f, data[f])
        db.session.commit()
        log_audit('partivia', 'PartiviaMeetingRoom', mid, 'update',
                  summary='Aggiornata sala meeting')
        return jsonify(ok=True)

    @app.put('/api/partivia/fb-option/<int:fid>')
    def partivia_update_fb_option(fid):
        fb = PartiviaFBOption.query.get_or_404(fid)
        data = request.get_json()
        for f in ('meal_type', 'price_per_person', 'menu_description'):
            if f in data:
                setattr(fb, f, data[f])
        db.session.commit()
        log_audit('partivia', 'PartiviaFBOption', fid, 'update',
                  summary='Aggiornata opzione F&B')
        return jsonify(ok=True)

    # ── Bulk set option deadline ───────────────────────────────────────────

    @app.post('/api/partivia/bulk-deadline')
    def partivia_bulk_deadline():
        data = request.get_json()
        deadline = data.get('deadline', '')
        quotes = PartiviaQuote.query.filter_by(deleted=False).all()
        for q in quotes:
            q.validity_date = deadline
            q.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify(ok=True, count=len(quotes))

    # ── Export Budget Excel with formulas ────────────────────────────────

    @app.get('/api/partivia/budget-export')
    def partivia_budget_export():
        import io
        import re
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

        wb = Workbook()
        ws = wb.active
        ws.title = 'Budget'

        # Styles
        hdr_font = Font(bold=True, color='FFFFFF', size=11)
        hdr_fill = PatternFill(start_color='A44227', end_color='A44227', fill_type='solid')
        param_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
        param_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        eur_fmt = '#,##0'
        pct_fmt = '0%'

        # ── Parameters section (rows 1-7) ──
        # Load saved params from budget overrides
        ov_row = BudgetOverride.query.first()
        ov_data = ov_row.data if ov_row else {}
        params = ov_data.get('_params', {})

        param_labels = [
            ('Rooms', int(params.get('rooms', 50))),
            ('Nights', int(params.get('nights', 2))),
            ('Participants', int(params.get('pax', 50))),
            ('Lunches', int(params.get('lunches', 2))),
            ('Half-day Meetings', int(params.get('meetings', 2))),
            ('VAT Rooms & F&B', 0.10),
            ('VAT Meeting Rooms', 0.21),
        ]
        for i, (label, val) in enumerate(param_labels, 1):
            ws.cell(row=i, column=1, value=label).font = param_font
            ws.cell(row=i, column=1).fill = param_fill
            c = ws.cell(row=i, column=2, value=val)
            c.fill = param_fill
            c.font = Font(bold=True, size=12, color='A44227')
            if isinstance(val, float):
                c.number_format = pct_fmt

        # Parameter cell references
        P_ROOMS = '$B$1'
        P_NIGHTS = '$B$2'
        P_PAX = '$B$3'
        P_LUNCHES = '$B$4'
        P_MEETINGS = '$B$5'
        P_VAT_FB = '$B$6'
        P_VAT_MTG = '$B$7'

        # ── Header row (row 9) ──
        HDR_ROW = 9
        headers = ['Hotel', 'City', 'Room/night', 'Rooms Subtotal',
                   'Lunch/pax', 'Lunch Subtotal',
                   'Meeting Room', 'Meeting Subtotal',
                   'Dinner/pax', 'Dinner Subtotal',
                   'VAT incl.', 'NET', 'VAT €', 'TOTAL']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=HDR_ROW, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            c.border = thin_border

        # ── Hotel rows ──
        quotes = (PartiviaQuote.query.filter_by(deleted=False)
                  .order_by(PartiviaQuote.city, PartiviaQuote.hotel_name)
                  .all())

        # Group by hotel (best = most complete)
        hotels_seen = {}
        for q in quotes:
            key = q.hotel_name.lower().replace("'", "")
            if key not in hotels_seen:
                hotels_seen[key] = q
            else:
                existing = hotels_seen[key]
                if len(q.room_rates) + len(q.meeting_rooms) + len(q.fb_options) > \
                   len(existing.room_rates) + len(existing.meeting_rooms) + len(existing.fb_options):
                    hotels_seen[key] = q

        def parse_price(s):
            if not s:
                return None
            cleaned = re.sub(r'[€$£\s]', '', str(s))
            cleaned = re.split(r'[/\(]', cleaned)[0].strip()
            if ',' in cleaned and '.' not in cleaned:
                cleaned = cleaned.replace(',', '')
            elif ',' in cleaned and '.' in cleaned:
                if cleaned.rindex(',') > cleaned.rindex('.'):
                    cleaned = cleaned.replace('.', '').replace(',', '.')
                else:
                    cleaned = cleaned.replace(',', '')
            try:
                return float(cleaned)
            except ValueError:
                return None

        def find_rate(rates):
            # Single use / DUS room rate
            single_kws = ['single', 'singol', 'sgl', 'dus', 'individual',
                          'single use', 'single occupancy']
            for rr in rates:
                if any(k in (rr.room_type or '').lower() for k in single_kws):
                    p = parse_price(rr.rate_per_night)
                    if p:
                        return p
            # Fallback: lowest available
            lowest = None
            for rr in rates:
                p = parse_price(rr.rate_per_night)
                if p is not None and (lowest is None or p < lowest):
                    lowest = p
            return lowest

        def find_fb(fb_opts, meal):
            kws = {'lunch': ['lunch', 'pranzo', 'almuerzo'],
                   'dinner': ['dinner', 'cena', 'gala', 'cocktail dinner']}
            keywords = kws.get(meal, [meal])
            for fb in fb_opts:
                if any(k in (fb.meal_type or '').lower() for k in keywords):
                    p = parse_price(fb.price_per_person)
                    if p:
                        return p
            return None

        def find_meeting(mrs):
            for mr in mrs:
                rate_str = (mr.rate or '').lower()
                # Try half-day price first
                half_match = re.search(r'€?\s?([\d.,]+)\s*[/\(]?\s*half', rate_str, re.I) \
                    or re.search(r'half[- ]?day[:\s]*€?\s?([\d.,]+)', rate_str, re.I)
                if half_match:
                    p = parse_price(half_match.group(1))
                    if p:
                        return p
                p = parse_price(mr.rate)
                if p:
                    return p
            return None

        row = HDR_ROW + 1
        for key, q in sorted(hotels_seen.items(), key=lambda x: (x[1].city, x[1].hotel_name)):
            hotel_ov = ov_data.get(key, {})

            room_rate = hotel_ov.get('room_rate') or find_rate(q.room_rates)
            lunch_pp = hotel_ov.get('lunch_pp') or find_fb(q.fb_options, 'lunch')
            meeting_rate = hotel_ov.get('meeting_rate') or find_meeting(q.meeting_rooms)
            dinner_pp = hotel_ov.get('dinner_pp') or find_fb(q.fb_options, 'dinner')
            vat = q.vat_included or 'unknown'

            # Col A: Hotel
            ws.cell(row=row, column=1, value=q.hotel_name).font = Font(bold=True)
            # Col B: City
            ws.cell(row=row, column=2, value=q.city)
            # Col C: Room/night (editable value)
            ws.cell(row=row, column=3, value=room_rate).number_format = eur_fmt
            # Col D: Rooms Subtotal = C * Rooms * Nights
            ws.cell(row=row, column=4).value = f'=IF(C{row}="","",C{row}*{P_ROOMS}*{P_NIGHTS})'
            ws.cell(row=row, column=4).number_format = eur_fmt
            # Col E: Lunch/pax
            ws.cell(row=row, column=5, value=lunch_pp).number_format = eur_fmt
            # Col F: Lunch Subtotal = E * Pax * Lunches
            ws.cell(row=row, column=6).value = f'=IF(E{row}="","",E{row}*{P_PAX}*{P_LUNCHES})'
            ws.cell(row=row, column=6).number_format = eur_fmt
            # Col G: Meeting Room
            ws.cell(row=row, column=7, value=meeting_rate).number_format = eur_fmt
            # Col H: Meeting Subtotal = G * Meetings
            ws.cell(row=row, column=8).value = f'=IF(G{row}="","",G{row}*{P_MEETINGS})'
            ws.cell(row=row, column=8).number_format = eur_fmt
            # Col I: Dinner/pax
            ws.cell(row=row, column=9, value=dinner_pp).number_format = eur_fmt
            # Col J: Dinner Subtotal = I * Pax
            ws.cell(row=row, column=10).value = f'=IF(I{row}="","",I{row}*{P_PAX})'
            ws.cell(row=row, column=10).number_format = eur_fmt
            # Col K: VAT included (Y/N)
            vat_label = 'Y' if vat == 'yes' else 'N' if vat == 'no' else '?'
            ws.cell(row=row, column=11, value=vat_label).alignment = Alignment(horizontal='center')
            # Col L: NET
            # If Y: sum of subtotals/(1+vat_rate)
            # If N: sum of subtotals as-is
            net_formula = (
                f'=IF(K{row}="Y",'
                f'IF(D{row}<>"",D{row}/(1+{P_VAT_FB}),0)+IF(F{row}<>"",F{row}/(1+{P_VAT_FB}),0)'
                f'+IF(H{row}<>"",H{row}/(1+{P_VAT_MTG}),0)+IF(J{row}<>"",J{row}/(1+{P_VAT_FB}),0),'
                f'IF(D{row}<>"",D{row},0)+IF(F{row}<>"",F{row},0)+IF(H{row}<>"",H{row},0)+IF(J{row}<>"",J{row},0))'
            )
            ws.cell(row=row, column=12).value = net_formula
            ws.cell(row=row, column=12).number_format = eur_fmt
            # Col M: VAT €
            # If Y: gross - net
            # If N: each subtotal * vat_rate
            vat_formula = (
                f'=IF(K{row}="Y",'
                f'(IF(D{row}<>"",D{row},0)+IF(F{row}<>"",F{row},0)+IF(H{row}<>"",H{row},0)+IF(J{row}<>"",J{row},0))-L{row},'
                f'IF(K{row}="N",'
                f'IF(D{row}<>"",D{row}*{P_VAT_FB},0)+IF(F{row}<>"",F{row}*{P_VAT_FB},0)'
                f'+IF(H{row}<>"",H{row}*{P_VAT_MTG},0)+IF(J{row}<>"",J{row}*{P_VAT_FB},0),'
                f'""))'
            )
            ws.cell(row=row, column=13).value = vat_formula
            ws.cell(row=row, column=13).number_format = eur_fmt
            # Col N: TOTAL = NET + VAT
            ws.cell(row=row, column=14).value = f'=IF(M{row}="",L{row},L{row}+M{row})'
            ws.cell(row=row, column=14).number_format = eur_fmt
            ws.cell(row=row, column=14).font = Font(bold=True, size=12, color='A44227')

            # Apply borders
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = thin_border

            row += 1

        # Column widths
        col_widths = [28, 14, 12, 14, 12, 14, 14, 14, 12, 14, 10, 14, 14, 14]
        for i, w in enumerate(col_widths, 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = f'C{HDR_ROW + 1}'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name='Partivia_Budget.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ── Export Excel comparativo ──────────────────────────────────────────

    @app.get('/api/partivia/export')
    def partivia_export():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        quotes = (PartiviaQuote.query.filter_by(deleted=False)
                  .order_by(PartiviaQuote.city, PartiviaQuote.hotel_name)
                  .all())

        wb = Workbook()
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='2F5496')
        city_fill = PatternFill('solid', fgColor='D6E4F0')
        city_font = Font(bold=True, size=12)
        link_font = Font(bold=True, size=12, color='0563C1', underline='single')
        section_font = Font(bold=True, color='2F5496')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        wrap = Alignment(wrap_text=True, vertical='top')

        # ── Tab 1: Hotel Comparison (client-facing) ──
        ws = wb.active
        ws.title = 'Hotel Comparison'
        headers = [
            'City', 'Hotel', 'Stars', 'Available Dates',
            'Rooms Available',
            'Single/night', 'Double/night', 'Suite/night',
            'Meeting Room', 'Capacity', 'Meeting Cost',
            'Lunch/pax', 'Dinner/pax', 'Coffee Break/pax',
            'Option Deadline', 'Notes',
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        for row, q in enumerate(quotes, 2):
            rates = {r.room_type.lower(): r for r in q.room_rates}
            single = next((r for k, r in rates.items()
                           if 'singol' in k or 'single' in k), None)
            double = next((r for k, r in rates.items()
                           if 'doppi' in k or 'double' in k or 'twin' in k), None)
            suite = next((r for k, r in rates.items()
                          if 'suite' in k or 'junior' in k), None)
            main_mr = q.meeting_rooms[0] if q.meeting_rooms else None

            # Collect all meeting room notes for technical details
            mr_notes = '; '.join(
                f"{mr.name}: {mr.notes}" for mr in q.meeting_rooms
                if mr.notes
            ) if q.meeting_rooms else ''

            fb = {o.meal_type.lower(): o for o in q.fb_options}
            lunch = next((o for k, o in fb.items()
                          if 'pranzo' in k or 'lunch' in k), None)
            dinner = next((o for k, o in fb.items()
                           if 'cena' in k or 'dinner' in k or 'gala' in k), None)
            coffee = next((o for k, o in fb.items()
                           if 'coffee' in k or 'break' in k), None)

            # Notes: room + meeting only
            note_parts = []
            if mr_notes:
                note_parts.append(mr_notes)
            # Room notes
            for rr in q.room_rates:
                if rr.notes:
                    note_parts.append(f"{rr.room_type}: {rr.notes}")
            if q.notes:
                note_parts.append(q.notes)

            vals = [
                q.city, q.hotel_name, q.stars, q.dates_proposed or '',
                q.rooms_available or '',
                single.rate_per_night if single else '',
                double.rate_per_night if double else '',
                suite.rate_per_night if suite else '',
                main_mr.name if main_mr else '',
                main_mr.capacity if main_mr else '',
                main_mr.rate if main_mr else '',
                lunch.price_per_person if lunch else '',
                dinner.price_per_person if dinner else '',
                coffee.price_per_person if coffee else '',
                q.validity_date or '',
                '; '.join(note_parts) if note_parts else '',
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border
                cell.alignment = wrap

            # Hotel name as hyperlink if website_url exists
            hotel_cell = ws.cell(row=row, column=2)
            if q.website_url:
                hotel_cell.hyperlink = q.website_url
                hotel_cell.font = Font(color='0563C1', underline='single')

            # Highlight option deadline
            deadline_cell = ws.cell(row=row, column=15)
            if q.validity_date:
                deadline_cell.font = Font(bold=True, color='E65100')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions[get_column_letter(16)].width = 45
        ws.freeze_panes = 'C2'

        # ── Tab 2+: Detail per city ──
        quotes_by_city = {}
        for q in quotes:
            quotes_by_city.setdefault(q.city.upper(), []).append(q)

        for city in sorted(quotes_by_city.keys()):
            cqs = quotes_by_city[city]
            ws_c = wb.create_sheet(title=city[:31])
            ws_c.merge_cells('A1:F1')
            cell = ws_c.cell(row=1, column=1, value=f'Hotel Quotes — {city}')
            cell.font = Font(bold=True, size=14, color='2F5496')

            r = 3
            for q in sorted(cqs, key=lambda x: x.hotel_name):
                ws_c.merge_cells(f'A{r}:F{r}')
                hotel_label = f"{q.hotel_name} {'★' * (q.stars or 0)}"
                cell = ws_c.cell(row=r, column=1, value=hotel_label)
                if q.website_url:
                    cell.hyperlink = q.website_url
                    cell.font = link_font
                else:
                    cell.font = city_font
                cell.fill = city_fill
                r += 1
                for label, val in [
                    ('Available Dates', q.dates_proposed or '-'),
                    ('Rooms Available', str(q.rooms_available) if q.rooms_available else '-'),
                    ('Option Deadline', q.validity_date or '-'),
                ]:
                    ws_c.cell(row=r, column=1, value=label).font = Font(bold=True)
                    val_cell = ws_c.cell(row=r, column=2, value=val)
                    if label == 'Option Deadline' and q.validity_date:
                        val_cell.font = Font(bold=True, color='E65100')
                    r += 1

                # ROOM RATES
                if q.room_rates:
                    r += 1
                    ws_c.cell(row=r, column=1,
                              value='ROOM RATES').font = section_font
                    r += 1
                    for hdr_col, hdr_val in enumerate(
                            ['Type', 'Rate/Night', 'Breakfast', 'Notes'], 1):
                        c = ws_c.cell(row=r, column=hdr_col, value=hdr_val)
                        c.font = Font(bold=True, size=10)
                        c.fill = PatternFill('solid', fgColor='E8ECF4')
                    r += 1
                    for rate in q.room_rates:
                        ws_c.cell(row=r, column=1, value=rate.room_type)
                        ws_c.cell(row=r, column=2, value=rate.rate_per_night)
                        ws_c.cell(row=r, column=3,
                                  value=rate.breakfast_included or '')
                        ws_c.cell(row=r, column=4, value=rate.notes or '')
                        r += 1

                # MEETING ROOMS
                if q.meeting_rooms:
                    r += 1
                    ws_c.cell(row=r, column=1,
                              value='MEETING ROOMS').font = section_font
                    r += 1
                    for hdr_col, hdr_val in enumerate(
                            ['Room', 'Capacity', 'Cost', 'Technical Details'], 1):
                        c = ws_c.cell(row=r, column=hdr_col, value=hdr_val)
                        c.font = Font(bold=True, size=10)
                        c.fill = PatternFill('solid', fgColor='E8ECF4')
                    r += 1
                    for mr in q.meeting_rooms:
                        ws_c.cell(row=r, column=1, value=mr.name)
                        ws_c.cell(row=r, column=2, value=mr.capacity or '')
                        ws_c.cell(row=r, column=3, value=mr.rate or '')
                        ws_c.cell(row=r, column=4,
                                  value=mr.notes or '').alignment = wrap
                        r += 1

                # F&B RATES
                if q.fb_options:
                    r += 1
                    ws_c.cell(row=r, column=1,
                              value='F&B RATES').font = section_font
                    r += 1
                    for hdr_col, hdr_val in enumerate(
                            ['Type', 'Price/pax', 'Description'], 1):
                        c = ws_c.cell(row=r, column=hdr_col, value=hdr_val)
                        c.font = Font(bold=True, size=10)
                        c.fill = PatternFill('solid', fgColor='E8ECF4')
                    r += 1
                    for fb in q.fb_options:
                        ws_c.cell(row=r, column=1, value=fb.meal_type)
                        ws_c.cell(row=r, column=2,
                                  value=fb.price_per_person or '')
                        ws_c.cell(row=r, column=3,
                                  value=fb.menu_description or '')
                        r += 1

                # NOTES (room + meeting only)
                if q.notes:
                    r += 1
                    ws_c.cell(row=r, column=1, value='Notes').font = Font(bold=True)
                    ws_c.cell(row=r, column=2,
                              value=q.notes).alignment = wrap
                    r += 1
                r += 2

            ws_c.column_dimensions['A'].width = 22
            ws_c.column_dimensions['B'].width = 35
            ws_c.column_dimensions['C'].width = 20
            ws_c.column_dimensions['D'].width = 40

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        today = datetime.now().strftime('%Y-%m-%d')
        return send_file(buf, as_attachment=True,
                         download_name=f'partivia_hotel_comparison_{today}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ── Aggiungi manualmente ──────────────────────────────────────────────

    @app.post('/api/partivia/quote')
    def partivia_add_quote():
        data = request.get_json()
        q = PartiviaQuote(
            hotel_name=data.get('hotel_name', ''),
            city=data.get('city', ''),
            stars=data.get('stars'),
            contact_name=data.get('contact_name'),
            contact_email=data.get('contact_email'),
            dates_proposed=data.get('dates_proposed'),
            rooms_available=data.get('rooms_available'),
            total_estimate=data.get('total_estimate'),
            validity_date=data.get('validity_date'),
            image_url=data.get('image_url'),
            website_url=data.get('website_url'),
            address=data.get('address'),
            notes=data.get('notes'),
            source='manual',
        )
        db.session.add(q)
        db.session.commit()
        log_audit('partivia', 'PartiviaQuote', q.id, 'create',
                  summary=f'Creato preventivo {q.hotel_name}')
        return jsonify(ok=True, id=q.id)

    # ── Deadline monitor ─────────────────────────────────────────────

    @app.post('/api/deadline-monitor/run')
    def deadline_monitor_run():
        """Run the deadline monitor: fetch emails, parse, update deadlines."""
        import threading

        dry_run = request.args.get('dry_run', '').lower() in ('1', 'true')
        days = int(request.args.get('days', '3'))

        def _run():
            from deadline_monitor.run import run as monitor_run
            monitor_run(days=days, dry_run=dry_run)

        t = threading.Thread(target=_run, daemon=True)
        t.start()
        return jsonify(ok=True, message='Deadline monitor started in background')

    @app.get('/api/deadline-monitor/logs')
    def deadline_monitor_logs():
        """Get deadline-related email logs."""
        logs = EmailLog.query.filter_by(log_type='deadline') \
            .order_by(EmailLog.created_at.desc()).limit(50).all()
        return jsonify([{
            'id': l.id,
            'summary': l.summary,
            'created_at': l.created_at.isoformat() if l.created_at else None,
        } for l in logs])

    # ── TOUR: Liqui Moly ────────────────────────────────────────────────

    @app.route('/api/tour/client-link', methods=['POST'])
    def tour_generate_client_link():
        import secrets
        data = request.json or {}
        label = (data.get('label') or 'Client').strip()
        token = secrets.token_urlsafe(24)
        ct = TourClientToken(label=label, token=token)
        db.session.add(ct)
        db.session.commit()
        return jsonify({'token': ct.token, 'url': f'/tour/client/{ct.token}',
                        'label': ct.label, 'id': ct.id})

    @app.route('/api/tour/client-links')
    def tour_client_links():
        tokens = TourClientToken.query.order_by(TourClientToken.created_at.desc()).all()
        return jsonify([{'id': t.id, 'label': t.label, 'token': t.token,
                         'created_at': t.created_at.strftime('%d/%m/%Y %H:%M') if t.created_at else ''}
                        for t in tokens])

    @app.route('/api/tour/client-link/<int:tid>', methods=['DELETE'])
    def tour_delete_client_link(tid):
        t = TourClientToken.query.get_or_404(tid)
        db.session.delete(t)
        db.session.commit()
        return jsonify({'ok': True})

    @app.route('/tour/client/<token>')
    def tour_client_view(token):
        ct = TourClientToken.query.filter_by(token=token).first_or_404()
        return _tour_render(client_view=True)

    @app.route('/tour')
    def tour_index():
        return _tour_render(client_view=False)

    def _tour_render(client_view=False):
        import re as _re_tour
        guests = TourGuest.query.filter_by(deleted=False).order_by(TourGuest.cognome, TourGuest.nome).all()
        hotels = TourHotel.query.order_by(TourHotel.night_date, TourHotel.hotel_name).all()

        # Build occupancy data per hotel
        # Non-suffixed codes (GS, KING) = 1 room per person
        # Suffixed codes (DBL-T4, XL-2) = 1 room per unique raw code
        hotel_occupancy = {}  # hotel.id → {base_code: {room_count, people, guests}}
        suffix_re = _re_tour.compile(r'-[A-Z]*\d+$')
        for h in hotels:
            assignments = TourRoomAssignment.query.filter_by(hotel_id=h.id).all()
            occ = {}
            for a in assignments:
                base_code = suffix_re.sub('', a.room_code)
                has_suffix = (base_code != a.room_code)
                if base_code not in occ:
                    occ[base_code] = {'shared_codes': set(), 'single_count': 0,
                                      'people': 0, 'guests': []}
                occ[base_code]['people'] += 1
                occ[base_code]['guests'].append((a.guest, a.room_code))
                if has_suffix:
                    occ[base_code]['shared_codes'].add(a.room_code)
                else:
                    occ[base_code]['single_count'] += 1
            # Convert to room_count
            for code in occ:
                occ[code]['room_count'] = (occ[code]['single_count']
                                           + len(occ[code]['shared_codes']))
                del occ[code]['shared_codes']
                del occ[code]['single_count']
            hotel_occupancy[h.id] = occ

        # Build per-hotel summary: rooms_used, people
        hotel_summary = {}
        for h in hotels:
            occ = hotel_occupancy.get(h.id, {})
            total_rooms = 0
            total_people = 0
            for code, data in occ.items():
                total_rooms += data['room_count']
                total_people += data['people']
            hotel_summary[h.id] = {'rooms_used': total_rooms, 'people': total_people}

        # Build per-night summary
        from collections import OrderedDict
        night_data = OrderedDict()
        for h in hotels:
            nl = h.night_label
            if nl not in night_data:
                night_data[nl] = {'hotels': [], 'total_blocked': 0,
                                  'total_rooms_used': 0, 'total_people': 0}
            hs = hotel_summary[h.id]
            night_data[nl]['hotels'].append(h)
            night_data[nl]['total_blocked'] += h.rooms_blocked
            night_data[nl]['total_rooms_used'] += hs['rooms_used']
            night_data[nl]['total_people'] += hs['people']

        # Tour stages info
        stages = [
            {'date': '2 September', 'night': '2Sep', 'city': 'Brescia',
             'route': 'Arrival & Network Dinner',
             'dinner_venue': 'Centro Paolo VI',
             'highlights': ['Guest arrivals', 'Network Dinner at Centro Paolo VI']},
            {'date': '3 September', 'night': '3Sep', 'city': 'Ferrara',
             'route': 'Brescia → Ferrara',
             'dinner_venue': 'Circolo dei Negozianti',
             'highlights': ['South Garda Karting Circuit', 'Lunch at Fabbrica Pedavena, Verona',
                            'Coffee Break at Monselice Incredibilia', 'Dinner at Circolo dei Negozianti']},
            {'date': '4 September', 'night': '4Sep', 'city': 'Maranello',
             'route': 'Ferrara → Maranello',
             'dinner_venue': 'Hotel Arthur Pool',
             'highlights': ['Golf Club Le Fonti', 'Lunch at Passo della Futa',
                            'Ferrari Plant visit', 'Dinner at Hotel Arthur Pool + DJ set']},
            {'date': '5 September', 'night': '5Sep', 'city': 'Brescia',
             'route': 'Maranello → Brescia',
             'dinner_venue': 'Antica Fratta, Franciacorta',
             'highlights': ['Lunch at Cologne Metelli',
                            'Dinner at Antica Fratta Winery']},
        ]

        # Payment stats
        payment_stats = {
            'total': len([g for g in guests if g.payment != 'PAID-CANCELLED']),
            'cancelled': len([g for g in guests if g.payment == 'PAID-CANCELLED']),
            'paid': len([g for g in guests if g.payment == 'PAID']),
            'to_collect': len([g for g in guests if g.payment and 'COLLECT' in (g.payment or '')
                               and g.payment != 'PAID-CANCELLED']),
            'no_need': len([g for g in guests if g.payment == 'NO NEED']),
            'on_site': len([g for g in guests if g.payment == 'PAY ON SITE']),
            'dinner_2sep': len([g for g in guests if g.dinner and g.payment != 'PAID-CANCELLED']),
        }

        return render_template('tour.html', guests=guests, hotels=hotels,
                               hotel_occupancy=hotel_occupancy,
                               hotel_summary=hotel_summary,
                               night_data=night_data,
                               stages=stages,
                               payment_stats=payment_stats,
                               client_view=client_view)

    @app.route('/api/tour/stats')
    def tour_stats():
        import re as _re_stats
        hotels = TourHotel.query.order_by(TourHotel.night_date, TourHotel.hotel_name).all()
        result = []
        for h in hotels:
            assignments = TourRoomAssignment.query.filter_by(hotel_id=h.id).all()
            # Count unique rooms: strip suffix (-T4, -D2, -1 etc.) then
            # count distinct raw codes per base code as rooms
            room_set = set()
            people = 0
            for a in assignments:
                room_set.add(a.room_code)
                people += 1
            result.append({
                'id': h.id,
                'column_key': h.column_key,
                'night_label': h.night_label,
                'hotel_name': h.hotel_name,
                'city': h.city,
                'rooms_blocked': h.rooms_blocked,
                'rooms_used': len(room_set),
                'people': people,
                'categories': [{
                    'code': c.code,
                    'category_name': c.category_name,
                    'rooms_available': c.rooms_available,
                } for c in h.categories],
            })
        return jsonify(result)

    @app.route('/api/tour/guest/<int:gid>', methods=['PUT'])
    def tour_update_guest(gid):
        g = TourGuest.query.get_or_404(gid)
        data = request.json
        str_fields = (
            'cognome', 'nome', 'email', 'telefono', 'nazionalita', 'titolo',
            'arrivo_mezzo', 'arrivo_data', 'room_with', 'car_number', 'car_with',
            'vip', 'client_room_note', 'payment', 'cloth_size', 'diet',
            'notes', 'email_requests',
        )
        bool_fields = ('dinner', 'sept2')
        changes = _diff(g, data, str_fields, bool_fields)
        for f in str_fields:
            if f in data:
                setattr(g, f, data[f] or None)
        for f in bool_fields:
            if f in data:
                setattr(g, f, _parse_bool(data[f]))
        db.session.commit()
        if changes:
            log_audit('tour', 'TourGuest', g.id, 'update',
                      changes=changes,
                      summary=f'Modificato {g.nome_completo}')
        return jsonify({'ok': True})

    @app.route('/api/tour/guest/<int:gid>/room', methods=['PUT'])
    def tour_assign_room(gid):
        """Assign or update a room for a guest at a specific hotel."""
        g = TourGuest.query.get_or_404(gid)
        data = request.json
        hotel_id = data['hotel_id']
        room_code = (data.get('room_code') or '').strip()

        existing = TourRoomAssignment.query.filter_by(
            guest_id=gid, hotel_id=hotel_id).first()

        if room_code:
            if existing:
                existing.room_code = room_code
            else:
                db.session.add(TourRoomAssignment(
                    guest_id=gid, hotel_id=hotel_id, room_code=room_code))
        else:
            if existing:
                db.session.delete(existing)

        db.session.commit()
        log_audit('tour', 'TourRoomAssignment', g.id, 'assign',
                  summary=f'Camera assegnata a {g.nome_completo}')
        return jsonify({'ok': True})

    # ── TOUR EXPORT XLSX ─────────────────────────────────────────────────

    @app.get('/api/tour/export')
    def tour_export_xlsx():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        guests = TourGuest.query.filter_by(deleted=False).order_by(TourGuest.cognome, TourGuest.nome).all()
        hotels = TourHotel.query.order_by(TourHotel.night_date, TourHotel.hotel_name).all()

        # Hotel lookup by column_key
        hotel_by_key = {h.column_key: h for h in hotels}

        wb = Workbook()
        ws = wb.active
        ws.title = 'Completed'

        # ── Colour palette (from original XLSX) ──
        FILL_NONE    = PatternFill()                                      # no fill
        FILL_GREY    = PatternFill('solid', fgColor='EDEDED')             # car group
        FILL_GREEN   = PatternFill('solid', fgColor='E2EFDA')             # 1Sep
        FILL_YELLOW  = PatternFill('solid', fgColor='FFF2CC')             # 2Sep
        FILL_BLUE    = PatternFill('solid', fgColor='DDEBF7')             # 3Sep
        FILL_PEACH   = PatternFill('solid', fgColor='FCE4D6')             # 4Sep
        FILL_PURPLE  = PatternFill('solid', fgColor='E4DFEC')             # 5Sep

        FONT_NORMAL  = Font(size=11)
        FONT_BOLD    = Font(bold=True, size=11)
        FONT_CANCEL  = Font(color='808080', size=11)
        FONT_HDR_BLK = Font(bold=True, size=12, color='000000')           # base headers

        # ── Exact column order matching original XLSX ──
        # (header_label, data_fn, header_fill, data_fill)
        COL_HOTEL_ORDER = [
            '1Sep_Paolovi', '2Sep_Paolovi',
            '3Sep_CasaEste', '3Sep_Carlton', '3Sep_Europa',
            '4Sep_Arthur', '4Sep_Arthurino', '4Sep_AcetaiaBoni', '4Sep_Village',
            '5Sep_Paolovi',
        ]

        HOTEL_FILL = {
            '1Sep_Paolovi':     FILL_GREEN,
            '2Sep_Paolovi':     FILL_YELLOW,
            '3Sep_CasaEste':    FILL_BLUE,
            '3Sep_Carlton':     FILL_BLUE,
            '3Sep_Europa':      FILL_BLUE,
            '4Sep_Arthur':      FILL_PEACH,
            '4Sep_Arthurino':   FILL_PEACH,
            '4Sep_AcetaiaBoni': FILL_PEACH,
            '4Sep_Village':     FILL_PEACH,
            '5Sep_Paolovi':     FILL_PURPLE,
        }

        # Build column spec: (header, getter_fn, header_fill, data_fill)
        def _hotel_getter(key):
            def fn(g, ra_map):
                hid = hotel_by_key[key].id if key in hotel_by_key else None
                return ra_map.get(hid, '') if hid else ''
            return fn

        columns = [
            # A-I: base (no fill)
            ('Surname',       lambda g, m: g.cognome,        FILL_NONE, FILL_NONE),
            ('Name',          lambda g, m: g.nome,           FILL_NONE, FILL_NONE),
            ('Email',         lambda g, m: g.email,          FILL_NONE, FILL_NONE),
            ('Arrival',       lambda g, m: g.arrivo_mezzo,   FILL_NONE, FILL_NONE),
            ('Nationality',   lambda g, m: g.nazionalita,    FILL_NONE, FILL_NONE),
            ('Sept2',         lambda g, m: 'yes' if g.sept2 else 'no', FILL_NONE, FILL_NONE),
            ('Telephone',     lambda g, m: g.telefono,       FILL_NONE, FILL_NONE),
            ('Title',         lambda g, m: g.titolo,         FILL_NONE, FILL_NONE),
            ('Arrival_date',  lambda g, m: g.arrivo_data,    FILL_NONE, FILL_NONE),
            # J: room_with (no fill)
            ('room_with',     lambda g, m: g.room_with,      FILL_NONE, FILL_NONE),
            # K-N: car group (grey fill)
            ('car',           lambda g, m: g.car_number,     FILL_GREY,  FILL_NONE),
            ('car_with',      lambda g, m: g.car_with,       FILL_GREY,  FILL_NONE),
            ('VIP',           lambda g, m: g.vip,            FILL_GREY,  FILL_NONE),
            ('client_room_note', lambda g, m: g.client_room_note, FILL_GREY, FILL_NONE),
            # O: 1Sep hotel (green)
            ('1Sep_Paolovi',  _hotel_getter('1Sep_Paolovi'), FILL_GREEN, FILL_GREEN),
            # P: dinner (no fill)
            ('dinner',        lambda g, m: 'X' if g.dinner else '', FILL_NONE, FILL_NONE),
            # Q: 2Sep hotel (yellow)
            ('2Sep_Paolovi',  _hotel_getter('2Sep_Paolovi'), FILL_YELLOW, FILL_YELLOW),
            # R: payment (no fill)
            ('payment',       lambda g, m: g.payment,        FILL_NONE, FILL_NONE),
            # S-U: 3Sep hotels (blue)
            ('3Sep_CasaEste', _hotel_getter('3Sep_CasaEste'), FILL_BLUE, FILL_BLUE),
            ('3Sep_Carlton',  _hotel_getter('3Sep_Carlton'),  FILL_BLUE, FILL_BLUE),
            ('3Sep_Europa',   _hotel_getter('3Sep_Europa'),   FILL_BLUE, FILL_BLUE),
            # V-Y: 4Sep hotels (peach)
            ('4Sep_Arthur',       _hotel_getter('4Sep_Arthur'),       FILL_PEACH, FILL_PEACH),
            ('4Sep_Arthurino',    _hotel_getter('4Sep_Arthurino'),    FILL_PEACH, FILL_PEACH),
            ('4Sep_AcetaiaBoni',  _hotel_getter('4Sep_AcetaiaBoni'),  FILL_PEACH, FILL_PEACH),
            ('4Sep_Village',      _hotel_getter('4Sep_Village'),      FILL_PEACH, FILL_PEACH),
            # Z: 5Sep hotel (purple)
            ('5Sep_Paolovi',  _hotel_getter('5Sep_Paolovi'), FILL_PURPLE, FILL_PURPLE),
            # AA-AD: post cols (no fill)
            ('cloth',         lambda g, m: g.cloth_size,      FILL_NONE, FILL_NONE),
            ('diet',          lambda g, m: g.diet,            FILL_NONE, FILL_NONE),
            ('notes',         lambda g, m: g.notes,           FILL_NONE, FILL_NONE),
            ('Email_Requests', lambda g, m: g.email_requests, FILL_NONE, FILL_NONE),
        ]

        # ── Write header row ──
        for c, (header, _, hdr_fill, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=c, value=header)
            cell.font = FONT_HDR_BLK
            if hdr_fill.fgColor and hdr_fill.fgColor.rgb and hdr_fill.fgColor.rgb != '00000000':
                cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # ── Write data rows ──
        for r, g in enumerate(guests, 2):
            ra_map = {ra.hotel_id: ra.room_code for ra in g.room_assignments}
            is_cancelled = g.payment == 'PAID-CANCELLED'

            for c, (_, getter, _, data_fill) in enumerate(columns, 1):
                val = getter(g, ra_map)
                cell = ws.cell(row=r, column=c, value=val if val is not None else '')

                # Apply column background fill
                has_fill = data_fill.fgColor and data_fill.fgColor.rgb and data_fill.fgColor.rgb != '00000000'
                if is_cancelled:
                    cell.font = FONT_CANCEL
                    if has_fill:
                        cell.fill = data_fill  # keep column colour, grey text
                else:
                    cell.font = FONT_NORMAL
                    if has_fill:
                        cell.fill = data_fill

        # ── Auto-width ──
        for col_cells in ws.columns:
            max_len = 0
            for cell in col_cells:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except Exception:
                    pass
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 40)

        # ── Sheet 2: Summary ─────────────────────────────────────────────
        ws2 = wb.create_sheet('Summary')
        ws2.cell(row=1, column=1, value='Liqui Moly Tour - Export Summary').font = Font(bold=True, size=13)

        row = 3
        total = len(guests)
        paid = sum(1 for g in guests if g.payment == 'PAID')
        collect = sum(1 for g in guests if g.payment and 'COLLECT' in (g.payment or ''))
        cancelled = sum(1 for g in guests if g.payment == 'PAID-CANCELLED')
        dinner = sum(1 for g in guests if g.dinner)

        for label, val in [
            ('Total participants', total),
            ('Paid', paid),
            ('To collect', collect),
            ('Paid-Cancelled', cancelled),
            ('Dinner (2 Sep)', dinner),
        ]:
            ws2.cell(row=row, column=1, value=label).font = FONT_BOLD
            ws2.cell(row=row, column=2, value=val)
            row += 1

        row += 1
        for ci, h in enumerate(['Night', 'Hotel', 'Blocked', 'Used', 'Left', 'People'], 1):
            ws2.cell(row=row, column=ci, value=h).font = FONT_BOLD
        row += 1

        for h in hotels:
            assignments = TourRoomAssignment.query.filter_by(hotel_id=h.id).all()
            room_set = set()
            for a in assignments:
                room_set.add(a.room_code)
            ws2.cell(row=row, column=1, value=h.night_label)
            ws2.cell(row=row, column=2, value=h.hotel_name)
            ws2.cell(row=row, column=3, value=h.rooms_blocked)
            ws2.cell(row=row, column=4, value=len(room_set))
            ws2.cell(row=row, column=5, value=h.rooms_blocked - len(room_set))
            ws2.cell(row=row, column=6, value=len(assignments))
            row += 1

        ws2.column_dimensions['A'].width = 18
        ws2.column_dimensions['B'].width = 30

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        today = datetime.now().strftime('%Y-%m-%d')
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'liquimoly_rooming_{today}.xlsx')

    # ── TOUR: dinner export ──────────────────────────────────────────────

    @app.get('/api/tour/export/dinners')
    def tour_export_dinners():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import re as _re_din

        guests = TourGuest.query.filter_by(deleted=False).order_by(TourGuest.cognome, TourGuest.nome).all()
        hotels = TourHotel.query.order_by(TourHotel.night_date, TourHotel.hotel_name).all()

        # Build assignment lookup: guest_id → {night_label → (hotel_name, city, room_code)}
        base_re = _re_din.compile(r'-[A-Z]*\d+$')
        cat_names_all = {}  # (hotel_id, code) → clean name
        for h in hotels:
            for c in h.categories:
                clean = c.category_name
                for sfx in (' - Single Use', ' for Single Use',
                            ' (twin beds on request)', ' with Single Bed',
                            ' / Double for Single Use'):
                    clean = clean.replace(sfx, '')
                cat_names_all[(h.id, c.code)] = clean

        guest_rooms = {}  # guest_id → {night_label → "Hotel, City - Category"}
        for h in hotels:
            assignments = TourRoomAssignment.query.filter_by(hotel_id=h.id).all()
            for a in assignments:
                base_code = base_re.sub('', a.room_code)
                cat_label = cat_names_all.get((h.id, base_code), base_code)
                sleeping = f'{h.hotel_name}, {h.city} - {cat_label}'
                guest_rooms.setdefault(a.guest_id, {})[h.night_label] = sleeping

        # Dinner config per night
        DINNERS = [
            {
                'night': '2Sep',
                'sheet_name': '2 Sept - Paolo VI',
                'title': 'Dinner 2 September 2026',
                'venue': 'Centro Paolo VI, Brescia',
                'use_dinner_flag': True,  # use guest.dinner column
            },
            {
                'night': '3Sep',
                'sheet_name': '3 Sept - Circolo Negozianti',
                'title': 'Dinner 3 September 2026',
                'venue': 'Circolo dei Negozianti, Ferrara',
                'use_dinner_flag': False,  # use room assignments
            },
            {
                'night': '4Sep',
                'sheet_name': '4 Sept - Arthur Pool',
                'title': 'Dinner 4 September 2026',
                'venue': 'Hotel Arthur - pool area, Maranello',
                'use_dinner_flag': False,
            },
            {
                'night': '5Sep',
                'sheet_name': '5 Sept - Antica Fratta',
                'title': 'Dinner 5 September 2026',
                'venue': 'Antica Fratta, Monticelli Brusati',
                'use_dinner_flag': False,
            },
        ]

        # Styles
        font_title = Font(bold=True, size=14, name='Calibri')
        font_sub = Font(size=10, name='Calibri')
        font_info = Font(bold=True, size=11, name='Calibri')
        font_hdr = Font(bold=True, size=11, name='Calibri')
        font_data = Font(size=10, name='Calibri')
        fill_hdr = PatternFill('solid', fgColor='D9E1F2')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        wb = Workbook()
        first_sheet = True

        for dinner in DINNERS:
            night = dinner['night']

            # Select guests for this dinner
            if dinner['use_dinner_flag']:
                dinner_guests = [g for g in guests
                                 if g.dinner and g.payment != 'PAID-CANCELLED']
            else:
                dinner_guests = [g for g in guests
                                 if night in guest_rooms.get(g.id, {})
                                 and g.payment != 'PAID-CANCELLED']

            # Also include guests with dinner flag who have no room but are dining
            # (for 2Sep this is already handled; for other nights, room = attending)

            diet_count = sum(1 for g in dinner_guests
                             if g.diet and g.diet.strip().lower()
                             not in ('no', 'none', 'no.', '', 'non'))

            if first_sheet:
                ws = wb.active
                ws.title = dinner['sheet_name']
                first_sheet = False
            else:
                ws = wb.create_sheet(dinner['sheet_name'])

            # Header rows
            ws.cell(row=1, column=1, value=dinner['title']).font = font_title
            ws.cell(row=2, column=1, value=dinner['venue']).font = font_sub
            ws.cell(row=3, column=1,
                    value='LIQUI MOLY NEXUS AUTO TOUR 2026 - group CARZILLA').font = font_sub
            ws.cell(row=4, column=1,
                    value=f'{len(dinner_guests)} guests - {diet_count} with dietary requirements').font = font_info

            # Column headers
            headers = ['#', 'Surname', 'Name', 'Nationality',
                       'Diet / allergies', 'Sleeping that night']
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=6, column=c, value=h)
                cell.font = font_hdr
                cell.fill = fill_hdr
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            # Data rows
            for i, g in enumerate(dinner_guests, 1):
                sleeping = guest_rooms.get(g.id, {}).get(night, 'no room this night')

                diet_val = g.diet
                if diet_val and diet_val.strip().lower() in ('no', 'none', 'no.', 'non'):
                    diet_val = None

                row_data = [
                    i,
                    g.cognome,
                    g.nome,
                    g.nazionalita,
                    diet_val,
                    sleeping,
                ]
                for c, v in enumerate(row_data, 1):
                    cell = ws.cell(row=6 + i, column=c, value=v if v is not None else '')
                    cell.font = font_data
                    cell.border = thin_border

            # Column widths
            widths = [5, 22, 18, 16, 30, 45]
            for c, w in enumerate(widths, 1):
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(c)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='dinners_LMNAT2026.xlsx')

    # ── TOUR: per-hotel export ───────────────────────────────────────────

    @app.get('/api/tour/export/hotel/<int:hotel_id>')
    def tour_export_hotel(hotel_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import re as _re_hex

        hotel = TourHotel.query.get_or_404(hotel_id)
        assignments = TourRoomAssignment.query.filter_by(hotel_id=hotel_id).all()

        # Group by room_code, resolve guests
        # Shared rooms: same base code with suffix → group together
        guest_map = {}  # guest_id → TourGuest
        for a in assignments:
            guest_map[a.guest_id] = a.guest

        # Build room list.
        # Codes WITH suffix (e.g. PD-D2) = shared room → group by raw_code.
        # Codes WITHOUT suffix (e.g. PD) = individual rooms → one room per person.
        # Clean category names for hotel rooming lists
        def _clean_cat(name):
            for suffix in (' - Single Use', ' for Single Use',
                           ' (twin beds on request)', ' with Single Bed',
                           ' / Double for Single Use'):
                name = name.replace(suffix, '')
            return name
        cat_names = {c.code: _clean_cat(c.category_name) for c in hotel.categories}

        cat_order = {c.code: c.sort_order for c in hotel.categories}
        base_re = _re_hex.compile(r'-[A-Z]*\d+$')

        rooms = []
        shared_groups = {}  # suffixed raw_code → [assignment, ...]
        for a in assignments:
            has_suffix = bool(base_re.search(a.room_code))
            if has_suffix:
                shared_groups.setdefault(a.room_code, []).append(a)
            else:
                # Individual room
                base = a.room_code
                rooms.append({
                    'raw_code': a.room_code,
                    'base_code': base,
                    'category_name': cat_names.get(base, base),
                    'sort_key': (cat_order.get(base, 999), a.guest.cognome),
                    'guests': [a.guest],
                    'shared': False,
                })

        # Add shared rooms
        for raw_code, assigns in shared_groups.items():
            base = base_re.sub('', raw_code)
            rooms.append({
                'raw_code': raw_code,
                'base_code': base,
                'category_name': cat_names.get(base, base),
                'sort_key': (cat_order.get(base, 999), assigns[0].guest.cognome),
                'guests': [a.guest for a in assigns],
                'shared': len(assigns) > 1,
            })

        rooms.sort(key=lambda r: r['sort_key'])

        # Determine bed type for shared rooms
        def _bed_type(guest, room_with):
            """'double bed' for couples, 'twin beds' otherwise."""
            if not room_with:
                return 'twin beds'
            rw = room_with.lower()
            if 'wife' in rw or 'husband' in rw or 'partner' in rw:
                return 'double bed'
            return 'twin beds'

        # ── Build workbook ──
        wb = Workbook()
        ws = wb.active
        ws.title = 'Rooming list'

        # Styles matching the template
        font_title = Font(bold=True, size=14, name='Calibri')
        font_sub = Font(size=10, name='Calibri')
        font_info = Font(bold=True, size=11, name='Calibri')
        font_hdr = Font(bold=True, size=11, name='Calibri')
        font_data = Font(size=10, name='Calibri')
        font_data_bold = Font(bold=True, size=10, name='Calibri')

        fill_hdr = PatternFill('solid', fgColor='D9E1F2')
        fill_shared = PatternFill('solid', fgColor='FFF2CC')
        fill_none = PatternFill()

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # Night label → readable date
        night_dates = {
            '1Sep': '1 September 2026', '2Sep': '2 September 2026',
            '3Sep': '3 September 2026', '4Sep': '4 September 2026',
            '5Sep': '5 September 2026',
        }
        checkin = night_dates.get(hotel.night_label, hotel.night_label)
        # Checkout = next day
        checkout_map = {
            '1Sep': '2 September 2026', '2Sep': '3 September 2026',
            '3Sep': '4 September 2026', '4Sep': '5 September 2026',
            '5Sep': '6 September 2026',
        }
        checkout = checkout_map.get(hotel.night_label, '')

        total_rooms = len(rooms)
        total_people = sum(len(r['guests']) for r in rooms)

        # Header rows
        ws.cell(row=1, column=1, value=f'{hotel.hotel_name} - {hotel.city}').font = font_title
        ws.cell(row=2, column=1, value='LIQUI MOLY NEXUS AUTO TOUR 2026 - group CARZILLA').font = font_sub
        ws.cell(row=3, column=1,
                value=f'Check-in {checkin}   /   Check-out {checkout}').font = font_info
        ws.cell(row=4, column=1,
                value=f'{total_rooms} rooms - {total_people} guests').font = font_sub

        # Column headers (row 6)
        headers = ['#', 'Room type', 'Room', 'Surname', 'Name', 'Nationality',
                   'Diet / notes', 'Beds']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=6, column=c, value=h)
            cell.font = font_hdr
            cell.fill = fill_hdr
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        row = 7
        room_num = 0
        for room in rooms:
            room_num += 1
            is_shared = room['shared']
            fill = fill_shared if is_shared else fill_none
            guests_in_room = room['guests']

            for gi, guest in enumerate(guests_in_room):
                is_first = (gi == 0)
                bed_text = ''
                if is_shared and is_first:
                    bed_text = _bed_type(guest, guest.room_with)

                data = [
                    room_num if is_first else None,           # #
                    room['category_name'] if is_first else None,  # Room type
                    None,                                      # Room (hotel fills in)
                    (guest.cognome or '').upper(),             # Surname
                    (guest.nome or '').upper(),                # Name
                    guest.nazionalita,                         # Nationality
                    guest.diet if guest.diet and guest.diet.lower() not in ('no', 'none', 'no.') else None,
                    bed_text or None,                          # Beds
                ]
                for c, v in enumerate(data, 1):
                    cell = ws.cell(row=row, column=c, value=v if v is not None else '')
                    cell.font = font_data
                    cell.border = thin_border
                    if is_shared:
                        cell.fill = fill

                row += 1

        # Column widths
        widths = [5, 20, 8, 22, 18, 16, 30, 14]
        for c, w in enumerate(widths, 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(c)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = hotel.hotel_name.replace(' ', '_').replace("'", '')
        filename = f'rooming_{hotel.night_label}_{safe_name}.xlsx'
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)

    # ── TOUR: serve documents from DB (admin) ──────────────────────────

    @app.route('/api/tour/doc/<int:guest_id>/<doc_type>')
    def tour_serve_doc(guest_id, doc_type):
        """Serve a document from DB inline."""
        doc = TourGuestDocument.query.filter_by(
            guest_id=guest_id, doc_type=doc_type).first()
        if not doc or not doc.data:
            return 'File not found', 404
        return send_file(BytesIO(doc.data), mimetype=doc.mime_type)

    # ── TOUR: hotel token management ────────────────────────────────────

    @app.route('/api/tour/hotel/<int:hotel_id>/generate-link', methods=['POST'])
    def tour_generate_hotel_link(hotel_id):
        """Generate (or return existing) access token for a hotel."""
        import secrets
        hotel = TourHotel.query.get_or_404(hotel_id)
        existing = TourHotelToken.query.filter_by(hotel_id=hotel_id).first()
        if existing:
            return jsonify({'token': existing.token,
                            'url': f'/tour/docs/{existing.token}'})
        token = secrets.token_urlsafe(24)
        ht = TourHotelToken(hotel_id=hotel_id, token=token)
        db.session.add(ht)
        db.session.commit()
        return jsonify({'token': ht.token, 'url': f'/tour/docs/{ht.token}'})

    @app.route('/api/tour/hotel-links')
    def tour_hotel_links():
        """List all hotel tokens and their access logs."""
        hotels = TourHotel.query.order_by(TourHotel.night_date, TourHotel.hotel_name).all()
        result = []
        for h in hotels:
            tok = TourHotelToken.query.filter_by(hotel_id=h.id).first()
            logs = []
            if tok:
                for log in (TourHotelAccessLog.query.filter_by(token_id=tok.id)
                            .order_by(TourHotelAccessLog.accessed_at.desc())
                            .limit(20).all()):
                    logs.append({
                        'action': log.action,
                        'detail': log.detail,
                        'ip': log.ip_address,
                        'at': log.accessed_at.strftime('%d/%m/%Y %H:%M') if log.accessed_at else '',
                    })
            result.append({
                'hotel_id': h.id,
                'hotel_name': h.hotel_name,
                'night_label': h.night_label,
                'city': h.city,
                'token': tok.token if tok else None,
                'logs': logs,
            })
        return jsonify(result)

    # ── TOUR: public hotel document portal (token-based) ─────────────

    def _hotel_guest_list(hotel):
        """Build sorted guest list with room info for a hotel."""
        import re as _re_gl
        base_re = _re_gl.compile(r'-[A-Z]*\d+$')
        cat_names = {}
        for c in hotel.categories:
            clean = c.category_name
            for sfx in (' - Single Use', ' for Single Use',
                        ' (twin beds on request)', ' with Single Bed',
                        ' / Double for Single Use'):
                clean = clean.replace(sfx, '')
            cat_names[c.code] = clean

        assignments = TourRoomAssignment.query.filter_by(hotel_id=hotel.id).all()
        guest_list = []
        for a in assignments:
            g = a.guest
            if g.payment == 'PAID-CANCELLED':
                continue
            base_code = base_re.sub('', a.room_code)
            guest_list.append({
                'guest': g,
                'room_code': a.room_code,
                'room_type': cat_names.get(base_code, base_code),
            })
        guest_list.sort(key=lambda x: (x['guest'].cognome, x['guest'].nome))
        return guest_list

    def _log_access(token_obj, action, detail=None):
        db.session.add(TourHotelAccessLog(
            token_id=token_obj.id,
            ip_address=request.remote_addr,
            user_agent=str(request.user_agent)[:500],
            action=action,
            detail=detail,
        ))
        db.session.commit()

    @app.route('/tour/docs/<token>')
    def tour_public_docs(token):
        tok = TourHotelToken.query.filter_by(token=token).first_or_404()
        hotel = tok.hotel
        guest_list = _hotel_guest_list(hotel)
        _log_access(tok, 'view')
        return render_template('tour_hotel_docs.html',
                               hotel=hotel, guest_list=guest_list, token=token)

    @app.route('/tour/docs/<token>/download/<int:guest_id>/<doc_type>')
    def tour_public_doc_download(token, guest_id, doc_type):
        """Download a single document from DB. doc_type = 'passport' or 'driving'."""
        tok = TourHotelToken.query.filter_by(token=token).first_or_404()
        TourRoomAssignment.query.filter_by(
            hotel_id=tok.hotel_id, guest_id=guest_id).first_or_404()

        doc = TourGuestDocument.query.filter_by(
            guest_id=guest_id, doc_type=doc_type).first()
        if not doc or not doc.data:
            return 'File not found', 404

        g = doc.guest
        _log_access(tok, f'download_{doc_type}', f'{g.cognome} {g.nome}')
        ext = os.path.splitext(doc.filename or '')[1]
        fname = f'{g.cognome}_{g.nome}_{doc_type}{ext}'
        return send_file(BytesIO(doc.data), mimetype=doc.mime_type,
                         as_attachment=True, download_name=fname)

    @app.route('/tour/docs/<token>/view/<int:guest_id>/<doc_type>')
    def tour_public_doc_view(token, guest_id, doc_type):
        """Serve a document inline (for image preview)."""
        tok = TourHotelToken.query.filter_by(token=token).first_or_404()
        TourRoomAssignment.query.filter_by(
            hotel_id=tok.hotel_id, guest_id=guest_id).first_or_404()

        doc = TourGuestDocument.query.filter_by(
            guest_id=guest_id, doc_type=doc_type).first()
        if not doc or not doc.data:
            return 'File not found', 404

        return send_file(BytesIO(doc.data), mimetype=doc.mime_type)

    @app.route('/tour/docs/<token>/download-all')
    def tour_public_docs_download_all(token):
        """Download all documents for this hotel as ZIP from DB."""
        import zipfile as _zf
        tok = TourHotelToken.query.filter_by(token=token).first_or_404()
        hotel = tok.hotel
        guest_list = _hotel_guest_list(hotel)

        buf = BytesIO()
        with _zf.ZipFile(buf, 'w', _zf.ZIP_DEFLATED) as zout:
            for item in guest_list:
                g = item['guest']
                folder = f'{g.cognome} {g.nome}'.strip()
                docs = TourGuestDocument.query.filter_by(guest_id=g.id).all()
                for doc in docs:
                    if not doc.data:
                        continue
                    ext = os.path.splitext(doc.filename or '')[1]
                    label = 'passport' if doc.doc_type == 'passport' else 'driving_licence'
                    zout.writestr(f'{folder}/{label}{ext}', doc.data)

        buf.seek(0)
        _log_access(tok, 'download_all')
        safe_name = hotel.hotel_name.replace(' ', '_').replace("'", '')
        return send_file(buf, mimetype='application/zip', as_attachment=True,
                         download_name=f'docs_{hotel.night_label}_{safe_name}.zip')

    # ── AUDIT LOG API ──────────────────────────────────────────────────────

    @app.get('/api/audit/<section>')
    def get_audit_log(section):
        if section not in ('rooming', 'partivia', 'tour', 'system'):
            return jsonify(ok=False, error='Sezione non valida'), 400
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        q = AuditLog.query.filter_by(section=section).order_by(AuditLog.timestamp.desc())

        entity_type = request.args.get('entity_type')
        if entity_type:
            q = q.filter_by(entity_type=entity_type)
        action = request.args.get('action')
        if action:
            q = q.filter_by(action=action)
        entity_id = request.args.get('entity_id', type=int)
        if entity_id:
            q = q.filter_by(entity_id=entity_id)

        total = q.count()
        entries = q.offset((page - 1) * per_page).limit(per_page).all()
        return jsonify(
            ok=True, total=total, page=page, per_page=per_page,
            entries=[{
                'id': e.id,
                'timestamp': e.timestamp.isoformat() if e.timestamp else None,
                'user_email': e.user_email,
                'entity_type': e.entity_type,
                'entity_id': e.entity_id,
                'action': e.action,
                'changes': e.changes,
                'summary': e.summary,
            } for e in entries]
        )

    @app.get('/api/audit/entity/<entity_type>/<int:entity_id>')
    def get_entity_audit(entity_type, entity_id):
        entries = AuditLog.query.filter_by(
            entity_type=entity_type, entity_id=entity_id
        ).order_by(AuditLog.timestamp.desc()).limit(200).all()
        return jsonify(
            ok=True,
            entries=[{
                'id': e.id,
                'timestamp': e.timestamp.isoformat() if e.timestamp else None,
                'user_email': e.user_email,
                'section': e.section,
                'action': e.action,
                'changes': e.changes,
                'summary': e.summary,
            } for e in entries]
        )

    @app.get('/api/audit/<section>/export')
    def export_audit_log(section):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        q = AuditLog.query.filter_by(section=section).order_by(AuditLog.timestamp.desc())
        entries = q.limit(5000).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Audit Log'
        headers = ['Data/Ora', 'Utente', 'Azione', 'Entità', 'ID', 'Riepilogo', 'Dettagli']
        hfont = Font(bold=True, size=11)
        hfill = PatternFill(fgColor='E0E0E0', fill_type='solid')
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hfont
            c.fill = hfill

        for ri, e in enumerate(entries, 2):
            ws.cell(row=ri, column=1, value=e.timestamp.strftime('%d/%m/%Y %H:%M') if e.timestamp else '')
            ws.cell(row=ri, column=2, value=e.user_email or '')
            ws.cell(row=ri, column=3, value=e.action or '')
            ws.cell(row=ri, column=4, value=e.entity_type or '')
            ws.cell(row=ri, column=5, value=e.entity_id)
            ws.cell(row=ri, column=6, value=e.summary or '')
            ws.cell(row=ri, column=7, value=json.dumps(e.changes, ensure_ascii=False) if e.changes else '')

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'audit_{section}.xlsx')

    # ── ROLLBACK SINGOLA AZIONE ───────────────────────────────────────────

    @app.post('/api/audit/<int:audit_id>/rollback')
    @superuser_required
    def rollback_audit(audit_id):
        entry = AuditLog.query.get_or_404(audit_id)
        changes = entry.changes or {}
        etype = entry.entity_type
        eid = entry.entity_id
        action = entry.action

        # Map entity_type to model
        MODEL_MAP = {
            'Guest': Guest,
            'PartiviaQuote': PartiviaQuote,
            'TourGuest': TourGuest,
        }
        Model = MODEL_MAP.get(etype)
        if not Model:
            return jsonify(ok=False, error=f'Rollback non supportato per {etype}'), 400

        if action == 'update':
            # Revert each field to its old value
            if not eid:
                return jsonify(ok=False, error='ID entità mancante'), 400
            obj = Model.query.get(eid)
            if not obj:
                return jsonify(ok=False, error=f'{etype} #{eid} non trovato'), 404
            reverted = {}
            for field, vals in changes.items():
                if isinstance(vals, dict) and 'old' in vals:
                    old_val = vals['old']
                    current_val = getattr(obj, field, None)
                    setattr(obj, field, old_val)
                    reverted[field] = {'old': current_val, 'new': old_val}
            if hasattr(obj, 'updated_at'):
                obj.updated_at = datetime.utcnow()
            db.session.commit()
            log_audit(entry.section, etype, eid, 'rollback',
                      changes=reverted,
                      summary=f'Rollback: annullata modifica #{audit_id} su {etype} #{eid}')
            return jsonify(ok=True, reverted=reverted)

        elif action == 'create':
            # Rollback di una creazione → soft-delete
            if not eid:
                return jsonify(ok=False, error='ID entità mancante'), 400
            obj = Model.query.get(eid)
            if not obj:
                return jsonify(ok=False, error=f'{etype} #{eid} non trovato'), 404
            obj.deleted = True
            obj.deleted_at = datetime.utcnow()
            db.session.commit()
            log_audit(entry.section, etype, eid, 'rollback',
                      summary=f'Rollback: eliminato {etype} #{eid} (annullata creazione #{audit_id})')
            return jsonify(ok=True, action='soft-deleted')

        elif action == 'delete':
            # Rollback di una cancellazione → restore
            if not eid:
                return jsonify(ok=False, error='ID entità mancante'), 400
            obj = Model.query.get(eid)
            if not obj:
                return jsonify(ok=False, error=f'{etype} #{eid} non trovato'), 404
            if not obj.deleted:
                return jsonify(ok=False, error='Il record non è eliminato'), 400
            obj.deleted = False
            obj.deleted_at = None
            # If we have a snapshot in changes, restore the fields
            for field, val in changes.items():
                if hasattr(obj, field) and not isinstance(val, dict):
                    setattr(obj, field, val)
            db.session.commit()
            log_audit(entry.section, etype, eid, 'rollback',
                      summary=f'Rollback: ripristinato {etype} #{eid} (annullata eliminazione #{audit_id})')
            return jsonify(ok=True, action='restored')

        elif action == 'import':
            return jsonify(ok=False, error='Rollback di import non supportato — usa il cestino per i singoli record'), 400

        else:
            return jsonify(ok=False, error=f'Rollback non supportato per azione "{action}"'), 400

    # ── RESTORE & DELETED LIST ─────────────────────────────────────────────

    @app.get('/api/guests/deleted')
    @superuser_required
    def list_deleted_guests():
        deleted = Guest.query.filter_by(deleted=True).order_by(Guest.deleted_at.desc()).all()
        return jsonify(ok=True, guests=[{
            'id': g.id,
            'cognome': g.cognome,
            'nome': g.nome,
            'email': g.email,
            'deleted_at': g.deleted_at.isoformat() if g.deleted_at else None,
        } for g in deleted])

    @app.post('/api/guest/<int:gid>/restore')
    @superuser_required
    def restore_guest(gid):
        g = Guest.query.get_or_404(gid)
        if not g.deleted:
            return jsonify(ok=False, error='Non è eliminato'), 400
        g.deleted = False
        g.deleted_at = None
        db.session.commit()
        log_audit('rooming', 'Guest', g.id, 'restore',
                  summary=f'Ripristinato {g.nome_completo}')
        return jsonify(ok=True)

    @app.get('/api/partivia/quotes/deleted')
    @superuser_required
    def list_deleted_quotes():
        deleted = PartiviaQuote.query.filter_by(deleted=True).order_by(
            PartiviaQuote.deleted_at.desc()).all()
        return jsonify(ok=True, quotes=[{
            'id': q.id,
            'hotel_name': q.hotel_name,
            'city': q.city,
            'deleted_at': q.deleted_at.isoformat() if q.deleted_at else None,
        } for q in deleted])

    @app.post('/api/partivia/quote/<int:qid>/restore')
    @superuser_required
    def restore_partivia_quote(qid):
        q = PartiviaQuote.query.get_or_404(qid)
        if not q.deleted:
            return jsonify(ok=False, error='Non è eliminato'), 400
        q.deleted = False
        q.deleted_at = None
        db.session.commit()
        log_audit('partivia', 'PartiviaQuote', q.id, 'restore',
                  summary=f'Ripristinato preventivo {q.hotel_name}')
        return jsonify(ok=True)

    @app.get('/api/tour/guests/deleted')
    @superuser_required
    def list_deleted_tour_guests():
        deleted = TourGuest.query.filter_by(deleted=True).order_by(
            TourGuest.deleted_at.desc()).all()
        return jsonify(ok=True, guests=[{
            'id': g.id,
            'cognome': g.cognome,
            'nome': g.nome,
            'email': g.email,
            'deleted_at': g.deleted_at.isoformat() if g.deleted_at else None,
        } for g in deleted])

    @app.post('/api/tour/guest/<int:gid>/restore')
    @superuser_required
    def restore_tour_guest(gid):
        g = TourGuest.query.get_or_404(gid)
        if not g.deleted:
            return jsonify(ok=False, error='Non è eliminato'), 400
        g.deleted = False
        g.deleted_at = None
        db.session.commit()
        log_audit('tour', 'TourGuest', g.id, 'restore',
                  summary=f'Ripristinato {g.nome_completo}')
        return jsonify(ok=True)

    return app


if __name__ == '__main__':
    app = create_app()
    app.run(debug=os.environ.get('FLASK_DEBUG', '1') == '1', port=5005)
