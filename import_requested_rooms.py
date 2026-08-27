"""
Parse hotel rooming XLSX files and populate rooms_requested on TourRoomCategory.

Each XLSX has columns: N, Tipo, Descrizione, Camera, Arrivo, Partenza, Cognome, Nome, ...
We count unique rooms per room type code (Tipo column) and update the DB.

File naming convention:
  rooming_3Sep_Hotel_Carlton_3.xlsx  → night_label=3Sep, hotel contains "Carlton"
  rooming_4Sep_Hotel_Arthur.xlsx     → night_label=4Sep, hotel contains "Arthur"
  rooming 2 sett paolo vi.xlsx      → night_label=2Sep, hotel contains "Paolo"

Usage:
  python import_requested_rooms.py [folder_path]

Default folder: C:\\Users\\ACER\\Downloads\\ROOMING 25 ago
"""
import os
import sys
import re
import openpyxl
from collections import Counter

# Map file patterns to (night_label, hotel_name_fragment)
FILE_PATTERNS = [
    (r'2 sett paolo vi', '2Sep', 'Paolo'),
    (r'5 ago paolo vi', '1Sep', 'Paolo'),      # 5 agosto = night 1Sep (arrival night)
    (r'3Sep_Hotel_Carlton', '3Sep', 'Carlton'),
    (r'3Sep_Hotel_Casa_dEste', '3Sep', "Casa d'Este"),
    (r'3Sep_Hotel_Europa', '3Sep', 'Europa'),
    (r'4Sep_Hotel_Arthur\.', '4Sep', 'Arthur'),
    (r'4Sep_Hotel_Arthurino', '4Sep', 'Arthurino'),
    (r'4Sep_Hotel_Maranello', '4Sep', 'Maranello'),
]


def parse_rooming_file(filepath):
    """Parse an XLSX rooming file and return room type counts."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Strategy 1: Find header row with known keywords
    header_row = None
    tipo_col = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=False), 1):
        vals = [str(c.value or '').strip().lower() for c in row]
        # Look for "tipo", "room type", "tipologia", "cat", "category"
        for i, v in enumerate(vals):
            if v in ('tipo', 'room type', 'tipologia', 'cat', 'category', 'cat.'):
                header_row = row_idx
                tipo_col = i
                break
            # Also check for "n" + "tipo" pattern
            if v == 'n' and i + 1 < len(vals) and vals[i + 1] in ('tipo', 'room type', 'tipologia'):
                header_row = row_idx
                tipo_col = i + 1
                break
        if header_row:
            break

    if not header_row:
        # Strategy 2: Scan all rows looking for known room codes
        known_codes = {'SGL', 'GSGL', 'DBL', 'KING', 'DLX', 'BAL', 'MON', 'APP', 'XL',
                       'TWN', 'TWIN', 'STD', 'SUP', 'JUN', 'SUITE', 'SINGLE', 'DOUBLE'}
        print(f'  No header found, scanning for room codes...')
        room_types = Counter()
        for row in ws.iter_rows(min_row=1, values_only=True):
            for cell in row:
                val = str(cell or '').strip().upper()
                if val in known_codes:
                    room_types[val] += 1
        if room_types:
            print(f'  Found codes by scanning: {dict(room_types)}')
            return dict(room_types)
        # Strategy 3: dump first rows for debugging
        print(f'  WARNING: No room type data found in {filepath}')
        print(f'  First 5 rows:')
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            print(f'    {list(row)}')
        return {}

    # Count rooms per type
    room_types = Counter()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row)
        tipo = str(vals[tipo_col] or '').strip().upper() if tipo_col < len(vals) else ''
        if tipo:
            room_types[tipo] += 1

    return dict(room_types)


def match_file_to_hotel(filename):
    """Match a filename to (night_label, hotel_name_fragment)."""
    for pattern, night, hotel_frag in FILE_PATTERNS:
        if re.search(pattern, filename, re.IGNORECASE):
            return night, hotel_frag
    return None, None


def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\ACER\Downloads\ROOMING 25 ago'

    # First, just parse and show what we found
    results = []
    for fname in os.listdir(folder):
        if not fname.endswith('.xlsx') or fname.startswith('~$'):
            continue
        night, hotel_frag = match_file_to_hotel(fname)
        if not night:
            print(f'SKIP: {fname} (no pattern match)')
            continue

        fpath = os.path.join(folder, fname)
        counts = parse_rooming_file(fpath)
        print(f'\n{fname}')
        print(f'  → night={night}, hotel=*{hotel_frag}*')
        print(f'  Room types: {dict(counts)}')
        total = sum(counts.values())
        print(f'  Total rooms: {total}')
        results.append((night, hotel_frag, counts))

    if not results:
        print('No files matched!')
        return

    # Now update DB
    print('\n--- Updating database ---')
    from app import create_app
    app = create_app()
    with app.app_context():
        from models import db, TourHotel, TourRoomCategory

        for night, hotel_frag, counts in results:
            # Find matching hotel
            hotels = TourHotel.query.filter_by(night_label=night).all()
            hotel = None
            for h in hotels:
                if hotel_frag.lower() in h.hotel_name.lower():
                    hotel = h
                    break

            if not hotel:
                print(f'  WARNING: No hotel found for night={night} fragment="{hotel_frag}"')
                print(f'    Available: {[h.hotel_name for h in hotels]}')
                continue

            print(f'\n  {hotel.hotel_name} ({hotel.night_label}):')
            categories = TourRoomCategory.query.filter_by(hotel_id=hotel.id).all()
            cat_map = {c.code.upper(): c for c in categories}

            for room_code, count in counts.items():
                cat = cat_map.get(room_code.upper())
                if cat:
                    old = cat.rooms_requested
                    cat.rooms_requested = count
                    print(f'    {room_code}: requested={count} (was {old})')
                else:
                    print(f'    {room_code}: no matching category (codes: {list(cat_map.keys())})')

        db.session.commit()
        print('\nDone!')


if __name__ == '__main__':
    main()
