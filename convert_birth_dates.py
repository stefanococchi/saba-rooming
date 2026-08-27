"""
Convert birth dates XLSX to JSON for /api/import/birth-dates endpoint.
Usage: python convert_birth_dates.py <xlsx_file> [output.json]
"""
import sys
import json
import openpyxl

def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\ACER\Downloads\Copia di equans data nascita mancanti ok.xlsx'
    out_path = sys.argv[2] if len(sys.argv) > 2 else 'birth_dates.json'

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    print(f'Headers: {headers}')
    print(f'Total rows: {ws.max_row - 1}')

    # Try to find name and birth date columns
    name_col = None
    date_col = None
    for i, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).lower().strip()
        if any(k in hl for k in ('cognome', 'nome', 'name', 'nominativo', 'partecipante')):
            if name_col is None:
                name_col = i
        if any(k in hl for k in ('nascita', 'birth', 'data di nascita', 'data nascita', 'dob')):
            date_col = i

    if name_col is None or date_col is None:
        # Try alternative: maybe columns are just index 0 and 1
        print(f'\nAuto-detect: name_col={name_col}, date_col={date_col}')
        print('First 5 rows:')
        for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
            print(f'  {row}')
        print('\nManual mode: edit the script to set name_col and date_col indices')
        if name_col is None:
            name_col = 0
        if date_col is None:
            date_col = 1
        print(f'Using: name_col={name_col} ({headers[name_col]}), date_col={date_col} ({headers[date_col]})')

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if name_col >= len(vals) or date_col >= len(vals):
            continue
        name = vals[name_col]
        date_val = vals[date_col]
        if not name or not date_val:
            continue
        name = str(name).strip()
        # Format date
        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%d/%m/%Y')
        else:
            date_str = str(date_val).strip()
        result[name] = date_str

    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f'\nConverted {len(result)} entries → {out_path}')
    print('Sample:')
    for k, v in list(result.items())[:5]:
        print(f'  {k}: {v}')

if __name__ == '__main__':
    main()
