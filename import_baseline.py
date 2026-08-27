"""
Import final request rooming lists as baseline snapshots into TourRoomBaseline.
Parses each hotel XLSX and stores every guest with their room type.

Usage:
  python import_baseline.py [folder_path]
"""
import os
import sys
import re
import openpyxl

# Normalize room names to DB codes
ROOM_NAME_TO_CODE = {
    # Paolo VI
    'DLX': 'KINGP', 'BAL': 'PAN', 'MON': 'MONO', 'APP': 'FLAT', 'GSGL': 'GS',
    # Carlton
    'TWIN ROOM': 'TWIN', 'DOUBLE ROOM': 'DBL', 'TO BE CONFIRMED': 'TBC',
    # Casa d'Este
    'ROYAL SUITE': 'ROYAL', 'JUNIOR SUITE': 'JS', 'DELUXE DOUBLE': 'DD',
    'SUPERIOR DOUBLE': 'SD', 'FAMILY ROOM': 'FAM', 'PREMIUM DOUBLE': 'PD',
    'CLASSIC DOUBLE': 'CD', 'CLASSIC SINGLE': 'CS',
    # Europa
    'SINGLE ROOM': 'SGL',
    # Arthur / generic
    'STANDARD ROOM': 'STD',
    # Arthurino
    'HOUSE ROOM': 'HOUSE',
    # Maranello
    'DELUXE': 'DLX',
}

FILE_PATTERNS = [
    (r'2 sett paolo vi', '2Sep', 'Paolo'),
    (r'5 ago paolo vi', '5Sep', 'Paolo'),
    (r'3Sep_Hotel_Carlton', '3Sep', 'Carlton'),
    (r'3Sep_Hotel_Casa_dEste', '3Sep', "Casa d'Este"),
    (r'3Sep_Hotel_Europa', '3Sep', 'Europa'),
    (r'4Sep_Hotel_Arthurino', '4Sep', 'Arthurino'),
    (r'4Sep_Hotel_Arthur', '4Sep', 'Arthur'),
    (r'4Sep_Hotel_Maranello', '4Sep', 'Maranello'),
]


def match_file(filename):
    for pattern, night, hotel_frag in FILE_PATTERNS:
        if re.search(pattern, filename, re.IGNORECASE):
            return night, hotel_frag
    return None, None


def parse_paolo_vi(filepath):
    """Parse Centro Paolo VI format (has Tipo, Camera, Cognome, Nome columns)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    # Find header
    header_row = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
        vals = [str(c.value or '').strip().lower() for c in row]
        if 'tipo' in vals and 'cognome' in vals:
            header_row = ri
            headers = vals
            break
    if not header_row:
        print(f'  WARNING: no header in {filepath}')
        return []

    tipo_col = headers.index('tipo')
    camera_col = headers.index('camera') if 'camera' in headers else None
    cog_col = headers.index('cognome')
    nom_col = headers.index('nome') if 'nome' in headers else cog_col + 1
    note_col = headers.index('note') if 'note' in headers else None

    guests = []
    current_room_code = None
    current_room_label = None
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row)
        tipo = str(vals[tipo_col] or '').strip().upper() if tipo_col < len(vals) else ''
        camera = str(vals[camera_col] or '').strip() if camera_col and camera_col < len(vals) else ''
        cognome = str(vals[cog_col] or '').strip().upper() if cog_col < len(vals) else ''
        nome = str(vals[nom_col] or '').strip() if nom_col < len(vals) else ''
        note = str(vals[note_col] or '').strip() if note_col and note_col < len(vals) else ''

        if tipo:
            current_room_code = tipo
            current_room_label = camera
        if cognome:
            guests.append({
                'room_code': current_room_code or '',
                'room_label': current_room_label or '',
                'cognome': cognome,
                'nome': nome,
                'notes': note,
            })
    return guests


def parse_hotel_export(filepath):
    """Parse hotel export format (has Room type, Room, Surname, Name columns)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    # Find header row
    header_row = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=False), 1):
        vals = [str(c.value or '').strip().lower() for c in row]
        if any(v in vals for v in ('room type', 'tipologia', 'surname', 'cognome')):
            header_row = ri
            headers = vals
            break
    if not header_row:
        # Try to find by known column patterns
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=False), 1):
            vals = [str(c.value or '').strip().lower() for c in row]
            if '#' in vals or 'n' in vals:
                header_row = ri
                headers = vals
                break

    if not header_row:
        print(f'  WARNING: no header in {filepath}')
        return []

    # Find columns
    def find_col(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    type_col = find_col('room type', 'tipo', 'tipologia', 'cat', 'category')
    room_col = find_col('room', 'camera', 'room no', 'n. camera')
    cog_col = find_col('surname', 'cognome', 'last name')
    nom_col = find_col('name', 'nome', 'first name')
    note_col = find_col('note', 'notes', 'beds')

    if cog_col is None:
        print(f'  WARNING: no surname column in {filepath}, headers={headers}')
        return []

    guests = []
    current_room_code = None
    current_room_label = None
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row)
        room_type = str(vals[type_col] or '').strip() if type_col is not None and type_col < len(vals) else ''
        room = str(vals[room_col] or '').strip() if room_col is not None and room_col < len(vals) else ''
        cognome = str(vals[cog_col] or '').strip().upper() if cog_col < len(vals) else ''
        nome = str(vals[nom_col] or '').strip() if nom_col is not None and nom_col < len(vals) else ''
        note = str(vals[note_col] or '').strip() if note_col is not None and note_col < len(vals) else ''

        if room_type:
            current_room_code = room_type
            current_room_label = room
        if cognome:
            guests.append({
                'room_code': current_room_code or '',
                'room_label': current_room_label or '',
                'cognome': cognome,
                'nome': nome,
                'notes': note,
            })
    return guests


def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\ACER\Downloads\ROOMING 25 ago'

    results = []
    for fname in os.listdir(folder):
        if not fname.endswith('.xlsx') or fname.startswith('~$'):
            continue
        night, hotel_frag = match_file(fname)
        if not night:
            continue

        fpath = os.path.join(folder, fname)
        # Paolo VI has different format
        if 'paolo vi' in fname.lower():
            guests = parse_paolo_vi(fpath)
        else:
            guests = parse_hotel_export(fpath)

        print(f'\n{fname}')
        print(f'  → night={night}, hotel=*{hotel_frag}*, {len(guests)} guests')
        results.append((night, hotel_frag, guests))

    if not results:
        print('No files matched!')
        return

    print('\n--- Updating database ---')
    from app import create_app
    app = create_app()
    with app.app_context():
        from models import db, TourHotel, TourRoomBaseline

        for night, hotel_frag, guests in results:
            hotels = TourHotel.query.filter_by(night_label=night).all()
            candidates = [h for h in hotels if hotel_frag.lower() in h.hotel_name.lower()]
            hotel = min(candidates, key=lambda h: len(h.hotel_name)) if candidates else None

            if not hotel:
                print(f'  WARNING: No hotel for night={night} fragment="{hotel_frag}"')
                continue

            # Clear existing baselines for this hotel
            TourRoomBaseline.query.filter_by(hotel_id=hotel.id).delete()

            # Insert new baselines (normalize room codes)
            for g in guests:
                raw_code = g['room_code']
                normalized = ROOM_NAME_TO_CODE.get(raw_code, raw_code)
                db.session.add(TourRoomBaseline(
                    hotel_id=hotel.id,
                    room_code=normalized,
                    room_label=g['room_label'],
                    cognome=g['cognome'],
                    nome=g['nome'],
                    notes=g['notes'],
                ))

            print(f'  {hotel.hotel_name} ({night}): {len(guests)} guests saved')

        db.session.commit()
        print('\nDone!')


if __name__ == '__main__':
    main()
